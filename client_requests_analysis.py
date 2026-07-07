from __future__ import annotations

import argparse
import asyncio
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bitrix24_api import Bitrix24API
from logging_setup import get_logger
from pipelines.calls import (
    _normalize_bitrix_text,
    extract_bitrix_gpt_summaries,
    fetch_deal_activities,
    fetch_open_next_step_activities,
    fetch_timeline_comments,
)
from pipelines.deals import deal_url_from_id
from pipelines.paths import REPORTS_DIR
from pipelines.transcription import _load_state_cache, load_cached_transcript

logger = get_logger(__name__)

SALES_DEPARTMENT_NAME = "Отдел продаж АРТ"
FIRED_DEPARTMENT_NAME = "Уволенные"
OP_CATEGORY_NAME = "ОП воронка"
DEFAULT_DOMAIN = "online-kassa.bitrix24.ru"


@dataclass(frozen=True)
class RequestCategory:
    name: str
    patterns: tuple[str, ...]
    recommendation: str


REQUEST_CATEGORIES: tuple[RequestCategory, ...] = (
    RequestCategory(
        "Маркировка / Честный знак",
        (
            r"\bмаркировк",
            r"честн(ый|ого)\s+знак",
            r"\bчз\b",
            r"разрешительн(ый|ого)\s+режим",
            r"табак|обув|молочн|бель[её]|пиво|кег",
        ),
        "Держать быстрый пресейл по маркировке: отрасль, товарная группа, схема работы, касса/ПО/сканер.",
    ),
    RequestCategory(
        "Касса / ККТ / онлайн-касса",
        (
            r"\bккт\b",
            r"\bкасс(а|ы|у|ой|овый|овая)",
            r"онлайн[\s-]?касс",
            r"эвотор|атол|меркурий|ньюджер|aqsi|акси|шtrih|штрих",
        ),
        "Разделить скрипт под подбор кассы: формат торговли, номенклатура, платежи, интеграции, бюджет.",
    ),
    RequestCategory(
        "Фискальный накопитель / ФН",
        (
            r"\bфн\b",
            r"фискальн(ый|ого|ые)\s+накопител",
            r"накопител",
        ),
        "Сразу уточнять срок ФН, систему налогообложения, маркировку/акциз и срочность замены.",
    ),
    RequestCategory(
        "Настройка / подключение / техподдержка",
        (
            r"настройк",
            r"подключ",
            r"тех\s?поддерж",
            r"\bтп\b",
            r"удаленн(ая|о|ую)\s+настрой",
            r"ошибк|не\s+работ|почин|помощ",
        ),
        "Отделять платную настройку от консультации: что сломано, модель, доступы, дедлайн, ответственный техник.",
    ),
    RequestCategory(
        "Оборудование: принтеры, сканеры, терминалы",
        (
            r"принтер\s+этикет",
            r"\bсканер",
            r"\bтсд\b",
            r"терминал",
            r"вес(ы|ов)",
            r"этикет",
            r"оборудован",
        ),
        "Вести подбор оборудования через чек-лист: нагрузка, интерфейсы, совместимость с кассой/1С/маркировкой.",
    ),
    RequestCategory(
        "Онлайн-оплата / эквайринг / СБП / QR",
        (
            r"эквайринг",
            r"онлайн[\s-]?оплат",
            r"\bсбп\b",
            r"\bqr\b|куар",
            r"оплат[аы]\s+по\s+ссылк",
            r"дистанционн(ая|ой)\s+оплат",
            r"кнопк[а-я]*\s+оплат",
        ),
        "Уточнять сценарий оплаты: сайт/мессенджер/ссылка, юрлицо, банк, фискализация, возвраты.",
    ),
    RequestCategory(
        "Интеграция с 1С / сайтом / CRM / складом",
        (
            r"\b1с\b",
            r"интеграц",
            r"сайт",
            r"интернет[\s-]?магазин",
            r"\bcrm\b|битрикс",
            r"мой\s?склад",
            r"автоматизац",
        ),
        "Подключать техпресейл раньше: текущая система, обмен номенклатурой, заказы, оплаты, ответственные API.",
    ),
    RequestCategory(
        "Регистрация / перерегистрация / ФНС / ОФД",
        (
            r"регистрац",
            r"перерегистрац",
            r"\bфнс\b|налогов",
            r"\bофд\b",
            r"снять\s+с\s+уч[её]та",
            r"поставить\s+на\s+уч[её]т",
        ),
        "Собирать обязательные данные заранее: ИНН, ОФД, ФН, модель ККТ, причина перерегистрации.",
    ),
    RequestCategory(
        "ЭДО / документы / УПД",
        (
            r"\bэдо\b",
            r"электронн(ый|ого)\s+документооборот",
            r"\bупд\b",
            r"диадок",
            r"закрыт(ые|ых|ым)?\s+документ",
        ),
        "Фиксировать документооборот отдельно от кассовой задачи: оператор ЭДО, участники, типы документов.",
    ),
    RequestCategory(
        "Облачная касса / аренда / сервис",
        (
            r"облачн(ая|ой)\s+касс",
            r"аренд[аы]",
            r"сервис",
            r"тариф",
            r"подписк",
            r"продлен",
        ),
        "Показывать экономику сервиса: срок, пакет, что входит, чем отличается аренда от покупки.",
    ),
    RequestCategory(
        "Консультация / подбор решения",
        (
            r"консультац",
            r"подбор",
            r"подобрать",
            r"что\s+нужно",
            r"какую\s+касс",
            r"интересует",
        ),
        "Переводить общий запрос в квалификацию: отрасль, способ продаж, товары, платежи, срок запуска.",
    ),
)


def _clean(value: Any) -> str:
    return _normalize_bitrix_text(str(value or ""))


def _is_context_noise(value: Any) -> bool:
    text = _clean(value)
    if not text:
        return True
    lowered = text.lower().replace("ё", "е")
    noise_markers = (
        "изменен ответственный за сделку",
        "изменён ответственный за сделку",
        "ответственный синхронизирован",
    )
    if any(marker in lowered for marker in noise_markers):
        return True
    if re.fullmatch(r"сумма документа:\s*[\d\s.,]*\s*(руб\.?|₽)?", lowered):
        return True
    return False


def _short(value: Any, limit: int = 900) -> str:
    text = _clean(value)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def _status_entity_id(category_id: str) -> str:
    return "DEAL_STAGE" if str(category_id) in {"0", ""} else f"DEAL_STAGE_{category_id}"


def _extract_text_from_activity(activity: dict[str, Any]) -> list[str]:
    parts: list[str] = []
    for key in ("SUBJECT", "DESCRIPTION", "RESULT_SUMMARY", "COMMENTS", "PROVIDER_TYPE_NAME"):
        text = _clean(activity.get(key))
        if text:
            parts.append(text)

    settings = activity.get("SETTINGS")
    if isinstance(settings, str) and settings.strip().startswith("{"):
        try:
            settings = json.loads(settings)
        except Exception:
            settings = None
    if isinstance(settings, dict):
        for key, value in settings.items():
            key_text = str(key or "").lower()
            if any(token in key_text for token in ("summary", "resume", "transcript", "резюме", "расшифр")):
                text = _clean(value)
                if text:
                    parts.append(text)
    return parts


def _dedupe_texts(values: list[str], max_items: int = 80) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = _clean(value)
        if not text or _is_context_noise(text):
            continue
        key = text.lower().replace("ё", "е")
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
        if len(out) >= max_items:
            break
    return out


def _find_evidence(texts: list[str], pattern: str) -> str:
    rx = re.compile(pattern, flags=re.IGNORECASE)
    for text in texts:
        match = rx.search(text)
        if not match:
            continue
        start = max(0, match.start() - 120)
        end = min(len(text), match.end() + 180)
        return _short(text[start:end], 360)
    return ""


def classify_request(texts: list[str]) -> dict[str, Any]:
    normalized = "\n".join(_clean(x).lower().replace("ё", "е") for x in texts if x)
    scored: list[tuple[int, str, str, str]] = []
    for category in REQUEST_CATEGORIES:
        score = 0
        evidence = ""
        for pattern in category.patterns:
            matches = re.findall(pattern, normalized, flags=re.IGNORECASE)
            if matches:
                score += len(matches)
                if not evidence:
                    evidence = _find_evidence(texts, pattern)
        if score:
            scored.append((score, category.name, evidence, category.recommendation))

    if not scored:
        return {
            "primary_request": "Недостаточно контекста",
            "secondary_requests": "",
            "request_score": 0,
            "request_evidence": "",
            "request_recommendation": "Нужна ручная проверка карточки или расшифровка звонка.",
        }

    scored.sort(key=lambda item: item[0], reverse=True)
    primary = scored[0]
    secondary = [name for _, name, _, _ in scored[1:4]]
    return {
        "primary_request": primary[1],
        "secondary_requests": "; ".join(secondary),
        "request_score": primary[0],
        "request_evidence": primary[2],
        "request_recommendation": primary[3],
    }


async def load_categories(api: Bitrix24API) -> tuple[str, dict[str, str]]:
    rows = (await api.call("crm.dealcategory.list", {})).get("result", []) or []
    category_id = "1"
    names: dict[str, str] = {}
    for row in rows:
        cid = str(row.get("ID") or "")
        name = str(row.get("NAME") or "")
        if cid:
            names[cid] = name
        if name.strip().lower() == OP_CATEGORY_NAME.lower():
            category_id = cid
    return category_id, names


async def load_stage_names(api: Bitrix24API, category_id: str) -> dict[str, str]:
    entity_id = _status_entity_id(category_id)
    rows = (await api.call("crm.status.list", {"filter": {"ENTITY_ID": entity_id}})).get("result", []) or []
    out: dict[str, str] = {}
    for row in rows:
        status_id = str(row.get("STATUS_ID") or "")
        name = str(row.get("NAME") or "")
        if status_id and name:
            out[f"C{category_id}:{status_id}"] = name
            out[status_id] = name
    return out


async def load_sales_users(api: Bitrix24API) -> tuple[list[str], dict[str, str]]:
    departments = (await api.call("department.get", {})).get("result", []) or []
    dep_by_id = {str(d.get("ID")): str(d.get("NAME") or "") for d in departments}
    sales_dep_ids = {
        str(d.get("ID"))
        for d in departments
        if str(d.get("NAME") or "").strip().lower() == SALES_DEPARTMENT_NAME.lower()
    }
    users = await api.get_all("user.get", {"FILTER": {"ACTIVE": True}})
    sales_ids: list[str] = []
    name_by_id: dict[str, str] = {}
    for user in users:
        uid = str(user.get("ID") or "")
        deps = [str(x) for x in (user.get("UF_DEPARTMENT") or [])]
        dep_names = [dep_by_id.get(x, "") for x in deps]
        if not uid or not (set(deps) & sales_dep_ids):
            continue
        if any(name.strip().lower() == FIRED_DEPARTMENT_NAME.lower() for name in dep_names):
            continue
        sales_ids.append(uid)
        name = f"{user.get('NAME', '')} {user.get('LAST_NAME', '')}".strip()
        name_by_id[uid] = name or uid
    return sales_ids, name_by_id


async def fetch_deals(
    api: Bitrix24API,
    *,
    category_id: str,
    sales_user_ids: list[str],
    date_from: str,
    date_to_exclusive: str,
) -> list[dict[str, Any]]:
    flt = {
        "CATEGORY_ID": category_id,
        ">=DATE_CREATE": date_from,
        "<DATE_CREATE": date_to_exclusive,
        "ASSIGNED_BY_ID": sales_user_ids,
    }
    return await api.get_all(
        "crm.deal.list",
        {
            "filter": flt,
            "select": [
                "ID",
                "TITLE",
                "ASSIGNED_BY_ID",
                "STAGE_ID",
                "DATE_CREATE",
                "DATE_MODIFY",
                "CLOSEDATE",
                "COMMENTS",
                "SOURCE_ID",
                "SOURCE_DESCRIPTION",
                "OPPORTUNITY",
                "CATEGORY_ID",
            ],
            "order": {"ID": "DESC"},
        },
    )


async def build_deal_row(
    api: Bitrix24API,
    deal: dict[str, Any],
    *,
    domain: str,
    manager_names: dict[str, str],
    stage_names: dict[str, str],
    state_cache: dict[str, Any],
    semaphore: asyncio.Semaphore,
) -> dict[str, Any]:
    async with semaphore:
        deal_id = str(deal.get("ID") or "")
        comments_task = asyncio.create_task(fetch_timeline_comments(api, deal_id))
        activities_task = asyncio.create_task(fetch_deal_activities(api, deal_id, limit=150))
        next_steps_task = asyncio.create_task(fetch_open_next_step_activities(api, deal_id))
        comments, activities, next_steps = await asyncio.gather(
            comments_task, activities_task, next_steps_task
        )

    texts: list[str] = [
        str(deal.get("TITLE") or ""),
        str(deal.get("COMMENTS") or ""),
        str(deal.get("SOURCE_DESCRIPTION") or ""),
    ]
    texts.extend(comments)

    call_count = 0
    cached_transcripts = 0
    chat_activity_count = 0
    task_activity_count = len(next_steps)

    for activity in activities:
        provider = str(activity.get("PROVIDER_ID") or "").upper()
        type_id = str(activity.get("TYPE_ID") or "")
        subject = str(activity.get("SUBJECT") or "")
        subject_low = subject.lower()
        if provider == "VOXIMPLANT_CALL" or type_id == "2":
            call_count += 1
            transcript, _path = load_cached_transcript(
                state_cache,
                str(activity.get("ORIGIN_ID") or ""),
                deal_id,
                activity.get("ID"),
            )
            if transcript:
                cached_transcripts += 1
                texts.append(transcript)
        if any(marker in subject_low for marker in ("чат", "telegram", "whatsapp", "openline")):
            chat_activity_count += 1
        texts.extend(_extract_text_from_activity(activity))

    for item in next_steps:
        texts.extend(_extract_text_from_activity(item))

    texts = _dedupe_texts(texts)
    summaries = extract_bitrix_gpt_summaries(comments + [str(deal.get("COMMENTS") or "")], activities)
    if summaries.get("bitrix_chat_summary"):
        texts.insert(0, str(summaries.get("bitrix_chat_summary")))
    if summaries.get("bitrix_call_summary"):
        texts.insert(0, str(summaries.get("bitrix_call_summary")))

    classified = classify_request(texts)
    context_chars = sum(len(x) for x in texts)
    context_sources = []
    if deal.get("COMMENTS"):
        context_sources.append("поле COMMENTS")
    if comments:
        context_sources.append("комментарии таймлайна")
    if activities:
        context_sources.append("активности")
    if cached_transcripts:
        context_sources.append("кэш расшифровок")
    if summaries.get("bitrix_summary_found"):
        context_sources.append("BitrixGPT")

    return {
        "deal_id": deal_id,
        "deal_url": deal_url_from_id(domain, deal_id),
        "date_create": deal.get("DATE_CREATE"),
        "manager_id": deal.get("ASSIGNED_BY_ID"),
        "manager_name": manager_names.get(str(deal.get("ASSIGNED_BY_ID") or ""), str(deal.get("ASSIGNED_BY_ID") or "")),
        "stage_id": deal.get("STAGE_ID"),
        "stage_name": stage_names.get(str(deal.get("STAGE_ID") or ""), str(deal.get("STAGE_ID") or "")),
        "title": _clean(deal.get("TITLE")),
        "source_id": deal.get("SOURCE_ID"),
        "source_description": _clean(deal.get("SOURCE_DESCRIPTION")),
        "primary_request": classified["primary_request"],
        "secondary_requests": classified["secondary_requests"],
        "request_score": classified["request_score"],
        "request_evidence": classified["request_evidence"],
        "request_recommendation": classified["request_recommendation"],
        "context_sources": "; ".join(context_sources),
        "context_chars": context_chars,
        "comments_count": len(comments),
        "activities_count": len(activities),
        "calls_count": call_count,
        "cached_transcripts_count": cached_transcripts,
        "chat_activity_count": chat_activity_count,
        "next_steps_count": task_activity_count,
        "bitrix_chat_summary": summaries.get("bitrix_chat_summary") or "",
        "bitrix_call_summary": summaries.get("bitrix_call_summary") or "",
        "context_excerpt": _short(" | ".join(texts[:8]), 1800),
        "needs_asr_review": bool(call_count and not cached_transcripts and context_chars < 260),
    }


def build_top_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total = len(rows)
    by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        category = str(row.get("primary_request") or "Недостаточно контекста")
        if category == "Недостаточно контекста":
            continue
        by_category[category].append(row)

    out: list[dict[str, Any]] = []
    for category, items in sorted(by_category.items(), key=lambda item: len(item[1]), reverse=True):
        examples = []
        for item in items[:8]:
            examples.append(f"{item.get('title')} ({item.get('deal_url')})")
        managers = Counter(str(item.get("manager_name") or "") for item in items)
        sources = Counter(str(item.get("source_id") or "") for item in items)
        recommendation = ""
        for cat in REQUEST_CATEGORIES:
            if cat.name == category:
                recommendation = cat.recommendation
                break
        if not recommendation:
            recommendation = "Разобрать вручную сделки без достаточного контекста и усилить обязательность комментария/резюме."
        out.append(
            {
                "rank": len(out) + 1,
                "request_category": category,
                "deals_count": len(items),
                "share_percent": round(len(items) * 100.0 / total, 2) if total else 0,
                "top_managers": "; ".join(f"{name}: {count}" for name, count in managers.most_common(5)),
                "top_sources": "; ".join(f"{name}: {count}" for name, count in sources.most_common(5)),
                "example_deals": "\n".join(examples),
                "recommendation": recommendation,
            }
        )
    return out[:10]


def build_manager_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], int] = Counter(
        (str(row.get("manager_name") or ""), str(row.get("primary_request") or "")) for row in rows
    )
    out: list[dict[str, Any]] = []
    for (manager, category), count in grouped.most_common():
        out.append({"manager_name": manager, "request_category": category, "deals_count": count})
    return out


def write_excel(path: Path, *, summary: list[dict[str, Any]], top: list[dict[str, Any]], deal_rows: list[dict[str, Any]], manager_rows: list[dict[str, Any]], weak_rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="Итоги", index=False)
        pd.DataFrame(top).to_excel(writer, sheet_name="Топ-10 запросов", index=False)
        pd.DataFrame(deal_rows).to_excel(writer, sheet_name="Сделки", index=False)
        pd.DataFrame(manager_rows).to_excel(writer, sheet_name="Менеджеры_запросы", index=False)
        pd.DataFrame(weak_rows).to_excel(writer, sheet_name="Нужна_ASR_проверка", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True, name="Arial")
        body_font = Font(name="Arial", size=10)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for idx, _col in enumerate(ws.columns, 1):
                header = str(ws.cell(row=1, column=idx).value or "")
                width = 18
                if any(token in header.lower() for token in ("url", "ссылка", "example", "context", "резюме", "evidence", "recommendation", "пример", "вывод")):
                    width = 42
                elif any(token in header.lower() for token in ("title", "название", "request", "запрос")):
                    width = 30
                ws.column_dimensions[get_column_letter(idx)].width = min(width, 60)


async def run(args: argparse.Namespace) -> tuple[Path, Path, list[dict[str, Any]]]:
    load_dotenv(args.env_path, override=True)
    api = Bitrix24API(readonly=True)
    try:
        category_id, category_names = await load_categories(api)
        sales_ids, manager_names = await load_sales_users(api)
        stage_names = await load_stage_names(api, category_id)
        deals = await fetch_deals(
            api,
            category_id=category_id,
            sales_user_ids=sales_ids,
            date_from=args.date_from,
            date_to_exclusive=args.date_to_exclusive,
        )
        logger.info(
            f"[REQUESTS] Сделок к анализу: {len(deals)}; "
            f"воронка={category_names.get(category_id, category_id)}; "
            f"менеджеров={len(sales_ids)}"
        )

        state_cache = _load_state_cache()
        semaphore = asyncio.Semaphore(max(1, int(args.concurrency or 6)))
        tasks = [
            build_deal_row(
                api,
                deal,
                domain=args.domain,
                manager_names=manager_names,
                stage_names=stage_names,
                state_cache=state_cache,
                semaphore=semaphore,
            )
            for deal in deals
        ]
        deal_rows = await asyncio.gather(*tasks)
    finally:
        await api.aclose()

    top_rows = build_top_rows(deal_rows)
    manager_rows = build_manager_rows(deal_rows)
    weak_rows = [
        row
        for row in deal_rows
        if row.get("needs_asr_review") or row.get("primary_request") == "Недостаточно контекста"
    ]

    summary = [
        {"metric": "Период", "value": f"{args.date_from} — {args.date_to_exclusive} (не включая верхнюю дату)"},
        {"metric": "Воронка", "value": OP_CATEGORY_NAME},
        {"metric": "Сделок проанализировано", "value": len(deal_rows)},
        {"metric": "Менеджеров отдела продаж", "value": len(set(row.get("manager_id") for row in deal_rows))},
        {"metric": "Сделок с кэшем расшифровок", "value": sum(1 for row in deal_rows if int(row.get("cached_transcripts_count") or 0) > 0)},
        {"metric": "Сделок с BitrixGPT/комментариями/активностями", "value": sum(1 for row in deal_rows if int(row.get("context_chars") or 0) >= 260)},
        {"metric": "Сделок для ASR/ручной проверки", "value": len(weak_rows)},
    ]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_out = REPORTS_DIR / f"client_requests_op_{args.date_from}_{args.date_to_exclusive}_{ts}.json"
    xlsx_out = REPORTS_DIR / f"client_requests_op_{args.date_from}_{args.date_to_exclusive}_{ts}.xlsx"
    payload = {
        "summary": summary,
        "top_10": top_rows,
        "deals": deal_rows,
        "manager_rows": manager_rows,
        "weak_context_rows": weak_rows,
    }
    json_out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(
        xlsx_out,
        summary=summary,
        top=top_rows,
        deal_rows=deal_rows,
        manager_rows=manager_rows,
        weak_rows=weak_rows,
    )
    return json_out, xlsx_out, top_rows


def parse_args() -> argparse.Namespace:
    today = date.today()
    default_to = (today + timedelta(days=1)).isoformat()
    parser = argparse.ArgumentParser(description="Анализ клиентских запросов по ОП воронке")
    parser.add_argument("--date-from", default="2026-05-01")
    parser.add_argument("--date-to-exclusive", default=default_to)
    parser.add_argument("--domain", default=DEFAULT_DOMAIN)
    parser.add_argument("--env-path", default=".env")
    parser.add_argument("--concurrency", type=int, default=6)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    json_out, xlsx_out, top_rows = asyncio.run(run(args))
    print(f"JSON: {json_out}")
    print(f"Excel: {xlsx_out}")
    print("TOP-10:")
    for row in top_rows:
        print(
            f"{row['rank']}. {row['request_category']}: "
            f"{row['deals_count']} ({row['share_percent']}%)"
        )


if __name__ == "__main__":
    main()
