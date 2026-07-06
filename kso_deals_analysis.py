from __future__ import annotations

import argparse
import asyncio
import json
import shutil
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bitrix24_api import Bitrix24API
from client_requests_analysis import _clean, _short
from pipelines.deals import deal_url_from_id
from pipelines.paths import REPORTS_DIR
from pipelines.stages import safe_int
from pipelines.transcription import _load_state_cache
from so_deals_analysis import (
    DEFAULT_DOMAIN,
    FINAL_NOT_WON_STAGES,
    as_float,
    classify_reason,
    enrich_context,
    has_meaningful_comment,
    lifetime_days,
    load_sales_users,
    parse_dt,
)

KSO_FIELD = "UF_CRM_1744698030"
KSO_VALUE_ID = "6867"
KSO_VALUE_NAME = "02.4 Кассы самообслуживания/Информационные киоски"


def is_won(deal: dict[str, Any]) -> bool:
    return str(deal.get("STAGE_SEMANTIC_ID") or "").upper() == "S" or str(
        deal.get("STAGE_ID") or ""
    ).endswith(":WON")


def is_failed(deal: dict[str, Any]) -> bool:
    return str(deal.get("STAGE_SEMANTIC_ID") or "").upper() == "F" or str(
        deal.get("STAGE_ID") or ""
    ) in FINAL_NOT_WON_STAGES


def stage_state(deal: dict[str, Any]) -> str:
    if is_won(deal):
        return "Выиграна"
    if is_failed(deal):
        return "Проиграна/провалена"
    return "В работе / не выиграна"


def days_since_update(value: Any) -> float | None:
    dt = parse_dt(value)
    if not dt:
        return None
    now = datetime.now(dt.tzinfo)
    return round(max(0.0, (now - dt).total_seconds() / 86400.0), 1)


async def load_stage_names(api: Bitrix24API) -> dict[str, str]:
    categories = (await api.call("crm.dealcategory.list", {})).get("result", []) or []
    out: dict[str, str] = {}
    for category in categories:
        cid = str(category.get("ID") or "")
        entity_id = "DEAL_STAGE" if cid in {"", "0"} else f"DEAL_STAGE_{cid}"
        rows = (await api.call("crm.status.list", {"filter": {"ENTITY_ID": entity_id}})).get(
            "result", []
        ) or []
        for row in rows:
            status_id = str(row.get("STATUS_ID") or "")
            name = str(row.get("NAME") or "")
            if status_id and name:
                out[status_id] = name
                if cid and ":" not in status_id:
                    out[f"C{cid}:{status_id}"] = name
    return out


async def fetch_kso_deals(
    api: Bitrix24API,
    sales_ids: list[str],
    *,
    date_from: str | None = None,
    date_to_exclusive: str | None = None,
) -> list[dict[str, Any]]:
    flt = {KSO_FIELD: KSO_VALUE_ID, "ASSIGNED_BY_ID": sales_ids}
    if date_from:
        flt[">=DATE_CREATE"] = date_from
    if date_to_exclusive:
        flt["<DATE_CREATE"] = date_to_exclusive
    select = [
        "ID",
        "TITLE",
        "ASSIGNED_BY_ID",
        "STAGE_ID",
        "STAGE_SEMANTIC_ID",
        "CATEGORY_ID",
        "DATE_CREATE",
        "DATE_MODIFY",
        "CLOSEDATE",
        "COMMENTS",
        "SOURCE_ID",
        "SOURCE_DESCRIPTION",
        "OPPORTUNITY",
        KSO_FIELD,
    ]
    out: list[dict[str, Any]] = []
    start: int | None = 0
    total: int | None = None
    while start is not None:
        res = await api.call(
            "crm.deal.list",
            {
                "filter": flt,
                "select": select,
                "order": {"ID": "DESC"},
                "start": start,
            },
        )
        chunk = res.get("result", []) or []
        total = safe_int(res.get("total")) or total
        out.extend(chunk)
        print(f"[KSO] deals {len(out)}/{total or '?'}", flush=True)
        next_start = res.get("next")
        start = safe_int(next_start) if next_start is not None else None
        await asyncio.sleep(0.2)
    return out


def base_row(deal: dict[str, Any], manager_names: dict[str, str], stage_names: dict[str, str]) -> dict[str, Any]:
    deal_id = str(deal.get("ID") or "")
    stage_id = str(deal.get("STAGE_ID") or "")
    won = is_won(deal)
    texts = [
        _clean(deal.get("TITLE")),
        _clean(deal.get("COMMENTS")),
        _clean(deal.get("SOURCE_DESCRIPTION")),
        stage_names.get(stage_id, stage_id),
    ]
    reason = (
        {
            "reason_category": "Выиграна",
            "reason_confidence": "Высокая",
            "reason_evidence": "",
            "recommendation": "Использовать как положительный пример для сравнения с невыигранными КСО-сделками.",
        }
        if won
        else classify_reason(texts, stage_id)
    )
    manager_id = str(deal.get("ASSIGNED_BY_ID") or "")
    return {
        "deal_id": deal_id,
        "deal_url": deal_url_from_id(DEFAULT_DOMAIN, deal_id),
        "title": _clean(deal.get("TITLE")),
        "manager_id": manager_id,
        "manager_name": manager_names.get(manager_id, manager_id),
        "category_id": deal.get("CATEGORY_ID"),
        "kso_category": KSO_VALUE_NAME,
        "stage_id": stage_id,
        "stage_name": stage_names.get(stage_id, stage_id),
        "stage_state": stage_state(deal),
        "is_won": won,
        "is_failed": is_failed(deal),
        "date_create": deal.get("DATE_CREATE"),
        "date_modify": deal.get("DATE_MODIFY"),
        "close_date": deal.get("CLOSEDATE"),
        "lifetime_days": lifetime_days(deal),
        "days_since_update": days_since_update(deal.get("DATE_MODIFY")),
        "amount": as_float(deal.get("OPPORTUNITY")),
        "source": _clean(deal.get("SOURCE_DESCRIPTION")) or _clean(deal.get("SOURCE_ID")),
        "comments": _clean(deal.get("COMMENTS")),
        "context_enriched": False,
        "comments_count": 0,
        "activities_count": 0,
        "next_steps_count": 0,
        "calls_count": 0,
        "cached_transcripts_count": 0,
        "chat_activity_count": 0,
        "bitrix_chat_summary": "",
        "bitrix_call_summary": "",
        "context_excerpt": "",
        "has_meaningful_comment": has_meaningful_comment(deal.get("COMMENTS")),
        "has_next_step": None,
        **reason,
    }


def score_manager_work(row: dict[str, Any]) -> tuple[int, str]:
    score = 0
    issues: list[str] = []
    if row.get("has_meaningful_comment"):
        score += 25
    else:
        issues.append("нет содержательного комментария")

    if row.get("context_enriched") and (
        row.get("comments_count") or row.get("activities_count") or row.get("context_excerpt")
    ):
        score += 15
    else:
        issues.append("мало контекста в карточке")

    if row.get("is_won"):
        score += 30
    elif row.get("reason_category") != "Причина не указана / мало контекста":
        score += 30
    else:
        issues.append("не зафиксирована понятная причина")

    if row.get("is_failed") or row.get("is_won"):
        score += 10
    elif row.get("context_enriched") and row.get("has_next_step"):
        score += 10
    elif row.get("stage_state") == "В работе / не выиграна":
        issues.append("следующий шаг не подтвержден")

    if row.get("calls_count") or row.get("chat_activity_count") or row.get("cached_transcripts_count"):
        score += 10
    else:
        issues.append("нет подтвержденных звонков/чатов в анализе")

    if row.get("amount") and float(row.get("amount") or 0) > 0:
        score += 10
    else:
        issues.append("нет суммы/потенциала")

    return min(100, score), "; ".join(issues) if issues else "замечаний нет"


def apply_context(base: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    out = {**base, **context}
    if out.get("is_won"):
        out["reason_category"] = "Выиграна"
        out["reason_confidence"] = "Высокая"
        out["recommendation"] = "Использовать как положительный пример для сравнения."
    score, issues = score_manager_work(out)
    out["manager_work_score"] = score
    out["manager_work_issues"] = issues
    out["chat_coverage_status"] = chat_coverage_status(out)
    out["deal_meaning"] = build_deal_meaning(out)
    return out


def chat_coverage_status(row: dict[str, Any]) -> str:
    if row.get("bitrix_chat_summary"):
        return "Есть резюме чата BitrixGPT"
    if int(row.get("chat_activity_count") or 0) > 0:
        return "Есть чат-активности, но нет резюме чата"
    return "Чат/мессенджер не найден"


def build_deal_meaning(row: dict[str, Any]) -> str:
    parts = [
        row.get("bitrix_chat_summary"),
        row.get("bitrix_call_summary"),
        row.get("reason_evidence"),
        row.get("context_excerpt"),
        row.get("comments"),
    ]
    text = " | ".join(_clean(part) for part in parts if _clean(part))
    return _short(text, 1600)


def build_summary(
    rows: list[dict[str, Any]],
    *,
    date_from: str | None = None,
    date_to_exclusive: str | None = None,
) -> list[dict[str, Any]]:
    total = len(rows)
    won = sum(1 for row in rows if row.get("is_won"))
    not_won = total - won
    failed = sum(1 for row in rows if row.get("is_failed"))
    no_reason = sum(
        1
        for row in rows
        if not row.get("is_won") and row.get("reason_category") == "Причина не указана / мало контекста"
    )
    avg_score = round(sum(float(row.get("manager_work_score") or 0) for row in rows) / max(1, total), 2)
    with_chat = sum(1 for row in rows if int(row.get("chat_activity_count") or 0) > 0)
    with_chat_summary = sum(1 for row in rows if row.get("bitrix_chat_summary"))
    with_calls = sum(1 for row in rows if int(row.get("calls_count") or 0) > 0)
    with_transcripts = sum(1 for row in rows if int(row.get("cached_transcripts_count") or 0) > 0)
    return [
        {"metric": "Категория КСО", "value": KSO_VALUE_NAME},
        {"metric": "Поле Bitrix", "value": f"{KSO_FIELD}={KSO_VALUE_ID}"},
        {
            "metric": "Период создания сделок",
            "value": f"{date_from or 'без нижней границы'} — {date_to_exclusive or 'без верхней границы'} (верхняя дата не включается)",
        },
        {"metric": "Всего КСО-сделок отдела продаж", "value": total},
        {"metric": "Выиграно", "value": won},
        {"metric": "Не выиграно", "value": not_won},
        {"metric": "Финально проиграно/провалено", "value": failed},
        {"metric": "Без понятной причины среди невыигранных", "value": no_reason},
        {"metric": "Средняя оценка работы менеджера", "value": avg_score},
        {"metric": "Сделок с чатами/мессенджерами", "value": with_chat},
        {"metric": "Сделок с резюме чата BitrixGPT", "value": with_chat_summary},
        {"metric": "Сделок со звонками", "value": with_calls},
        {"metric": "Сделок с кэшированной расшифровкой звонка", "value": with_transcripts},
        {"metric": "Дата отчета", "value": datetime.now().isoformat(timespec="seconds")},
    ]


def build_reason_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    not_won = [row for row in rows if not row.get("is_won")]
    total = max(1, len(not_won))
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in not_won:
        grouped[str(row.get("reason_category") or "Причина не указана / мало контекста")].append(row)
    out: list[dict[str, Any]] = []
    for reason, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        managers = Counter(str(row.get("manager_name") or "") for row in items)
        stages = Counter(str(row.get("stage_name") or "") for row in items)
        examples = "\n".join(f"{row.get('title')} ({row.get('deal_url')})" for row in items[:8])
        rec = next((str(row.get("recommendation") or "") for row in items if row.get("recommendation")), "")
        out.append(
            {
                "reason_category": reason,
                "deals_count": len(items),
                "share_percent": round(len(items) * 100.0 / total, 2),
                "top_managers": "; ".join(f"{name}: {count}" for name, count in managers.most_common(5)),
                "top_stages": "; ".join(f"{name}: {count}" for name, count in stages.most_common(5)),
                "avg_manager_work_score": round(
                    sum(float(row.get("manager_work_score") or 0) for row in items) / max(1, len(items)),
                    2,
                ),
                "recommendation": rec,
                "example_deals": examples,
            }
        )
    return out


def build_manager_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("manager_name") or "")].append(row)
    out: list[dict[str, Any]] = []
    for manager, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        won = sum(1 for row in items if row.get("is_won"))
        not_won = len(items) - won
        reasons = Counter(str(row.get("reason_category") or "") for row in items if not row.get("is_won"))
        out.append(
            {
                "manager_name": manager,
                "deals_count": len(items),
                "won_count": won,
                "not_won_count": not_won,
                "win_rate_percent": round(won * 100.0 / max(1, len(items)), 2),
                "avg_manager_work_score": round(
                    sum(float(row.get("manager_work_score") or 0) for row in items) / max(1, len(items)),
                    2,
                ),
                "no_reason_count": sum(
                    1
                    for row in items
                    if not row.get("is_won") and row.get("reason_category") == "Причина не указана / мало контекста"
                ),
                "no_comment_count": sum(1 for row in items if not row.get("has_meaningful_comment")),
                "top_not_won_reasons": "; ".join(
                    f"{name}: {count}" for name, count in reasons.most_common(5) if name
                ),
            }
        )
    return out


def build_recommendations(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reason_rows = build_reason_rows(rows)
    manager_rows = build_manager_rows(rows)
    out: list[dict[str, Any]] = []
    if reason_rows:
        top = reason_rows[0]
        out.append(
            {
                "priority": "Критично",
                "area": "Причины невыигрыша КСО",
                "finding": f"{top['reason_category']}: {top['deals_count']} сделок ({top['share_percent']}%).",
                "action": top["recommendation"],
            }
        )
    no_reason = sum(
        1
        for row in rows
        if not row.get("is_won") and row.get("reason_category") == "Причина не указана / мало контекста"
    )
    if no_reason:
        out.append(
            {
                "priority": "Критично",
                "area": "CRM-дисциплина",
                "finding": f"{no_reason} невыигранных КСО-сделок без понятной причины.",
                "action": "При закрытии/провале КСО сделать обязательным: потребность, модель КСО/киоска, бюджет, техническое ограничение, следующий шаг или причина отказа.",
            }
        )
    for manager in manager_rows:
        if manager["not_won_count"] and float(manager["avg_manager_work_score"]) < 70:
            out.append(
                {
                    "priority": "Высокий",
                    "area": f"Менеджер: {manager['manager_name']}",
                    "finding": f"КСО-сделок: {manager['deals_count']}, выиграно: {manager['won_count']}, средняя оценка: {manager['avg_manager_work_score']}.",
                    "action": "Разобрать каждую невыигранную КСО-сделку: где клиент остановился, что предложено, был ли тех. пресейл, какой следующий шаг.",
                }
            )
    return out


def build_source_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        out.append(
            {
                "Сделка": row.get("deal_url"),
                "Название": row.get("title"),
                "Менеджер": row.get("manager_name"),
                "Стадия": row.get("stage_name"),
                "Состояние": row.get("stage_state"),
                "Причина": row.get("reason_category"),
                "Комментариев в таймлайне": row.get("comments_count"),
                "Активностей": row.get("activities_count"),
                "Следующих дел": row.get("next_steps_count"),
                "Звонков": row.get("calls_count"),
                "Расшифровок в кэше": row.get("cached_transcripts_count"),
                "Чат-активностей": row.get("chat_activity_count"),
                "Покрытие чата": row.get("chat_coverage_status"),
                "Есть комментарий": "да" if row.get("has_meaningful_comment") else "нет",
                "Есть следующий шаг": "да" if row.get("has_next_step") else "нет",
                "Оценка работы менеджера": row.get("manager_work_score"),
            }
        )
    return out


def build_chat_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        if int(row.get("chat_activity_count") or 0) <= 0 and not row.get("bitrix_chat_summary"):
            continue
        out.append(
            {
                "Сделка": row.get("deal_url"),
                "Название": row.get("title"),
                "Менеджер": row.get("manager_name"),
                "Стадия": row.get("stage_name"),
                "Причина": row.get("reason_category"),
                "Чат-активностей": row.get("chat_activity_count"),
                "Покрытие чата": row.get("chat_coverage_status"),
                "Резюме чата BitrixGPT": row.get("bitrix_chat_summary"),
                "Резюме звонка BitrixGPT": row.get("bitrix_call_summary"),
                "Общий смысл сделки": row.get("deal_meaning"),
                "Что мешает продаже": row.get("reason_evidence"),
                "Рекомендация": row.get("recommendation"),
            }
        )
    return out


def build_context_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        out.append(
            {
                "Сделка": row.get("deal_url"),
                "Название": row.get("title"),
                "Менеджер": row.get("manager_name"),
                "Стадия": row.get("stage_name"),
                "Состояние": row.get("stage_state"),
                "Причина": row.get("reason_category"),
                "Уверенность причины": row.get("reason_confidence"),
                "Общий смысл сделки": row.get("deal_meaning"),
                "Доступный контекст": row.get("context_excerpt"),
                "Проблемы работы менеджера": row.get("manager_work_issues"),
                "Рекомендация": row.get("recommendation"),
            }
        )
    return out


def write_excel(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(payload["summary"]).to_excel(writer, sheet_name="Итоги", index=False)
        pd.DataFrame(payload["reasons"]).to_excel(writer, sheet_name="Причины", index=False)
        pd.DataFrame(payload["managers"]).to_excel(writer, sheet_name="Менеджеры", index=False)
        pd.DataFrame(payload["sources"]).to_excel(writer, sheet_name="Источники", index=False)
        pd.DataFrame(payload["chats"]).to_excel(writer, sheet_name="Чаты", index=False)
        pd.DataFrame(payload["contexts"]).to_excel(writer, sheet_name="Контекст сделок", index=False)
        pd.DataFrame(payload["deals"]).to_excel(writer, sheet_name="Сделки", index=False)
        pd.DataFrame(payload["recommendations"]).to_excel(writer, sheet_name="Рекомендации", index=False)
        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True, name="Arial")
        body_font = Font(name="Arial", size=10)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=idx).value or "").lower()
                width = 18
                if any(token in header for token in ("url", "summary", "excerpt", "evidence", "recommendation", "issues", "finding", "action", "example", "comment")):
                    width = 48
                elif any(token in header for token in ("title", "reason", "stage", "manager")):
                    width = 30
                ws.column_dimensions[get_column_letter(idx)].width = min(width, 65)


def copy_latest(src: Path, dst: Path) -> None:
    try:
        shutil.copy2(src, dst)
    except PermissionError as exc:
        print(f"[WARN] Не смог обновить {dst}: файл занят другим процессом. Новый отчет сохранен: {src}. {exc}", flush=True)


async def run(args: argparse.Namespace) -> tuple[Path, Path, dict[str, Any]]:
    load_dotenv(args.env_path, override=True)
    api = Bitrix24API(readonly=True)
    try:
        sales_ids, manager_names = await load_sales_users(api)
        stage_names = await load_stage_names(api)
        deals = await fetch_kso_deals(
            api,
            sales_ids,
            date_from=args.date_from,
            date_to_exclusive=args.date_to_exclusive,
        )
        state_cache = _load_state_cache()
        semaphore = asyncio.Semaphore(max(1, args.concurrency))
        context_tasks = [
            enrich_context(
                api,
                deal,
                manager_names=manager_names,
                state_cache=state_cache,
                semaphore=semaphore,
            )
            for deal in deals
        ]
        contexts: list[dict[str, Any]] = []
        for index, task in enumerate(asyncio.as_completed(context_tasks), 1):
            contexts.append(await task)
            print(f"[KSO] context {index}/{len(context_tasks)}", flush=True)
    finally:
        await api.aclose()

    context_by_id = {row["deal_id"]: row for row in contexts}
    rows = [
        apply_context(base_row(deal, manager_names, stage_names), context_by_id[str(deal.get("ID") or "")])
        for deal in deals
    ]
    rows.sort(key=lambda row: (row.get("is_won") is True, str(row.get("manager_name")), str(row.get("date_create"))))

    payload = {
        "summary": build_summary(
            rows,
            date_from=args.date_from,
            date_to_exclusive=args.date_to_exclusive,
        ),
        "reasons": build_reason_rows(rows),
        "managers": build_manager_rows(rows),
        "sources": build_source_rows(rows),
        "chats": build_chat_rows(rows),
        "contexts": build_context_rows(rows),
        "deals": rows,
        "recommendations": build_recommendations(rows),
    }
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_out = REPORTS_DIR / f"kso_deals_analysis_{ts}.json"
    xlsx_out = REPORTS_DIR / f"kso_deals_analysis_{ts}.xlsx"
    json_out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(xlsx_out, payload)
    copy_latest(json_out, REPORTS_DIR / "latest_kso_deals_analysis.json")
    copy_latest(xlsx_out, REPORTS_DIR / "latest_kso_deals_analysis.xlsx")
    return json_out, xlsx_out, payload


def parse_args() -> argparse.Namespace:
    default_to = (date.today() + timedelta(days=1)).isoformat()
    parser = argparse.ArgumentParser(description="Анализ КСО-сделок отдела продаж")
    parser.add_argument("--env-path", default=".env")
    parser.add_argument("--concurrency", type=int, default=3)
    parser.add_argument("--date-from", default="2026-01-01")
    parser.add_argument("--date-to-exclusive", default=default_to)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    json_out, xlsx_out, payload = asyncio.run(run(args))
    print(f"JSON: {json_out}")
    print(f"Excel: {xlsx_out}")
    for row in payload["summary"]:
        print(f"{row['metric']}: {row['value']}")
    print("Top reasons:")
    for row in payload["reasons"][:10]:
        print(f"- {row['reason_category']}: {row['deals_count']} ({row['share_percent']}%)")


if __name__ == "__main__":
    main()
