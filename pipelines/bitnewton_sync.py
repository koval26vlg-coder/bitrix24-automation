from __future__ import annotations

import argparse
import html
import hashlib
import json
import os
import re
import shutil
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests

from asr.bitnewton import BitNewtonError, env_bitnewton_asr
from bitrix.api import Bitrix24API
from bitrix.recordings import resolve_call_recording
from download_resolver import download_best_effort
from pipelines.paths import LATEST_JSON_REPORT, LATEST_XLSX_REPORT, REPORTS_DIR, TRANSCRIPTS_DIR
from pipelines.reporting import publish_latest_report
from pipelines.scoring import (
    HANDLING_RE,
    OBJECTION_RULES,
    _clean_text_for_report,
    _context,
    evaluate_call_checklist,
    evaluate_call_text,
    recalculate_overall_score,
)


AUDIO_EXTENSIONS = {".mp3", ".wav", ".ogg", ".m4a", ".mp4", ".webm", ".aac", ".opus", ".flac", ".wma"}
FIRST_RESPONSE_SLA_HOURS = 0.5
MAX_GAP_BETWEEN_CALLS_HOURS = 72.0
STAGE_STUCK_THRESHOLD_HOURS = 72.0
DEAL_TOTAL_WORK_WARNING_MINUTES = 10 * 24 * 60
DEAL_TOTAL_WORK_CRITICAL_MINUTES = 14 * 24 * 60
DEFAULT_STAGE_WARNING_MINUTES = 3 * 24 * 60
DEFAULT_STAGE_CRITICAL_MINUTES = 5 * 24 * 60
STAGE_DURATION_THRESHOLDS_MINUTES: Dict[str, Tuple[int, int]] = {
    "C1:NEW": (30, 3 * 24 * 60),
    "C1:PREPARATION": (24 * 60, 2 * 24 * 60),
    "C1:PREPAYMENT_INVOICE": (2 * 24 * 60, 4 * 24 * 60),
    "C1:FINAL_INVOICE": (5 * 24 * 60, 8 * 24 * 60),
    "C1:EXECUTING": (2 * 24 * 60, 3 * 24 * 60),
    "C1:UC_9NU15J": (3 * 24 * 60, 5 * 24 * 60),
    "C1:UC_50AD9V": (5 * 24 * 60, 7 * 24 * 60),
    "C1:UC_KLTOFA": (7 * 24 * 60, 10 * 24 * 60),
    "C1:UC_6SX2WL": (24 * 60, 2 * 24 * 60),
}
DEFAULT_KPI_CONFIG: Dict[str, Any] = {
    "profile": {"name": "default", "version": "1"},
    "sla": {"first_response_hours": 0.5, "max_gap_between_calls_hours": 72.0},
    "weights": {
        "overall": {"call_quality": 0.50, "discipline": 0.0, "crm_alignment": 0.50},
        "discipline_split": {"first_response": 1.0, "cadence": 0.0},
        "crm_alignment_split": {"deal_quality": 0.6, "alignment": 0.4},
    },
    "deal_quality_weights": {"has_contact": 25, "has_amount": 25, "has_title": 25, "has_comments": 25},
    "call_quality_weights": {"greeting": 25, "needs_discovery": 30, "objection_work": 20, "next_step": 25},
    "alignment_weights": {"title_hit": 10, "title_hit_cap": 40, "amount_mentioned": 30, "next_step_synced": 30},
    "patterns": {
        "greeting": [r"\b(добрый|здравств\w*|привет)\b"],
        "needs_discovery": [r"\b(потребност\w*|задач\w*|нужно|необходим\w*|интересу\w*|уточн\w*|подскаж\w*|что нужно|что необходимо)\b"],
        "objection_work": [r"\b(дорог\w*|возраж\w*|сомнен\w*|не подходит|не смож\w*|не получится|нет возможности|подума\w*)\b"],
        "next_step": [r"\b(договор\w*|следующ\w*|перезвон\w*|отправ\w*|вышл\w*|встреч\w*|созвон\w*|уточн\w*|согласу\w*)\b"],
    },
}

STATE_CACHE_PATH = REPORTS_DIR / "state_cache.json"


RU_COLUMNS: Dict[str, str] = {
    "deal_id": "ID сделки",
    "deal_url": "Ссылка на сделку",
    "stage_id": "Стадия",
    "stage_name": "Стадия",
    "stage_history_event_id": "ID события стадии",
    "stage_history_event_type": "Тип события стадии",
    "stage_history_created_at": "Дата смены стадии",
    "stage_history_count": "Кол-во переходов по стадиям",
    "stage_history_path": "Маршрут стадий",
    "stage_duration_minutes": "Время до следующей стадии, мин.",
    "stage_duration_hours": "Время до следующей стадии, ч.",
    "stage_is_current": "Текущая стадия",
    "stage_last_change_at": "Последняя смена стадии",
    "stage_current_age_minutes": "Время в текущей стадии, мин.",
    "stage_current_age_days": "Время в текущей стадии, дн.",
    "stage_warning_threshold": "Порог предупреждения по стадии",
    "stage_critical_threshold": "Порог тревоги по стадии",
    "stage_current_age_status": "Статус срока в стадии",
    "deal_total_work_minutes": "Общее время в работе, мин.",
    "deal_total_work_days": "Общее время в работе, дн.",
    "deal_total_work_warning_threshold": "Порог предупреждения по сделке",
    "deal_total_work_critical_threshold": "Порог тревоги по сделке",
    "deal_total_work_status": "Статус общего срока сделки",
    "stage_return_count": "Возвраты на предыдущие стадии",
    "stage_reached_final": "Достигнута финальная стадия",
    "stage_movement_risk": "Риск движения по воронке",
    "stage_movement_recommendation": "Рекомендация по движению сделки",
    "stage_deals_total": "Всего сделок",
    "stage_deals_passed": "Сделок прошло стадию",
    "stage_sr": "SR стадии, %",
    "stage_order": "Порядок стадии",
    "manager_id": "ID менеджера",
    "manager_name": "Менеджер",
    "kpi_profile": "Профиль KPI",
    "kpi_profile_ru": "Профиль KPI",
    "kpi_explanation": "Расшифровка KPI",
    "kpi_version": "Версия KPI",
    "kpi_profile_cmp": "Профиль KPI сравнения",
    "kpi_version_cmp": "Версия KPI сравнения",
    "activity_id": "ID активности",
    "origin_id": "ID звонка Bitrix",
    "subject": "Тема звонка",
    "duration_minutes": "Длительность звонка, мин.",
    "disk_file_id": "ID файла записи",
    "download_url": "Ссылка на запись",
    "audio_path": "Путь к аудио",
    "local_audio_source_used": "Использован локальный источник аудио",
    "local_audio_source_path": "Локальный источник аудио",
    "bitnewton_task_id": "ID задачи Bit.Newton",
    "attach_result": "Результат прикрепления в Bitrix",
    "error": "Ошибка",
    "calls_count": "Кол-во звонков по сделке",
    "first_response_minutes": "Время до первого звонка, мин.",
    "reaction_speed_label": "Скорость реакции менеджера",
    "first_response_sla_ok": "Первый контакт в срок",
    "first_response_explanation": "Что значит первый контакт",
    "deal_quality_score": "Оценка заполнения сделки",
    "deal_quality_details": "Критерии заполнения сделки",
    "has_contact": "Есть контакт/компания",
    "has_amount": "Указана сумма",
    "has_title": "Есть название",
    "has_comments": "Есть комментарии/следующие шаги",
    "recording_diagnostics": "Диагностика записи",
    "download_attempts": "Попытки REST-скачивания",
    "ui_download_errors": "Ошибки UI-скачивания",
    "ui_download_used": "Использовано UI-скачивание",
    "ui_download_path": "Путь UI-скачивания",
    "audio_size_bytes": "Размер аудио, байт",
    "transcript_excerpt": "Фрагмент расшифровки",
    "transcript_text": "Полная расшифровка",
    "transcript_marked": "Расшифровка с пометками",
    "transcript_path": "Файл расшифровки",
    "transcript_hash": "Хеш расшифровки",
    "call_quality_score": "Оценка качества разговора",
    "call_quality_details": "Критерии качества разговора",
    "call_checklist_total_score": "Баллы по чек-листу звонка",
    "call_checklist_max_score": "Максимум чек-листа звонка",
    "call_checklist_percent": "Чек-лист звонка, %",
    "call_checklist_block_details": "Оценка этапов продаж",
    "checklist_block_name": "Этап продаж",
    "checklist_criterion": "Критерий чек-листа",
    "checklist_score": "Оценка критерия",
    "checklist_max_score": "Максимум критерия",
    "checklist_evidence": "Доказательство из расшифровки",
    "checklist_comment": "Комментарий оценки",
    "checklist_code": "Код критерия",
    "sales_stage_block_name": "Этап продаж",
    "sales_stage_score": "Баллы этапа",
    "sales_stage_max_score": "Максимум этапа",
    "sales_stage_percent": "Выполнение этапа, %",
    "sales_stage_missing": "Что усилить по этапу",
    "manager_stage_calls": "Звонков с оценкой этапа",
    "manager_stage_deals": "Сделок с оценкой этапа",
    "manager_stage_avg_percent": "Среднее выполнение этапа, %",
    "manager_stage_weak_calls": "Звонков ниже нормы",
    "manager_stage_weak_rate": "Доля слабых звонков, %",
    "criterion_calls": "Звонков с критерием",
    "criterion_avg_score": "Средняя оценка критерия",
    "criterion_completion_percent": "Выполнение критерия, %",
    "criterion_failed_count": "Провалов критерия",
    "criterion_partial_count": "Частично выполнено",
    "criterion_fail_rate": "Доля проблем, %",
    "training_recommendation": "Рекомендация для обучения",
    "top_checklist_gaps": "Главные провалы чек-листа",
    "has_greeting": "Есть приветствие",
    "has_needs_discovery": "Выявлены потребности",
    "has_objection_work": "Есть работа с возражениями",
    "has_next_step_phrase": "Озвучен следующий шаг",
    "alignment_score": "Связь звонка с CRM",
    "crm_work_score": "Оценка ведения CRM",
    "alignment_details": "Что значит связь звонка с CRM",
    "title_mentions": "Совпадения с названием сделки",
    "amount_mentioned": "Сумма упомянута в разговоре",
    "next_step_synced": "Следующий шаг синхронизирован с CRM",
    "next_step_synced_details": "Что дает синхронизация следующего шага",
    "crm_checklist_total_score": "Баллы по CRM-чек-листу",
    "crm_checklist_max_score": "Максимум CRM-чек-листа",
    "crm_checklist_percent": "CRM-чек-лист, %",
    "crm_checklist_details": "Критерии ведения CRM",
    "crm_checklist_block_name": "Блок CRM",
    "crm_checklist_criterion": "Критерий CRM",
    "crm_checklist_score": "Оценка CRM-критерия",
    "crm_checklist_max_score": "Максимум CRM-критерия",
    "crm_checklist_comment": "Комментарий CRM-оценки",
    "crm_checklist_code": "Код CRM-критерия",
    "crm_criterion_deals": "Сделок с критерием",
    "crm_criterion_avg_score": "Средняя оценка CRM-критерия",
    "crm_criterion_completion_percent": "Выполнение CRM-критерия, %",
    "crm_criterion_failed_count": "Провалов CRM-критерия",
    "crm_criterion_partial_count": "Частично выполнено CRM",
    "crm_criterion_fail_rate": "Доля CRM-проблем, %",
    "control_priority": "Приоритет контроля",
    "control_priority_order": "Порядок приоритета",
    "control_reason": "Причина контроля",
    "control_next_action": "Что сделать дальше",
    "quality_control_score": "Оценка риска контроля",
    "training_source": "Источник обучения",
    "training_topic": "Тема обучения",
    "coaching_priority": "Приоритет обучения",
    "coaching_metric": "Метрика провала",
    "coaching_affected_count": "Сколько кейсов затронуто",
    "summary_section": "Раздел",
    "summary_metric": "Показатель",
    "summary_value": "Значение",
    "summary_comment": "Комментарий",
    "manager_rank": "Место",
    "manager_focus": "Фокус руководителя",
    "manager_critical_deals": "Критичных сделок",
    "manager_high_deals": "Высокий приоритет сделок",
    "manager_training_topics": "Тем обучения",
    "manager_top_training_topic": "Главная тема обучения",
    "manager_next_action": "Следующее действие руководителя",
    "overall_score": "Итоговая оценка",
    "overall_score_details": "Из чего складывается итоговая оценка",
    "overall_score_cmp": "Итоговая оценка сравнения",
    "overall_score_delta": "Разница итоговой оценки",
    "call_quality_score_cmp": "Качество разговора сравнение",
    "deal_quality_score_cmp": "Заполнение сделки сравнение",
    "alignment_score_cmp": "Связь звонка с CRM сравнение",
    "call_quality_conclusion": "Вывод по разговору",
    "client_work_conclusion": "Вывод по проработке клиента",
    "improvement_moments": "Моменты для улучшения",
    "objections_count": "Кол-во возражений",
    "unhandled_objections_count": "Кол-во неотработанных возражений",
    "objections_handled": "Возражения отработаны",
    "objections_found": "Найденные возражения",
    "unhandled_objections": "Неотработанные возражения",
    "objection_recommendations": "Варианты отработки возражений",
    "objection_status": "Статус отработки",
    "objection_fragment": "Фрагмент с возражением",
    "recommendations": "Рекомендации",
    "calls_total": "Всего звонков",
    "calls_ok": "Успешно обработано",
    "overall_score_sum": "Сумма итоговых оценок",
    "growth_deal_data": "Зона роста: данные CRM",
    "growth_call_structure": "Зона роста: структура разговора",
    "growth_alignment": "Зона роста: связь с CRM",
    "avg_overall_score": "Средняя итоговая оценка",
    "top_growth_zones": "Главные зоны роста",
    "avg_call_quality_score": "Средняя оценка разговоров",
    "client_work_score": "Оценка проработки клиента",
    "client_work_quality": "Качество проработки клиента",
    "conversation_quality": "Качество разговоров",
    "transcripts_count": "Кол-во расшифровок",
    "transcript_paths": "Файлы расшифровок",
    "no_calls": "Нет звонков по сделке",
    "deals_without_calls": "Сделок без звонков",
    "deals_total": "Всего сделок",
    "calls_failed": "Ошибок по звонкам",
    "ignored_call_center_calls": "Исключено звонков Call-центра",
    "calls_breakdown": "Звонки внутри сделки",
    "transcripts_combined": "Все расшифровки по сделке",
    "bitrix_card_transcript": "Расшифровка из карточки Bitrix",
    "bitrix_card_transcript_status": "Статус расшифровки Bitrix",
    "transcript_match_score": "Совпадение Bit.Newton/Bitrix, %",
    "conversation_meaning": "Смысл разговора",
    "improvement_moments_combined": "Общие моменты для улучшения",
    "objections_combined": "Возражения по сделке",
    "call_number": "Номер звонка в сделке",
    "call_summary": "Кратко по звонку",
    "call_has_error": "Ошибка по звонку",
    "deal_row_type": "Тип строки",
}

DEFAULT_STAGE_NAMES: Dict[str, str] = {
    "C1:NEW": "Ответственный назначен",
    "C1:PREPARATION": "Потребность выявлена",
    "C1:PREPAYMENT_INVOICE": "Тех. пресейл назначен",
    "C1:FINAL_INVOICE": "Тех. пресейл проведен",
    "C1:EXECUTING": "Защита КП проведена",
    "C1:UC_9NU15J": "Счет отправлен",
    "C1:UC_50AD9V": "Оплата получена",
    "C1:UC_KLTOFA": "Передан в Тех деп.",
    "C1:UC_6SX2WL": "Возврат сделки в работу от РОП",
    "C1:WON": "Успешно завершена",
    "C1:UC_AHHIBG": "Сделка проиграна на проверке РОП",
    "C1:LOSE": "Сделка проиграна",
    "C1:UC_62ISSK": "ВОЗВРАТ",
    "C1:UC_2EN0IE": "ДУБЛЬ",
    "NEW": "Новая",
    "WON": "Сделка успешна",
    "LOSE": "Сделка проиграна",
}

KPI_PROFILE_DESCRIPTIONS: Dict[str, Tuple[str, str]] = {
    "example": ("Базовый контроль звонка и CRM", "Сбалансированная оценка разговора, заполнения сделки, дисциплины касаний и связи звонка с CRM."),
    "default": ("Базовый контроль звонка и CRM", "Сбалансированная оценка разговора, заполнения сделки, дисциплины касаний и связи звонка с CRM."),
    "sales": ("Продажи: стандартный контроль менеджера", "Основной рабочий профиль для продаж: выявление потребностей, следующий шаг, работа с возражениями, SLA и связь с CRM."),
    "sales_soft": ("Продажи: мягкая оценка диалога", "Больше веса качеству разговора и следующему шагу, мягче требования к скорости первого контакта и паузам между касаниями."),
    "sales_strict": ("Продажи: строгая дисциплина и результат", "Жесткий контроль скорости реакции, регулярности касаний, выявления потребностей, возражений и фиксации следующего шага."),
}


def _looks_like_html_prefix(data: bytes) -> bool:
    if not data:
        return True
    head = data.lstrip()[:200].lower()
    return head.startswith(b"<!doctype") or head.startswith(b"<html") or head.startswith(b"<head") or head.startswith(b"<body")


def validate_audio_file(path: Path) -> Optional[str]:
    """
    Быстрая валидация, чтобы не отправлять в ASR HTML/пустышки.
    """
    try:
        if not path.exists() or path.stat().st_size <= 0:
            return "файл не создан или пустой"
        if path.stat().st_size < 2048:
            return f"файл слишком маленький ({path.stat().st_size} bytes)"
        head = path.read_bytes()[:512]
        if _looks_like_html_prefix(head):
            return "вместо аудио скачался HTML (логин/нет прав)"
    except Exception:
        return None
    return None


def parse_audio_source_dirs(values: Any) -> List[Path]:
    out: List[Path] = []
    raw_values = values if isinstance(values, list) else ([values] if values else [])
    for raw in raw_values:
        for part in str(raw or "").split(";"):
            text = part.strip().strip('"')
            if text:
                out.append(Path(text))
    return out


def build_audio_source_index(source_dirs: List[Path]) -> List[Path]:
    files: List[Path] = []
    seen: set[str] = set()
    for source_dir in source_dirs:
        try:
            base = Path(source_dir)
            if not base.exists() or not base.is_dir():
                print(f"[WARN] Локальная папка аудио не найдена: {base}", flush=True)
                continue
            for path in base.rglob("*"):
                if not path.is_file() or path.suffix.lower() not in AUDIO_EXTENSIONS:
                    continue
                key = str(path.resolve()).lower()
                if key in seen:
                    continue
                seen.add(key)
                files.append(path)
        except Exception as e:
            print(f"[WARN] Не удалось просканировать папку аудио {source_dir}: {e}", flush=True)
    return sorted(files, key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)


def find_local_audio_source(
    audio_index: List[Path],
    deal_id: Any,
    activity_id: Any,
    disk_file_id: Any,
    call_id: Any,
    disk_file_name: Any = None,
) -> Optional[Path]:
    if not audio_index:
        return None

    disk_name = str(disk_file_name or "").strip().lower()
    strong_tokens = [
        str(disk_file_id or "").strip(),
        str(activity_id or "").strip(),
        str(call_id or "").strip(),
        str(call_id or "").replace("VI_", "").strip(),
    ]
    weak_tokens = [str(deal_id or "").strip()]
    strong_tokens = [t.lower() for t in strong_tokens if len(t.strip()) >= 3]
    weak_tokens = [t.lower() for t in weak_tokens if len(t.strip()) >= 4]

    ranked: List[Tuple[int, float, Path]] = []
    for path in audio_index:
        try:
            name = path.name.lower()
            stem = path.stem.lower()
            score = 0
            if disk_name and name == disk_name:
                score += 120
            elif disk_name and (disk_name in name or name in disk_name):
                score += 90
            if any(token and token in name for token in strong_tokens):
                score += 80
            if any(token and token in stem for token in strong_tokens):
                score += 40
            if any(token and token in name for token in weak_tokens):
                score += 20
            if score >= 80:
                ranked.append((score, path.stat().st_mtime if path.exists() else 0, path))
        except Exception:
            continue

    if not ranked:
        return None
    ranked.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return ranked[0][2]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Bit.Newton: звонок → аудио → ASR → attachtranscription")
    parser.add_argument("--mode", choices=["single", "filter"], default=None)
    parser.add_argument("--deal-id", default=None, help="ID сделки (для mode=single)")
    parser.add_argument("--deal-url", default=None, help="URL сделки (для mode=single, если ID не задан)")
    parser.add_argument("--filter-json", default=None, help="Путь к JSON фильтру для crm.deal.list (для mode=filter)")
    parser.add_argument("--limit", type=int, default=200)
    parser.add_argument("--diarize", action="store_true", help="Bit.Newton diarize")
    parser.add_argument("--use-bitnewton", action="store_true", help="Включить Bit.Newton (нужен BITNEWTON_TOKEN)")
    parser.add_argument("--bitnewton-flow", action="store_true", help="Идти строго: call→audio→bitnewton→attach")
    parser.add_argument("--download-audio", action="store_true", help="Сохранять аудио в reports/audio/")
    parser.add_argument(
        "--audio-source-dir",
        action="append",
        default=[],
        help="Локальная папка с уже доступными аудиозаписями. Можно передать несколько раз или через ;",
    )
    parser.add_argument("--ui-download", action="store_true", help="Если REST-скачивание недоступно — попытаться скачать через Chrome (нужен логин в профиле)")
    parser.add_argument("--ui-browser", choices=["chrome", "edge"], default="chrome", help="Какой браузер использовать для UI-fallback (edge безопаснее для рабочего Chrome)")
    parser.add_argument("--ui-download-mode", choices=["direct", "timeline", "auto"], default="auto", help="UI способ: direct=точная CRM-ссылка записи, timeline=кнопка Скачать в таймлайне, auto=direct->timeline")
    parser.add_argument("--ui-timeout-sec", type=int, default=20, help="Сколько ждать обычное UI-скачивание; ручной вход в Bitrix ждём отдельно минимум 120 секунд")
    parser.add_argument("--rest-timeout-sec", type=int, default=20, help="Сколько ждать REST-скачивание одного URL записи")
    parser.add_argument("--ui-download-dir", default=None, help="Куда скачивать через UI (по умолчанию reports/audio_ui)")
    parser.add_argument("--browser-profile-directory", default="Default", help="Имя профиля браузера внутри User Data, например Default или Profile 1")
    parser.add_argument(
        "--chrome-profile-dir",
        default=None,
        help="Папка профиля Chrome для UI-скачивания. Можно указать 'system' чтобы использовать обычный профиль Chrome (LOCALAPPDATA/Google/Chrome/User Data).",
    )
    parser.add_argument("--domain", default=os.getenv("BITRIX24_DOMAIN", "online-kassa.bitrix24.ru"))
    parser.add_argument("--kpi-config", default=None, help="Путь к JSON с порогами/весами/паттернами оценки")
    parser.add_argument("--kpi-config-compare", default=None, help="Второй KPI JSON для сравнения в этом же отчёте")
    parser.add_argument("--force-attach", action="store_true", help="Всегда делать attachtranscription, игнорируя локальный state_cache")
    parser.add_argument("--no-reuse-transcripts", action="store_true", help="Не использовать ранее сохранённые расшифровки; заново скачивать аудио и отправлять в Bit.Newton")
    parser.add_argument("--retry-errors-from", default=None, help="JSON отчета, из которого нужно повторить только строки с ошибками")
    parser.add_argument("--reevaluate-from", default=None, help="JSON отчета, который нужно переоценить без скачивания аудио и Bit.Newton")
    parser.add_argument("--max-calls-per-deal", type=int, default=0, help="Ограничить количество звонков для анализа по каждой сделке; 0 = все")
    parser.add_argument("--include-call-center", action="store_true", help="Не исключать звонки операторов Call-центра из анализа")
    parser.add_argument("--fetch-bitrix-card-transcript", action="store_true", help="Пробовать читать расшифровку из карточки звонка Bitrix через UI и сопоставлять с Bit.Newton")
    parser.add_argument("--cleanup-output-days", type=int, default=30, help="Автоудаление отчетов, расшифровок и аудио старше N дней; 0 отключает")
    parser.add_argument("--cleanup-chrome-tmp-days", type=int, default=7, help="Удалять старые reports/chrome_profile_tmp_* (дней хранения)")
    return parser


def deal_url_from_id(domain: str, deal_id: str) -> str:
    domain = (domain or "").strip().replace("https://", "").replace("http://", "").strip("/")
    return f"https://{domain}/crm/deal/details/{deal_id}/"


def safe_int(x: Any) -> Optional[int]:
    try:
        return int(x)
    except Exception:
        return None


def fetch_deals_by_filter(api: Bitrix24API, flt: Dict[str, Any], limit: int = 200) -> List[Dict[str, Any]]:
    deals: List[Dict[str, Any]] = []
    start: Optional[int] = 0
    page_size = 50
    max_items = max(0, int(limit or 0))
    if max_items <= 0:
        return deals

    while start is not None and len(deals) < max_items:
        res = api.call(
            "crm.deal.list",
            {
                "filter": flt,
                "select": ["ID", "TITLE", "ASSIGNED_BY_ID", "STAGE_ID", "DATE_CREATE"],
                "order": {"ID": "DESC"},
                "start": start,
            },
        )
        chunk = res.get("result", []) or []
        if not chunk:
            break

        remaining = max_items - len(deals)
        deals.extend(chunk[:remaining])
        if len(deals) >= max_items:
            break

        next_start = res.get("next")
        if next_start is not None:
            start = safe_int(next_start)
        else:
            total = safe_int(res.get("total"))
            start = start + page_size
            if total is None or start >= total:
                start = None

    return deals


def normalize_deal_filter_dates(flt: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(flt or {})
    date_to = out.pop("<=DATE_CREATE", None)
    if isinstance(date_to, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", date_to.strip()):
        next_day = datetime.strptime(date_to.strip(), "%Y-%m-%d") + timedelta(days=1)
        out["<DATE_CREATE"] = next_day.date().isoformat()
    elif date_to is not None:
        out["<=DATE_CREATE"] = date_to
    return out


def deal_get(api: Bitrix24API, deal_id: str) -> Dict[str, Any]:
    return api.call("crm.deal.get", {"id": int(deal_id)}).get("result", {}) or {}


def stage_entity_id_from_stage(stage_id: str) -> str:
    stage_id = str(stage_id or "")
    m = re.match(r"^C(\d+):", stage_id)
    if m:
        return f"DEAL_STAGE_{m.group(1)}"
    return "DEAL_STAGE"


def fetch_stage_name_map(api: Bitrix24API, stage_ids: List[str]) -> Dict[str, str]:
    out = dict(DEFAULT_STAGE_NAMES)
    entities = sorted({stage_entity_id_from_stage(s) for s in stage_ids if s})
    for entity in entities:
        try:
            res = api.call("crm.status.list", {"filter": {"ENTITY_ID": entity}, "order": {"SORT": "ASC"}})
            for row in res.get("result") or []:
                sid = str(row.get("STATUS_ID") or "").strip()
                name = str(row.get("NAME") or "").strip()
                if sid and name:
                    out[sid] = name
        except Exception:
            continue
    return out


def stage_display_name(stage_id: Any, stage_map: Optional[Dict[str, str]] = None) -> str:
    sid = str(stage_id or "").strip()
    if not sid:
        return ""
    return (stage_map or DEFAULT_STAGE_NAMES).get(sid, sid)


def stage_order_map(stage_map: Optional[Dict[str, str]] = None) -> Dict[str, int]:
    ordered_ids = list(DEFAULT_STAGE_NAMES.keys())
    if stage_map:
        for stage_id in stage_map.keys():
            if stage_id not in ordered_ids:
                ordered_ids.append(stage_id)
    return {stage_id: index for index, stage_id in enumerate(ordered_ids, start=1)}


def stage_history_type_label(type_id: Any) -> str:
    labels = {
        1: "Создание/попадание на стадию",
        2: "Смена стадии",
        3: "Финальная стадия",
    }
    return labels.get(safe_int(type_id), str(type_id or ""))


def _minutes_between(a: Optional[datetime], b: Optional[datetime]) -> Optional[float]:
    if not a or not b:
        return None
    try:
        return round(max(0.0, (b - a).total_seconds() / 60.0), 2)
    except Exception:
        return None


def _minutes_since(dt: Optional[datetime]) -> Optional[float]:
    if not dt:
        return None
    try:
        now = datetime.now(dt.tzinfo) if dt.tzinfo else datetime.now()
        return round(max(0.0, (now - dt).total_seconds() / 60.0), 2)
    except Exception:
        return None


def _format_minutes_for_threshold(minutes: Optional[float]) -> str:
    if minutes is None:
        return ""
    try:
        m = float(minutes)
    except Exception:
        return ""
    if m >= 24 * 60 and m % (24 * 60) == 0:
        return f"{int(m // (24 * 60))} д."
    if m >= 60 and m % 60 == 0:
        return f"{int(m // 60)} ч."
    return f"{int(round(m))} мин."


def stage_duration_thresholds(stage_id: Any) -> Tuple[int, int]:
    sid = str(stage_id or "").strip()
    return STAGE_DURATION_THRESHOLDS_MINUTES.get(sid, (DEFAULT_STAGE_WARNING_MINUTES, DEFAULT_STAGE_CRITICAL_MINUTES))


def threshold_status(value_minutes: Optional[float], warning_minutes: Optional[float], critical_minutes: Optional[float]) -> str:
    if value_minutes is None or warning_minutes is None or critical_minutes is None:
        return "Нет данных"
    value = float(value_minutes)
    if value >= float(critical_minutes):
        return "Тревога"
    if value >= float(warning_minutes):
        return "Предупреждение"
    return "OK"


def _stage_history_items_from_response(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    result = data.get("result")
    if isinstance(result, dict):
        items = result.get("items") or []
    else:
        items = result or []
    return [item for item in items if isinstance(item, dict)]


def _stage_history_next_start(data: Dict[str, Any], current_start: int, items_count: int) -> Optional[int]:
    result = data.get("result")
    next_value = data.get("next")
    if next_value is None and isinstance(result, dict):
        next_value = result.get("next")
    if next_value is not None:
        return safe_int(next_value)

    total = safe_int(data.get("total") or (result.get("total") if isinstance(result, dict) else None))
    if total is not None and current_start + items_count < total:
        return current_start + max(1, items_count)
    return None


def fetch_stage_history_by_deals(api: Bitrix24API, deal_ids: List[str], chunk_size: int = 50) -> Dict[str, List[Dict[str, Any]]]:
    out: Dict[str, List[Dict[str, Any]]] = {str(deal_id): [] for deal_id in deal_ids if deal_id}
    clean_ids = [int(d) for d in dict.fromkeys(str(deal_id) for deal_id in deal_ids if str(deal_id).isdigit())]
    for chunk_start in range(0, len(clean_ids), max(1, chunk_size)):
        chunk = clean_ids[chunk_start : chunk_start + max(1, chunk_size)]
        start: Optional[int] = 0
        while start is not None:
            params: Dict[str, Any] = {
                "entityTypeId": 2,
                "filter": {"OWNER_ID": chunk},
                "order": {"OWNER_ID": "ASC", "ID": "ASC"},
                "start": start,
            }
            data = api.call("crm.stagehistory.list", params)
            items = _stage_history_items_from_response(data)
            for item in items:
                owner_id = str(item.get("OWNER_ID") or "").strip()
                if owner_id:
                    out.setdefault(owner_id, []).append(item)
            start = _stage_history_next_start(data, int(start or 0), len(items))
            if start is not None:
                time.sleep(0.1)
    return out


def summarize_stage_history(
    items: List[Dict[str, Any]],
    stage_map: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    ranks = stage_order_map(stage_map)
    sorted_items = sorted(
        [item for item in items if isinstance(item, dict)],
        key=lambda item: (str(item.get("CREATED_TIME") or ""), safe_int(item.get("ID")) or 0),
    )
    normalized: List[Dict[str, Any]] = []
    return_count = 0
    reached_final = False
    previous_rank: Optional[int] = None

    for index, item in enumerate(sorted_items):
        stage_id = str(item.get("STAGE_ID") or "").strip()
        created_at = parse_dt(item.get("CREATED_TIME"))
        next_dt = parse_dt(sorted_items[index + 1].get("CREATED_TIME")) if index + 1 < len(sorted_items) else None
        duration_minutes = _minutes_between(created_at, next_dt)
        rank = ranks.get(stage_id)
        if previous_rank is not None and rank is not None and rank < previous_rank:
            return_count += 1
        if rank is not None:
            previous_rank = rank
        semantic = str(item.get("STAGE_SEMANTIC_ID") or "").upper()
        if semantic in {"S", "F"} or stage_id.endswith(":WON") or stage_id.endswith(":LOSE") or stage_id in {"WON", "LOSE"}:
            reached_final = True
        normalized.append(
            {
                "stage_history_event_id": item.get("ID"),
                "stage_history_event_type": stage_history_type_label(item.get("TYPE_ID")),
                "stage_history_created_at": item.get("CREATED_TIME"),
                "stage_id": stage_id,
                "stage_name": stage_display_name(stage_id, stage_map=stage_map),
                "stage_order": rank,
                "stage_duration_minutes": duration_minutes,
                "stage_duration_hours": round(duration_minutes / 60.0, 2) if duration_minutes is not None else None,
                "stage_is_current": index == len(sorted_items) - 1,
            }
        )

    current = normalized[-1] if normalized else {}
    current_stage_id = current.get("stage_id") if current else None
    stage_warning_minutes, stage_critical_minutes = stage_duration_thresholds(current_stage_id)
    current_age_minutes = _minutes_since(parse_dt(current.get("stage_history_created_at"))) if current else None
    current_age_status = threshold_status(current_age_minutes, stage_warning_minutes, stage_critical_minutes)
    first_dt = parse_dt(normalized[0].get("stage_history_created_at")) if normalized else None
    last_dt = parse_dt(normalized[-1].get("stage_history_created_at")) if normalized else None
    if reached_final:
        total_work_minutes = _minutes_between(first_dt, last_dt)
    else:
        total_work_minutes = _minutes_since(first_dt)
    total_work_status = threshold_status(
        total_work_minutes,
        DEAL_TOTAL_WORK_WARNING_MINUTES,
        DEAL_TOTAL_WORK_CRITICAL_MINUTES,
    )
    if reached_final:
        current_age_status = "Финал"
    path_names = [str(item.get("stage_name") or item.get("stage_id") or "").strip() for item in normalized]
    compact_path: List[str] = []
    for name in path_names:
        if name and (not compact_path or compact_path[-1] != name):
            compact_path.append(name)

    if not normalized:
        risk = "Нет истории"
        recommendation = "История перемещения по стадиям не найдена. Проверьте права вебхука и корректность выборки."
    elif reached_final:
        risk = "Финал"
        recommendation = "Сделка достигла финальной стадии. Для анализа важно сопоставить итог со звонками и причиной результата."
    elif current_age_status == "Тревога":
        risk = "Тревога"
        recommendation = (
            f"Сделка находится на текущей стадии дольше критического порога "
            f"({_format_minutes_for_threshold(stage_critical_minutes)}). Нужен следующий шаг или актуализация стадии."
        )
    elif total_work_status == "Тревога":
        risk = "Тревога по сроку сделки"
        recommendation = (
            f"Общее время сделки в работе выше критического порога "
            f"({_format_minutes_for_threshold(DEAL_TOTAL_WORK_CRITICAL_MINUTES)}). Проверьте причину задержки по воронке."
        )
    elif current_age_status == "Предупреждение":
        risk = "Предупреждение"
        recommendation = (
            f"Сделка приближается к критическому сроку на текущей стадии. "
            f"Предупреждение после {_format_minutes_for_threshold(stage_warning_minutes)}, "
            f"тревога после {_format_minutes_for_threshold(stage_critical_minutes)}."
        )
    elif return_count > 0:
        risk = "Возвраты"
        recommendation = "Есть возвраты на предыдущие стадии. Проверьте, не переводится ли сделка вперед без готовности клиента."
    elif len(normalized) <= 1:
        risk = "Нет продвижения"
        recommendation = "Сделка пока не двигалась по воронке. Проверьте, есть ли звонок, выявленная потребность и запланированный следующий шаг."
    else:
        risk = "OK"
        recommendation = "Движение по стадиям выглядит рабочим. Сопоставьте переходы с качеством звонков и фиксацией следующего шага."

    return {
        "stage_history_items": normalized,
        "stage_history_count": len(normalized),
        "stage_history_path": " → ".join(compact_path),
        "stage_last_change_at": current.get("stage_history_created_at"),
        "stage_current_age_minutes": current_age_minutes,
        "stage_current_age_days": round(current_age_minutes / 1440.0, 2) if current_age_minutes is not None else None,
        "stage_warning_threshold": _format_minutes_for_threshold(stage_warning_minutes),
        "stage_critical_threshold": _format_minutes_for_threshold(stage_critical_minutes),
        "stage_current_age_status": current_age_status,
        "deal_total_work_minutes": total_work_minutes,
        "deal_total_work_days": round(total_work_minutes / 1440.0, 2) if total_work_minutes is not None else None,
        "deal_total_work_warning_threshold": _format_minutes_for_threshold(DEAL_TOTAL_WORK_WARNING_MINUTES),
        "deal_total_work_critical_threshold": _format_minutes_for_threshold(DEAL_TOTAL_WORK_CRITICAL_MINUTES),
        "deal_total_work_status": total_work_status,
        "stage_return_count": return_count,
        "stage_reached_final": reached_final,
        "stage_movement_risk": risk,
        "stage_movement_recommendation": recommendation,
    }


def attach_stage_history_metrics(
    rows: List[Dict[str, Any]],
    stage_history_by_deal: Dict[str, List[Dict[str, Any]]],
    stage_map: Optional[Dict[str, str]] = None,
) -> None:
    for row in rows:
        deal_id = deal_id_from_report_row(row)
        if not deal_id:
            continue
        row.update(summarize_stage_history(stage_history_by_deal.get(str(deal_id), []), stage_map=stage_map))


def activity_get(api: Bitrix24API, activity_id: int) -> Dict[str, Any]:
    return api.call("crm.activity.get", {"id": int(activity_id)}).get("result", {}) or {}


def list_deal_call_activities(api: Bitrix24API, deal_id: str) -> List[Dict[str, Any]]:
    res = api.call(
        "crm.activity.list",
        {
            "filter": {"OWNER_TYPE_ID": 2, "OWNER_ID": str(deal_id), "TYPE_ID": 2, "PROVIDER_ID": "VOXIMPLANT_CALL"},
            "select": ["ID", "CREATED", "START_TIME", "END_TIME", "SUBJECT", "ORIGIN_ID", "DIRECTION", "PROVIDER_ID", "PROVIDER_TYPE_ID", "AUTHOR_ID", "RESPONSIBLE_ID"],
            "order": {"START_TIME": "ASC"},
            "start": 0,
        },
    )
    return res.get("result", []) or []


def user_profile(api: Bitrix24API, user_id: Any, cache: Dict[int, Dict[str, Any]]) -> Dict[str, Any]:
    uid = safe_int(user_id)
    if uid is None:
        return {}
    if uid in cache:
        return cache[uid]
    try:
        arr = api.call("user.get", {"ID": int(uid)}).get("result") or []
        cache[uid] = arr[0] if arr and isinstance(arr[0], dict) else {}
    except Exception:
        cache[uid] = {}
    return cache[uid]


def load_department_chain(api: Bitrix24API, department_ids: List[Any], cache: Dict[int, Dict[str, Any]]) -> None:
    pending = {int(x) for x in [safe_int(v) for v in department_ids] if x is not None and int(x) > 0 and int(x) not in cache}
    while pending:
        current = sorted(pending)
        pending.clear()
        try:
            rows = api.call("department.get", {"ID": current}).get("result") or []
        except Exception:
            rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            did = safe_int(row.get("ID"))
            if did is None:
                continue
            cache[did] = row
            parent = safe_int(row.get("PARENT"))
            if parent is not None and parent > 0 and parent not in cache:
                pending.add(parent)


def department_is_call_center(department_id: Any, cache: Dict[int, Dict[str, Any]]) -> bool:
    did = safe_int(department_id)
    seen: set[int] = set()
    while did is not None and did > 0 and did not in seen:
        seen.add(did)
        row = cache.get(did) or {}
        name = str(row.get("NAME") or "").lower()
        if "call центр" in name or "call center" in name or "колл" in name or "контакт центр" in name:
            return True
        did = safe_int(row.get("PARENT"))
    return False


def is_call_center_operator(
    api: Bitrix24API,
    user_id: Any,
    user_cache: Dict[int, Dict[str, Any]],
    department_cache: Dict[int, Dict[str, Any]],
) -> bool:
    user = user_profile(api, user_id, user_cache)
    if not user:
        return False
    position = str(user.get("WORK_POSITION") or "").lower()
    if "оператор" not in position and "operator" not in position:
        return False
    departments = user.get("UF_DEPARTMENT") or []
    if not isinstance(departments, list):
        departments = [departments]
    load_department_chain(api, departments, department_cache)
    return any(department_is_call_center(dept_id, department_cache) for dept_id in departments)


def split_call_center_operator_activities(
    api: Bitrix24API,
    activities: List[Dict[str, Any]],
    user_cache: Dict[int, Dict[str, Any]],
    department_cache: Dict[int, Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    kept: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    for act in activities:
        responsible_id = act.get("RESPONSIBLE_ID") or act.get("AUTHOR_ID")
        if is_call_center_operator(api, responsible_id, user_cache, department_cache):
            skipped.append(act)
        else:
            kept.append(act)
    return kept, skipped


def user_name_map(api: Bitrix24API, user_ids: List[int]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for uid in sorted(set(user_ids)):
        try:
            arr = api.call("user.get", {"ID": int(uid)}).get("result") or []
            if arr and isinstance(arr[0], dict):
                u = arr[0]
                out[int(uid)] = f"{u.get('NAME', '')} {u.get('LAST_NAME', '')}".strip() or str(uid)
        except Exception:
            out[int(uid)] = str(uid)
    return out


def fetch_timeline_comments(api: Bitrix24API, deal_id: str) -> List[str]:
    try:
        res = api.call("crm.timeline.comment.list", {"filter": {"ENTITY_TYPE": "deal", "ENTITY_ID": int(deal_id)}})
        rows = res.get("result", []) or []
        comments: List[str] = []
        for r in rows:
            txt = str((r or {}).get("COMMENT") or "").strip()
            if txt:
                comments.append(txt)
        return comments
    except Exception:
        return []


def parse_dt(raw: Any) -> Optional[datetime]:
    try:
        if not raw:
            return None
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except Exception:
        return None


def guess_duration_sec(act: Dict[str, Any]) -> int:
    try:
        st = act.get("START_TIME")
        en = act.get("END_TIME")
        if st and en:
            dt1 = datetime.fromisoformat(str(st).replace("Z", "+00:00"))
            dt2 = datetime.fromisoformat(str(en).replace("Z", "+00:00"))
            sec = int(max(1.0, (dt2 - dt1).total_seconds()))
            return sec
    except Exception:
        pass
    return 60


def guess_duration_minutes(act: Dict[str, Any]) -> float:
    return round(guess_duration_sec(act) / 60.0, 2)


def reaction_speed_label(first_delay_min: Optional[float]) -> str:
    if first_delay_min is None:
        return "Нет звонка менеджера"
    if first_delay_min <= 15:
        return "Быстрая реакция"
    if first_delay_min <= 30:
        return "В срок"
    if first_delay_min <= 60:
        return "Поздно"
    return "Критически поздно"


def compute_discipline_metrics(deal: Dict[str, Any], calls: List[Dict[str, Any]], kpi: Dict[str, Any]) -> Dict[str, Any]:
    sla_cfg = kpi.get("sla", {})
    first_response_sla = float(sla_cfg.get("first_response_hours", FIRST_RESPONSE_SLA_HOURS))
    created = parse_dt(deal.get("DATE_CREATE"))
    call_times = [parse_dt(c.get("START_TIME")) for c in calls]
    call_times = [t for t in call_times if t is not None]
    call_times.sort()

    first_delay_h = None
    first_delay_min = None
    if created and call_times:
        first_delay_h = round(max(0.0, (call_times[0] - created).total_seconds() / 3600.0), 2)
        first_delay_min = round(max(0.0, (call_times[0] - created).total_seconds() / 60.0), 1)

    first_sla_min = round(first_response_sla * 60.0, 1)
    first_ok = first_delay_h is not None and first_delay_h <= first_response_sla

    return {
        "calls_count": len(call_times),
        "first_response_hours": first_delay_h,
        "first_response_minutes": first_delay_min,
        "first_response_sla_minutes": first_sla_min,
        "first_response_sla_ok": first_ok,
        "reaction_speed_label": reaction_speed_label(first_delay_min),
        "first_response_explanation": (
            f"Скорость реакции — сколько минут прошло от создания сделки до первого звонка менеджера. "
            f"Единая норма KPI: до 30 мин.; факт: {first_delay_min:g} мин."
            if first_delay_min is not None
            else "Скорость реакции — время от создания сделки до первого звонка менеджера. Единая норма KPI: до 30 мин.; звонков менеджера нет."
        ),
    }


def compute_deal_quality(deal: Dict[str, Any], comments: List[str], kpi: Dict[str, Any]) -> Dict[str, Any]:
    w = kpi.get("deal_quality_weights", {})
    has_contact = bool(deal.get("CONTACT_ID") or deal.get("COMPANY_ID"))
    has_amount = bool(str(deal.get("OPPORTUNITY") or "").strip() not in {"", "0", "0.00"})
    has_title = bool(str(deal.get("TITLE") or "").strip())
    has_next_step = bool(comments)
    score = 0
    score += int(w.get("has_contact", 25)) if has_contact else 0
    score += int(w.get("has_amount", 25)) if has_amount else 0
    score += int(w.get("has_title", 25)) if has_title else 0
    score += int(w.get("has_comments", 25)) if has_next_step else 0
    details = (
        f"Контакт/компания: {'да' if has_contact else 'нет'} ({int(w.get('has_contact', 25))} баллов); "
        f"сумма: {'да' if has_amount else 'нет'} ({int(w.get('has_amount', 25))} баллов); "
        f"название сделки: {'да' if has_title else 'нет'} ({int(w.get('has_title', 25))} баллов); "
        f"комментарии или следующий шаг в CRM: {'да' if has_next_step else 'нет'} ({int(w.get('has_comments', 25))} баллов)."
    )
    return {
        "deal_quality_score": score,
        "deal_quality_details": details,
        "has_contact": has_contact,
        "has_amount": has_amount,
        "has_title": has_title,
        "has_comments": has_next_step,
    }


def _crm_item(code: str, block: str, criterion: str, score: float, comment: str) -> Dict[str, Any]:
    score = max(0.0, min(1.0, float(score or 0.0)))
    return {
        "crm_checklist_block_name": block,
        "crm_checklist_code": code,
        "crm_checklist_criterion": criterion,
        "crm_checklist_score": score,
        "crm_checklist_max_score": 1,
        "crm_checklist_comment": comment,
    }


def _status_score(status: Any) -> Tuple[float, str]:
    text = str(status or "").strip()
    lowered = text.lower()
    if not text:
        return 0.5, "Нет данных о статусе срока, критерий учтен как частично выполненный."
    if "тревога" in lowered:
        return 0.0, f"Статус: {text}. Нужна управленческая реакция по сделке."
    if "предупреждение" in lowered:
        return 0.5, f"Статус: {text}. Сделка приближается к критическому сроку."
    return 1.0, f"Статус: {text}."


def evaluate_crm_checklist(row: Dict[str, Any], suffix: str = "", include_stage: bool = True) -> Dict[str, Any]:
    def get(key: str) -> Any:
        if suffix and f"{key}{suffix}" in row:
            return row.get(f"{key}{suffix}")
        return row.get(key)

    items: List[Dict[str, Any]] = []
    items.append(
        _crm_item(
            "crm_has_contact",
            "Заполнение сделки",
            "В сделке указан контакт или компания",
            1.0 if get("has_contact") else 0.0,
            "Контакт/компания есть." if get("has_contact") else "В сделке нет контакта или компании.",
        )
    )
    items.append(
        _crm_item(
            "crm_has_amount",
            "Заполнение сделки",
            "В сделке указана сумма",
            1.0 if get("has_amount") else 0.0,
            "Сумма заполнена." if get("has_amount") else "Сумма сделки не заполнена или равна нулю.",
        )
    )
    items.append(
        _crm_item(
            "crm_has_title",
            "Заполнение сделки",
            "Название сделки заполнено",
            1.0 if get("has_title") else 0.0,
            "Название заполнено." if get("has_title") else "Название сделки пустое.",
        )
    )
    items.append(
        _crm_item(
            "crm_has_comments",
            "Заполнение сделки",
            "В CRM есть комментарий или следующий шаг",
            1.0 if get("has_comments") else 0.0,
            "В таймлайне есть комментарии/следующие действия." if get("has_comments") else "В CRM не найден комментарий или следующий шаг.",
        )
    )

    has_call = not bool(row.get("no_calls")) and int(row.get("calls_count") or 0) > 0
    items.append(
        _crm_item(
            "crm_has_manager_call",
            "Активность по сделке",
            "Есть звонок менеджера после исключения Call-центра",
            1.0 if has_call else 0.0,
            "Звонок менеджера найден." if has_call else "Звонков менеджера не найдено.",
        )
    )

    next_step_synced = bool(get("next_step_synced"))
    next_step_in_call = bool(get("has_next_step_phrase"))
    items.append(
        _crm_item(
            "crm_next_step_synced",
            "Связь звонка с CRM",
            "Следующий шаг из разговора синхронизирован с CRM",
            1.0 if next_step_synced else (0.5 if next_step_in_call else 0.0),
            (
                "Следующий шаг прозвучал и зафиксирован в CRM."
                if next_step_synced
                else (
                    "Следующий шаг прозвучал, но не подтвержден в CRM."
                    if next_step_in_call
                    else "Следующий шаг не найден ни в разговоре, ни в CRM."
                )
            ),
        )
    )

    amount_mentioned = bool(get("amount_mentioned"))
    has_amount = bool(get("has_amount"))
    items.append(
        _crm_item(
            "crm_amount_aligned",
            "Связь звонка с CRM",
            "Сумма сделки подтверждается разговором",
            1.0 if amount_mentioned else (0.5 if has_amount else 0.0),
            (
                "Сумма из сделки встречается в разговоре."
                if amount_mentioned
                else (
                    "Сумма есть в CRM, но в разговоре явно не подтверждена."
                    if has_amount
                    else "Сумма не заполнена и не подтверждена разговором."
                )
            ),
        )
    )

    if include_stage:
        history_count = row.get("stage_history_count")
        if history_count is not None:
            history_ok = int(history_count or 0) > 0
            items.append(
                _crm_item(
                    "crm_stage_history",
                    "Движение по воронке",
                    "Есть история движения сделки по стадиям",
                    1.0 if history_ok else 0.0,
                    "История стадий найдена." if history_ok else "История движения по стадиям не найдена.",
                )
            )
            score, comment = _status_score(row.get("stage_current_age_status"))
            items.append(
                _crm_item(
                    "crm_stage_age_ok",
                    "Движение по воронке",
                    "Сделка не зависла на текущей стадии",
                    score,
                    comment,
                )
            )
            score, comment = _status_score(row.get("deal_total_work_status"))
            items.append(
                _crm_item(
                    "crm_total_work_ok",
                    "Движение по воронке",
                    "Общий срок сделки в работе без тревоги",
                    score,
                    comment,
                )
            )
            returns = int(row.get("stage_return_count") or 0)
            items.append(
                _crm_item(
                    "crm_no_stage_returns",
                    "Движение по воронке",
                    "Нет возвратов на предыдущие стадии",
                    1.0 if returns == 0 else (0.5 if returns == 1 else 0.0),
                    f"Возвратов на предыдущие стадии: {returns}.",
                )
            )
            risk = str(row.get("stage_movement_risk") or "").strip()
            risk_l = risk.lower()
            if not risk:
                risk_score = 0.5
                risk_comment = "Риск движения по воронке не рассчитан."
            elif risk in {"OK", "Финал"}:
                risk_score = 1.0
                risk_comment = f"Риск движения: {risk}."
            elif "тревога" in risk_l:
                risk_score = 0.0
                risk_comment = f"Риск движения: {risk}. Нужна корректировка работы со сделкой."
            else:
                risk_score = 0.5
                risk_comment = f"Риск движения: {risk}. Требуется проверка руководителем."
            items.append(
                _crm_item(
                    "crm_stage_movement_ok",
                    "Движение по воронке",
                    "Движение сделки по воронке без критических рисков",
                    risk_score,
                    risk_comment,
                )
            )

    total_score = round(sum(float(item.get("crm_checklist_score") or 0.0) for item in items), 2)
    total_max = len(items)
    percent = round(total_score * 100.0 / max(1, total_max), 2)
    blocks: Dict[str, List[Dict[str, Any]]] = {}
    for item in items:
        blocks.setdefault(str(item.get("crm_checklist_block_name") or ""), []).append(item)
    details = "; ".join(
        f"{block}: {round(sum(float(x.get('crm_checklist_score') or 0.0) for x in block_items), 2)}/{len(block_items)}"
        for block, block_items in blocks.items()
    )
    return {
        "crm_checklist_total_score": total_score,
        "crm_checklist_max_score": total_max,
        "crm_checklist_percent": percent,
        "crm_checklist_details": details,
        "crm_checklist_items": items,
        "crm_work_score": percent,
    }


def _match_any(text: str, patterns: List[str]) -> bool:
    for p in patterns:
        try:
            if re.search(p, text):
                return True
        except re.error:
            continue
    return False



def crm_call_alignment(deal: Dict[str, Any], text: str, comments: List[str], kpi: Dict[str, Any]) -> Dict[str, Any]:
    w = kpi.get("alignment_weights", {})
    t = (text or "").lower()
    title_words = [w for w in re.findall(r"[a-zA-Zа-яА-Я0-9]{4,}", str(deal.get("TITLE") or "").lower())[:6]]
    title_hits = sum(1 for w in title_words if w in t)
    amount = str(deal.get("OPPORTUNITY") or "").split(".")[0]
    amount_mentioned = bool(amount and amount != "0" and amount in t)
    comments_text = " ".join(comments).lower()
    next_step_re = r"\b(перезвон\w*|встреч\w*|созвон\w*|отправ\w*|вышл\w*|уточн\w*|согласу\w*|кп|договор\w*)\b"
    next_step_synced = bool(re.search(next_step_re, t) and re.search(next_step_re, comments_text))
    amount_weight = int(w.get("amount_mentioned", 30))
    next_step_weight = int(w.get("next_step_synced", 40))
    raw_score = (amount_weight if amount_mentioned else 0) + (next_step_weight if next_step_synced else 0)
    total_weight = max(1, amount_weight + next_step_weight)
    align_score = round(raw_score * 100.0 / total_weight, 2)
    next_step_details = (
        "Да: следующий шаг звучит в разговоре и зафиксирован в комментариях/таймлайне CRM."
        if next_step_synced
        else "Нет: следующий шаг либо не прозвучал в разговоре, либо не зафиксирован в CRM. Это мешает контролировать дальнейшее действие по клиенту."
    )
    details = (
        "Связь звонка с CRM показывает, совпадает ли содержание разговора с данными сделки: "
        f"сумма упомянута: {'да' if amount_mentioned else 'нет'}; "
        f"следующий шаг синхронизирован с CRM: {'да' if next_step_synced else 'нет'}. "
        "Совпадения с названием сделки больше не выводятся в отчет как отдельная метрика."
    )
    return {
        "alignment_score": align_score,
        "alignment_details": details,
        "title_mentions": title_hits,
        "amount_mentioned": amount_mentioned,
        "next_step_synced": next_step_synced,
        "next_step_synced_details": next_step_details,
    }


def quality_label(score: float) -> str:
    if score >= 80:
        return "сильное"
    if score >= 60:
        return "нормальное"
    if score >= 40:
        return "слабое"
    return "критически слабое"


def call_quality_conclusion(row: Dict[str, Any]) -> Tuple[str, str]:
    score = float(row.get("call_quality_score") or 0.0)
    missing: List[str] = []
    blocks = row.get("call_checklist_blocks") or []
    if isinstance(blocks, list) and blocks:
        weak_blocks = [
            str(block.get("sales_stage_block_name") or "")
            for block in blocks
            if isinstance(block, dict) and float(block.get("sales_stage_percent") or 0.0) < 70
        ]
        missing.extend([x for x in weak_blocks if x])
    else:
        if not row.get("has_greeting"):
            missing.append("приветствие")
        if not row.get("has_needs_discovery"):
            missing.append("выявление потребностей")
        if int(row.get("objections_count") or 0) > 0 and not row.get("has_objection_work"):
            missing.append("работа с возражениями")
        if not row.get("has_next_step_phrase"):
            missing.append("фиксированный следующий шаг")

    if score >= 80:
        conclusion = "Разговор структурный: есть основные элементы продажного диалога."
    elif score >= 60:
        conclusion = "Разговор рабочий, но часть обязательных элементов требует усиления."
    elif score >= 40:
        conclusion = "Разговор частично проработан: клиенту не хватает ясной структуры и завершения."
    else:
        conclusion = "Разговор слабый: ключевые элементы проработки клиента почти не зафиксированы."

    recs = []
    if missing:
        recs.append("Усилить: " + ", ".join(missing) + ".")
    if int(row.get("unhandled_objections_count") or 0) > 0:
        recs.append("Вернуться к неотработанным возражениям и дать клиенту альтернативу/ценность/следующий шаг.")
    if not row.get("next_step_synced"):
        recs.append("Зафиксировать следующий шаг в разговоре и в CRM.")
    objection_recs = str(row.get("objection_recommendations") or "").strip()
    if objection_recs:
        recs.append(objection_recs)
    if not recs:
        recs.append("Поддерживать текущую структуру разговора.")
    return conclusion, " ".join(recs)



def analyze_transcript_improvements(text: str, row: Dict[str, Any]) -> Dict[str, Any]:
    raw = text or ""
    lower = raw.lower()
    objections: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str]] = set()

    for label, pattern, suggestion in OBJECTION_RULES:
        for match in re.finditer(pattern, lower, flags=re.IGNORECASE):
            fragment = _context(raw, match.start(), match.end())
            key = (label, fragment[:160])
            if key in seen:
                continue
            seen.add(key)
            after = lower[match.end() : match.end() + 420]
            handled = bool(HANDLING_RE.search(after))
            objections.append(
                {
                    "objection_type": label,
                    "objection_fragment": fragment,
                    "objection_status": "отработано" if handled else "не отработано",
                    "objection_recommendation": suggestion,
                    "handled": handled,
                }
            )

    improvement_items: List[str] = []
    if not row.get("has_greeting"):
        improvement_items.append("Нет явного приветствия и представления.")
    if not row.get("has_needs_discovery"):
        improvement_items.append("Слабо выявлены потребности: не видно уточняющих вопросов о задаче клиента.")
    if not row.get("has_next_step_phrase"):
        improvement_items.append("Не зафиксирован конкретный следующий шаг.")

    unhandled = [o for o in objections if not o["handled"]]
    if unhandled:
        improvement_items.append("Есть неотработанные возражения: " + "; ".join(o["objection_type"] for o in unhandled[:5]) + ".")
    elif objections:
        improvement_items.append("Возражения найдены и в целом отработаны, но стоит фиксировать следующий шаг после ответа.")

    marked_lines: List[str] = []
    for objection in objections:
        status = objection["objection_status"]
        marker = "ТРЕБУЕТ ОТРАБОТКИ" if status == "не отработано" else "ОТРАБОТАНО"
        marked_lines.append(
            f"[ВОЗРАЖЕНИЕ: {objection['objection_type']} | {marker}]\n"
            f"{objection['objection_fragment']}\n"
            f"[ВАРИАНТ ОТРАБОТКИ] {objection['objection_recommendation']}"
        )

    if not marked_lines and improvement_items:
        marked_lines.append("[МОМЕНТЫ ДЛЯ УЛУЧШЕНИЯ]\n" + "\n".join(f"- {x}" for x in improvement_items))

    transcript_marked = raw
    if marked_lines:
        transcript_marked = "\n\n".join(marked_lines) + "\n\n[ПОЛНАЯ РАСШИФРОВКА]\n" + raw

    recommendations = [o["objection_recommendation"] for o in unhandled]
    if not recommendations and objections:
        recommendations = ["После ответа на возражение обязательно закрепить следующий шаг: дата, действие, ответственный."]

    return {
        "objections_count": len(objections),
        "unhandled_objections_count": len(unhandled),
        "objections_handled": bool(objections and not unhandled),
        "objections_found": "\n".join(
            f"{o['objection_type']}: {o['objection_fragment']} ({o['objection_status']})" for o in objections
        ),
        "unhandled_objections": "\n".join(f"{o['objection_type']}: {o['objection_fragment']}" for o in unhandled),
        "objection_recommendations": "\n".join(dict.fromkeys(recommendations)),
        "improvement_moments": "\n".join(improvement_items),
        "transcript_marked": transcript_marked,
        "objection_rows": objections,
    }


def save_transcript_file(deal_id: str, activity_id: Any, task_id: str, text: str) -> Path:
    TRANSCRIPTS_DIR.mkdir(parents=True, exist_ok=True)
    safe_task = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(task_id or ""))[:36] or "no_task"
    safe_deal = re.sub(r"\D+", "", str(deal_id or "")) or "unknown_deal"
    safe_activity = re.sub(r"\D+", "", str(activity_id or "")) or "unknown_activity"
    path = TRANSCRIPTS_DIR / f"deal{safe_deal}_activity{safe_activity}_{safe_task}.txt"
    path.write_text(text or "", encoding="utf-8")
    return path


def find_latest_transcript_file(deal_id: Any, activity_id: Any) -> Optional[Path]:
    safe_deal = re.sub(r"\D+", "", str(deal_id or ""))
    safe_activity = re.sub(r"\D+", "", str(activity_id or ""))
    if not safe_deal or not safe_activity or not TRANSCRIPTS_DIR.exists():
        return None
    pattern = f"deal{safe_deal}_activity{safe_activity}_*.txt"
    files = sorted(TRANSCRIPTS_DIR.glob(pattern), key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
    return files[0] if files else None


def load_cached_transcript(
    state_cache: Dict[str, Any],
    call_id: str,
    deal_id: Any,
    activity_id: Any,
) -> Tuple[Optional[str], Optional[Path]]:
    candidates: List[Path] = []
    cached = state_cache.get(call_id)
    if isinstance(cached, dict):
        cached_path = cached.get("transcript_path")
        if cached_path:
            candidates.append(Path(str(cached_path)))
    latest = find_latest_transcript_file(deal_id, activity_id)
    if latest:
        candidates.append(latest)

    seen: set[str] = set()
    for path in candidates:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        try:
            if path.exists() and path.stat().st_size > 0:
                text = path.read_text(encoding="utf-8", errors="ignore").strip()
                if text:
                    return text, path
        except Exception:
            continue
    return None, None


def finalize_transcript_analysis(
    row: Dict[str, Any],
    deal: Dict[str, Any],
    comments: List[Dict[str, Any]],
    bitnewton_text: str,
    bitrix_text: str,
    kpi: Dict[str, Any],
    kpi_cmp: Optional[Dict[str, Any]],
) -> None:
    analysis_text = merged_transcript_text(bitnewton_text or "", bitrix_text or "")
    row["combined_transcript_text"] = analysis_text
    apply_scores(row, deal, comments, analysis_text, kpi, suffix="")
    if kpi_cmp is not None:
        apply_scores(row, deal, comments, analysis_text, kpi_cmp, suffix="_cmp")
        row["overall_score_delta"] = round(float(row.get("overall_score_cmp") or 0) - float(row.get("overall_score") or 0), 2)
    row.update(analyze_transcript_improvements(analysis_text, row))
    call_conclusion, call_recommendations = call_quality_conclusion(row)
    row["call_quality_conclusion"] = call_conclusion
    row["recommendations"] = call_recommendations
    row["conversation_meaning"] = conversation_meaning(analysis_text, row)


def build_deal_conclusions(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("deal_id") or "unknown"), []).append(row)

    out: List[Dict[str, Any]] = []
    for deal_id, deal_rows in grouped.items():
        ok_rows = [r for r in deal_rows if not r.get("error")]
        scored = ok_rows or deal_rows
        calls_total = len(deal_rows)
        calls_ok = len(ok_rows)
        avg_overall = round(sum(float(r.get("overall_score") or 0.0) for r in scored) / max(1, len(scored)), 2)
        avg_call = round(sum(float(r.get("call_quality_score") or 0.0) for r in scored) / max(1, len(scored)), 2)
        deal_quality = max(float(r.get("deal_quality_score") or 0.0) for r in scored) if scored else 0.0
        needs_ratio = sum(1 for r in scored if r.get("has_needs_discovery")) / max(1, len(scored))
        next_step_ratio = sum(1 for r in scored if r.get("has_next_step_phrase")) / max(1, len(scored))
        objection_ratio = sum(1 for r in scored if r.get("has_objection_work")) / max(1, len(scored))
        synced_ratio = sum(1 for r in scored if r.get("next_step_synced")) / max(1, len(scored))
        avg_crm_work = round(sum(float(r.get("crm_work_score") or 0.0) for r in scored) / max(1, len(scored)), 2)

        client_work_score = round(
            avg_crm_work if avg_crm_work else (
                0.60 * deal_quality
                + 0.40 * synced_ratio * 100
            ),
            2,
        )

        issues: List[str] = []
        if deal_quality < 100:
            issues.append("не полностью заполнена сделка")
        if needs_ratio < 0.5:
            issues.append("слабо выявлены потребности")
        if next_step_ratio < 0.5:
            issues.append("не фиксируется следующий шаг")
        if synced_ratio < 0.5:
            issues.append("следующий шаг плохо синхронизирован с CRM")
        if objection_ratio == 0 and calls_total > 0:
            issues.append("работа с возражениями не прослеживается")

        first = deal_rows[0]
        if client_work_score >= 80:
            client_conclusion = "Клиент проработан качественно: CRM и разговоры дают понятную картину дальнейших действий."
        elif client_work_score >= 60:
            client_conclusion = "Проработка клиента нормальная, но есть зоны для усиления."
        elif client_work_score >= 40:
            client_conclusion = "Проработка клиента слабая: часть важных элементов не закреплена в разговоре или CRM."
        else:
            client_conclusion = "Проработка клиента критически слабая: не хватает структуры, фиксации потребностей и следующего шага."

        if avg_call >= 80:
            conversation_conclusion = "Качество разговоров сильное."
        elif avg_call >= 60:
            conversation_conclusion = "Качество разговоров нормальное, но нестабильное."
        elif avg_call >= 40:
            conversation_conclusion = "Качество разговоров слабое: менеджеру нужна более четкая структура диалога."
        else:
            conversation_conclusion = "Качество разговоров критически слабое."

        recommendations = "Усилить: " + ", ".join(issues) + "." if issues else "Сохранять текущий подход и фиксировать следующий шаг после каждого контакта."
        transcript_paths = [str(r.get("transcript_path")) for r in deal_rows if r.get("transcript_path")]

        out.append(
            {
                "deal_url": first.get("deal_url"),
                "stage_name": first.get("stage_name") or stage_display_name(first.get("stage_id")),
                "manager_name": first.get("manager_name"),
                "calls_total": calls_total,
                "calls_ok": calls_ok,
                "transcripts_count": len(transcript_paths),
                "avg_overall_score": avg_overall,
                "avg_call_quality_score": avg_call,
                "client_work_score": client_work_score,
                "client_work_quality": quality_label(client_work_score),
                "conversation_quality": quality_label(avg_call),
                "client_work_conclusion": client_conclusion,
                "call_quality_conclusion": conversation_conclusion,
                "recommendations": recommendations,
                "transcript_paths": "\n".join(transcript_paths),
            }
        )

    return sorted(out, key=lambda x: float(x.get("client_work_score") or 0.0))


def _deal_group_key(row: Dict[str, Any]) -> str:
    return str(row.get("deal_id") or row.get("deal_url") or "unknown")


def _join_unique(values: List[Any], sep: str = "\n") -> str:
    seen: set[str] = set()
    out: List[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return sep.join(out)


def _call_sort_key(row: Dict[str, Any]) -> Tuple[str, str]:
    return (str(row.get("start_time") or ""), str(row.get("activity_id") or ""))


def _call_heading(row: Dict[str, Any], number: int) -> str:
    subject = str(row.get("subject") or "Звонок").strip()
    duration = row.get("duration_minutes")
    duration_text = f"{duration:g} мин." if isinstance(duration, (int, float)) else ""
    if duration_text:
        return f"Звонок {number}: {subject} ({duration_text})"
    return f"Звонок {number}: {subject}"


def _short_error(text: Any, limit: int = 550) -> str:
    raw = _clean_text_for_report(str(text or ""))
    if not raw:
        return ""
    if "не дождался входа в Bitrix" in raw or "oauth/authorize" in raw:
        return "Не удалось открыть Bitrix через выбранный профиль браузера: требуется вход в Bitrix или выбран не тот профиль Edge/Chrome."
    if "UI download timeout" in raw:
        return "Не удалось скачать запись через интерфейс Bitrix: файл не появился в папке загрузок."
    if "download failed" in raw:
        return "Не удалось скачать запись звонка через REST/UI."
    return raw[:limit] + ("..." if len(raw) > limit else "")


def transcript_match_score(bitnewton_text: str, bitrix_text: str) -> Optional[float]:
    a = set(re.findall(r"[a-zA-Zа-яА-Я0-9]{4,}", (bitnewton_text or "").lower()))
    b = set(re.findall(r"[a-zA-Zа-яА-Я0-9]{4,}", (bitrix_text or "").lower()))
    if not a or not b:
        return None
    return round((len(a & b) / max(1, len(a | b))) * 100.0, 1)


def merged_transcript_text(bitnewton_text: str, bitrix_text: str) -> str:
    bn = (bitnewton_text or "").strip()
    bx = (bitrix_text or "").strip()
    if not bx:
        return bn
    score = transcript_match_score(bn, bx)
    if not bn:
        return bx
    if score is not None and score >= 85:
        return bn
    return f"[Bit.Newton]\n{bn}\n\n[Bitrix карточка]\n{bx}"


def conversation_meaning(text: str, row: Dict[str, Any]) -> str:
    t = (text or "").lower()
    topics: List[str] = []
    topic_rules = [
        ("цена/стоимость", r"\b(цен\w*|стоимост\w*|дорог\w*|бюджет)\b"),
        ("счет/оплата", r"\b(счет|оплат\w*|предоплат\w*|платеж)\b"),
        ("КП/договор", r"\b(кп|коммерческ\w*|договор\w*)\b"),
        ("технические условия", r"\b(тех\w*|интеграц\w*|настройк\w*|касс\w*|оборудован\w*)\b"),
        ("сроки/следующий контакт", r"\b(срок\w*|перезвон\w*|созвон\w*|встреч\w*|завтра|сегодня)\b"),
    ]
    for label, pattern in topic_rules:
        if re.search(pattern, t):
            topics.append(label)
    if not topics:
        topics.append("общая консультация по сделке")

    parts = [f"Смысл разговора: обсуждались {', '.join(dict.fromkeys(topics))}."]
    if row.get("has_needs_discovery"):
        parts.append("Потребности клиента в разговоре частично выявлены.")
    else:
        parts.append("Потребности клиента выражены слабо: нужно больше уточняющих вопросов.")
    if row.get("has_objection_work"):
        parts.append("Возражения или сомнения клиента в разговоре затронуты.")
    if row.get("has_next_step_phrase"):
        parts.append("Следующий шаг в разговоре прозвучал.")
    else:
        parts.append("Следующий шаг не закреплен явно.")
    if row.get("next_step_synced"):
        parts.append("Следующий шаг совпадает с фиксацией в CRM.")
    else:
        parts.append("Следующий шаг нужно отдельно зафиксировать в CRM.")
    unhandled = str(row.get("unhandled_objections") or "").strip()
    if unhandled:
        parts.append("Есть неотработанные возражения: " + _clean_text_for_report(unhandled)[:220])
    return " ".join(parts)


def build_deal_report_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        first = deal_rows[0]
        no_call_rows = [r for r in deal_rows if r.get("no_calls")]
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        ok_calls = [r for r in call_rows if not r.get("error")]
        failed_calls = [r for r in call_rows if r.get("error")]
        scored = ok_calls

        calls_total = len(call_rows)
        calls_ok = len(ok_calls)
        calls_failed = len(failed_calls)
        deal_quality = max(float(r.get("deal_quality_score") or 0.0) for r in deal_rows) if deal_rows else 0.0
        if scored:
            avg_overall = round(sum(float(r.get("overall_score") or 0.0) for r in scored) / len(scored), 2)
            avg_call = round(sum(float(r.get("call_quality_score") or 0.0) for r in scored) / len(scored), 2)
            avg_checklist = round(sum(float(r.get("call_checklist_percent") or r.get("call_quality_score") or 0.0) for r in scored) / len(scored), 2)
            avg_alignment = round(sum(float(r.get("alignment_score") or 0.0) for r in scored) / len(scored), 2)
            avg_crm_work = round(sum(float(r.get("crm_work_score") or 0.0) for r in scored) / len(scored), 2)
            avg_crm_checklist = round(sum(float(r.get("crm_checklist_percent") or r.get("crm_work_score") or 0.0) for r in scored) / len(scored), 2)
            needs_ratio = sum(1 for r in scored if r.get("has_needs_discovery")) / len(scored)
            next_step_ratio = sum(1 for r in scored if r.get("has_next_step_phrase")) / len(scored)
            objection_ratio = sum(1 for r in scored if r.get("has_objection_work")) / len(scored)
            synced_ratio = sum(1 for r in scored if r.get("next_step_synced")) / len(scored)
        else:
            avg_overall = 0.0
            avg_call = 0.0
            avg_checklist = 0.0
            avg_alignment = 0.0
            avg_crm_work = 0.0
            avg_crm_checklist = 0.0
            needs_ratio = 0.0
            next_step_ratio = 0.0
            objection_ratio = 0.0
            synced_ratio = 0.0

        client_work_score = round(
            avg_crm_work if avg_crm_work else (
                0.60 * deal_quality
                + 0.40 * synced_ratio * 100
            ),
            2,
        ) if scored else 0.0

        issues: List[str] = []
        if no_call_rows or calls_total == 0:
            issues.append("по сделке не найдено звонков")
        if calls_failed:
            issues.append(f"не обработано звонков: {calls_failed}")
        if deal_quality < 100:
            issues.append("не полностью заполнена сделка")
        if scored and needs_ratio < 0.5:
            issues.append("слабо выявлены потребности")
        if scored and next_step_ratio < 0.5:
            issues.append("не фиксируется следующий шаг")
        if scored and synced_ratio < 0.5:
            issues.append("следующий шаг плохо синхронизирован с CRM")
        if scored and objection_ratio == 0:
            issues.append("работа с возражениями не прослеживается")

        if not scored:
            client_conclusion = "По сделке нет обработанных звонков: оценка клиента невозможна до появления записи или успешной расшифровки."
            conversation_conclusion = "Нет обработанных разговоров для оценки."
        elif client_work_score >= 80:
            client_conclusion = "Клиент проработан качественно: CRM и разговоры дают понятную картину дальнейших действий."
            conversation_conclusion = "Качество разговоров сильное." if avg_call >= 80 else "Разговоры в целом рабочие, но есть отдельные зоны для усиления."
        elif client_work_score >= 60:
            client_conclusion = "Проработка клиента нормальная, но есть зоны для усиления."
            conversation_conclusion = "Качество разговоров нормальное, но нестабильное." if avg_call >= 60 else "Качество разговоров слабое: менеджеру нужна более четкая структура диалога."
        elif client_work_score >= 40:
            client_conclusion = "Проработка клиента слабая: часть важных элементов не закреплена в разговоре или CRM."
            conversation_conclusion = "Качество разговоров слабое: менеджеру нужна более четкая структура диалога."
        else:
            client_conclusion = "Проработка клиента критически слабая: не хватает структуры, фиксации потребностей и следующего шага."
            conversation_conclusion = "Качество разговоров критически слабое."

        call_blocks: List[str] = []
        transcript_blocks: List[str] = []
        for index, row in enumerate(call_rows, start=1):
            heading = _call_heading(row, index)
            if row.get("error"):
                call_blocks.append(f"{heading}\nСтатус: ошибка обработки — {row.get('error')}")
                continue
            call_blocks.append(
                "\n".join(
                    [
                        heading,
                        f"Качество разговора: {row.get('call_quality_score') or 0}",
                        f"Потребности: {'да' if row.get('has_needs_discovery') else 'нет'}; "
                        f"возражения: {'да' if row.get('has_objection_work') else 'нет'}; "
                        f"следующий шаг: {'да' if row.get('has_next_step_phrase') else 'нет'}",
                        str(row.get("improvement_moments") or "").strip(),
                    ]
                ).strip()
            )
            transcript = str(row.get("transcript_marked") or row.get("transcript_text") or "").strip()
            if transcript:
                transcript_blocks.append(f"{heading}\n{transcript}")

        if no_call_rows and not call_blocks:
            call_blocks.append("Звонков по сделке не найдено.")

        out.append(
            {
                "deal_url": first.get("deal_url"),
                "stage_name": first.get("stage_name") or stage_display_name(first.get("stage_id")),
                "manager_name": first.get("manager_name"),
                "kpi_profile_ru": first.get("kpi_profile_ru"),
                "kpi_explanation": first.get("kpi_explanation"),
                "calls_total": calls_total,
                "calls_ok": calls_ok,
                "calls_failed": calls_failed,
                "ignored_call_center_calls": int(first.get("ignored_call_center_calls") or 0),
                "no_calls": bool(no_call_rows or calls_total == 0),
                "transcripts_count": len([r for r in ok_calls if r.get("transcript_path")]),
                "avg_overall_score": avg_overall,
                "overall_score_details": _join_unique([r.get("overall_score_details") for r in ok_calls]) or first.get("overall_score_details"),
                "avg_call_quality_score": avg_call,
                "call_quality_details": _join_unique([r.get("call_quality_details") for r in ok_calls]),
                "call_checklist_percent": avg_checklist,
                "call_checklist_block_details": _join_unique([r.get("call_checklist_block_details") for r in ok_calls]),
                "deal_quality_score": deal_quality,
                "deal_quality_details": first.get("deal_quality_details"),
                "alignment_score": avg_alignment,
                "crm_work_score": client_work_score,
                "crm_checklist_percent": avg_crm_checklist,
                "crm_checklist_details": _join_unique([r.get("crm_checklist_details") for r in ok_calls]),
                "alignment_details": _join_unique([r.get("alignment_details") for r in ok_calls]),
                "stage_history_count": first.get("stage_history_count"),
                "stage_history_path": first.get("stage_history_path"),
                "stage_last_change_at": first.get("stage_last_change_at"),
                "stage_current_age_minutes": first.get("stage_current_age_minutes"),
                "stage_current_age_days": first.get("stage_current_age_days"),
                "stage_warning_threshold": first.get("stage_warning_threshold"),
                "stage_critical_threshold": first.get("stage_critical_threshold"),
                "stage_current_age_status": first.get("stage_current_age_status"),
                "deal_total_work_minutes": first.get("deal_total_work_minutes"),
                "deal_total_work_days": first.get("deal_total_work_days"),
                "deal_total_work_warning_threshold": first.get("deal_total_work_warning_threshold"),
                "deal_total_work_critical_threshold": first.get("deal_total_work_critical_threshold"),
                "deal_total_work_status": first.get("deal_total_work_status"),
                "stage_return_count": first.get("stage_return_count"),
                "stage_reached_final": first.get("stage_reached_final"),
                "stage_movement_risk": first.get("stage_movement_risk"),
                "stage_movement_recommendation": first.get("stage_movement_recommendation"),
                "client_work_score": client_work_score,
                "client_work_quality": quality_label(client_work_score) if scored else "Нет данных",
                "conversation_quality": quality_label(avg_call) if scored else "Нет данных",
                "client_work_conclusion": client_conclusion,
                "call_quality_conclusion": conversation_conclusion,
                "recommendations": "Усилить: " + ", ".join(issues) + "." if issues else "Сохранять текущий подход и фиксировать следующий шаг после каждого контакта.",
                "calls_breakdown": "\n\n".join(call_blocks),
                "transcripts_combined": "\n\n---\n\n".join(transcript_blocks),
                "conversation_meaning": _join_unique([r.get("conversation_meaning") for r in ok_calls], sep="\n\n"),
                "improvement_moments_combined": _join_unique([r.get("improvement_moments") for r in ok_calls]),
                "objections_combined": _join_unique([r.get("objections_found") for r in ok_calls]),
                "unhandled_objections": _join_unique([r.get("unhandled_objections") for r in ok_calls]),
                "objection_recommendations": _join_unique([r.get("objection_recommendations") for r in ok_calls]),
                "transcript_paths": _join_unique([r.get("transcript_path") for r in ok_calls]),
                "error": _short_error(_join_unique([r.get("error") for r in deal_rows if r.get("error")])),
            }
        )

    return sorted(out, key=lambda x: (str(x.get("manager_name") or ""), str(x.get("deal_url") or "")))


def build_call_detail_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        no_call_rows = [r for r in deal_rows if r.get("no_calls")]
        first = deal_rows[0]
        if not call_rows and no_call_rows:
            out.append(
                {
                    "deal_url": first.get("deal_url"),
                    "call_number": "",
                    "subject": "Звонков не найдено",
                    "duration_minutes": "",
                    "call_has_error": True,
                    "error": _short_error(first.get("error") or "По сделке не найдено звонков"),
                }
            )
            continue
        for index, row in enumerate(call_rows, start=1):
            out.append(
                {
                    "deal_url": row.get("deal_url"),
                    "call_number": index,
                    "subject": row.get("subject"),
                    "duration_minutes": row.get("duration_minutes"),
                    "call_quality_score": row.get("call_quality_score"),
                    "call_quality_details": row.get("call_quality_details"),
                    "call_checklist_total_score": row.get("call_checklist_total_score"),
                    "call_checklist_max_score": row.get("call_checklist_max_score"),
                    "call_checklist_percent": row.get("call_checklist_percent"),
                    "call_checklist_block_details": row.get("call_checklist_block_details"),
                    "has_greeting": row.get("has_greeting"),
                    "has_needs_discovery": row.get("has_needs_discovery"),
                    "has_objection_work": row.get("has_objection_work"),
                    "has_next_step_phrase": row.get("has_next_step_phrase"),
                    "alignment_score": row.get("alignment_score"),
                    "alignment_details": row.get("alignment_details"),
                    "next_step_synced": row.get("next_step_synced"),
                    "next_step_synced_details": row.get("next_step_synced_details"),
                    "call_quality_conclusion": row.get("call_quality_conclusion"),
                    "conversation_meaning": row.get("conversation_meaning"),
                    "improvement_moments": row.get("improvement_moments"),
                    "unhandled_objections": row.get("unhandled_objections"),
                    "objection_recommendations": row.get("objection_recommendations"),
                    "transcript_path": row.get("transcript_path"),
                    "transcript_marked": row.get("transcript_marked"),
                    "transcript_text": row.get("transcript_text"),
                    "bitrix_card_transcript_status": row.get("bitrix_card_transcript_status"),
                    "transcript_match_score": row.get("transcript_match_score"),
                    "bitrix_card_transcript": row.get("bitrix_card_transcript"),
                    "call_has_error": bool(row.get("error")),
                    "error": _short_error(row.get("error")),
                }
            )
    return out


def build_objection_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        for objection in row.get("objection_rows") or []:
            if not isinstance(objection, dict):
                continue
            out.append(
                {
                    "deal_url": row.get("deal_url"),
                    "stage_name": row.get("stage_name") or stage_display_name(row.get("stage_id")),
                    "manager_name": row.get("manager_name"),
                    "subject": row.get("subject"),
                    "objection_fragment": objection.get("objection_fragment"),
                    "objection_status": objection.get("objection_status"),
                    "objection_recommendations": objection.get("objection_recommendation"),
                }
            )
    return out


def build_call_checklist_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        for index, row in enumerate(call_rows, start=1):
            for item in row.get("call_checklist_items") or []:
                if not isinstance(item, dict):
                    continue
                out.append(
                    {
                        "deal_url": row.get("deal_url"),
                        "stage_name": row.get("stage_name") or stage_display_name(row.get("stage_id")),
                        "manager_name": row.get("manager_name"),
                        "call_number": index,
                        "subject": row.get("subject"),
                        "duration_minutes": row.get("duration_minutes"),
                        "checklist_block_name": item.get("checklist_block_name"),
                        "checklist_criterion": item.get("checklist_criterion"),
                        "checklist_score": item.get("checklist_score"),
                        "checklist_max_score": item.get("checklist_max_score"),
                        "checklist_evidence": item.get("checklist_evidence"),
                        "checklist_comment": item.get("checklist_comment"),
                        "checklist_code": item.get("checklist_code"),
                    }
                )
    return out


def build_sales_stage_score_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        for index, row in enumerate(call_rows, start=1):
            for block in row.get("call_checklist_blocks") or []:
                if not isinstance(block, dict):
                    continue
                out.append(
                    {
                        "deal_url": row.get("deal_url"),
                        "stage_name": row.get("stage_name") or stage_display_name(row.get("stage_id")),
                        "manager_name": row.get("manager_name"),
                        "call_number": index,
                        "subject": row.get("subject"),
                        "duration_minutes": row.get("duration_minutes"),
                        "sales_stage_block_name": block.get("sales_stage_block_name"),
                        "sales_stage_score": block.get("sales_stage_score"),
                        "sales_stage_max_score": block.get("sales_stage_max_score"),
                        "sales_stage_percent": block.get("sales_stage_percent"),
                        "sales_stage_missing": block.get("sales_stage_missing"),
                    }
                )
    return out


def _training_recommendation(block_name: Any, criterion: Any) -> str:
    block = str(block_name or "").lower()
    crit = str(criterion or "").strip()
    if "контакт" in block:
        return f"Отработать старт звонка: {crit}. Сделать короткий обязательный скрипт первых 20 секунд."
    if "потребност" in block:
        return f"Отработать выявление потребности: {crit}. Добавить 3-5 обязательных уточняющих вопросов до презентации."
    if "презентац" in block:
        return f"Усилить презентацию: {crit}. Привязывать предложение к задаче клиента и говорить языком выгоды."
    if "возраж" in block:
        return f"Отработать возражения: {crit}. Использовать схему: признать, уточнить причину, дать аргумент, закрепить следующий шаг."
    if "закрытие" in block:
        return f"Усилить закрытие звонка: {crit}. Фиксировать договоренность, срок и следующий контакт в разговоре и CRM."
    if "заполнение" in block:
        return f"Навести порядок в карточке сделки: {crit}. Сделать поле обязательным или добавить контроль перед переводом стадии."
    if "активность" in block:
        return f"Проверить дисциплину касаний: {crit}. Звонок менеджера должен быть виден в CRM после исключения Call-центра."
    if "связь" in block:
        return f"Синхронизировать разговор и CRM: {crit}. После звонка фиксировать следующий шаг и ключевые договоренности в карточке."
    if "движение" in block:
        return f"Контролировать движение по воронке: {crit}. Проверить зависшие стадии, возвраты и актуальность текущей стадии."
    return f"Разобрать на обучении: {crit}. Проверить фрагменты звонков с низкой оценкой."


def build_manager_stage_summary_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        if row.get("no_calls") or row.get("error"):
            continue
        manager = str(row.get("manager_name") or row.get("manager_id") or "Без менеджера")
        deal_key = _deal_group_key(row)
        for block in row.get("call_checklist_blocks") or []:
            if not isinstance(block, dict):
                continue
            block_name = str(block.get("sales_stage_block_name") or "")
            if not block_name:
                continue
            key = (manager, block_name)
            item = agg.setdefault(
                key,
                {
                    "manager_name": manager,
                    "sales_stage_block_name": block_name,
                    "manager_stage_calls": 0,
                    "_deals": set(),
                    "_percent_sum": 0.0,
                    "manager_stage_weak_calls": 0,
                },
            )
            percent = float(block.get("sales_stage_percent") or 0.0)
            item["manager_stage_calls"] += 1
            item["_percent_sum"] += percent
            item["_deals"].add(deal_key)
            if percent < 70:
                item["manager_stage_weak_calls"] += 1

    out: List[Dict[str, Any]] = []
    for item in agg.values():
        calls = max(1, int(item["manager_stage_calls"]))
        weak = int(item["manager_stage_weak_calls"])
        out.append(
            {
                "manager_name": item["manager_name"],
                "sales_stage_block_name": item["sales_stage_block_name"],
                "manager_stage_calls": calls,
                "manager_stage_deals": len(item["_deals"]),
                "manager_stage_avg_percent": round(float(item["_percent_sum"]) / calls, 2),
                "manager_stage_weak_calls": weak,
                "manager_stage_weak_rate": round(weak * 100.0 / calls, 2),
            }
        )
    return sorted(out, key=lambda r: (float(r.get("manager_stage_avg_percent") or 0.0), str(r.get("manager_name") or "")))


def build_manager_criterion_gap_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for row in rows:
        if row.get("no_calls") or row.get("error"):
            continue
        manager = str(row.get("manager_name") or row.get("manager_id") or "Без менеджера")
        for item in row.get("call_checklist_items") or []:
            if not isinstance(item, dict):
                continue
            block_name = str(item.get("checklist_block_name") or "")
            criterion = str(item.get("checklist_criterion") or "")
            if not criterion:
                continue
            key = (manager, block_name, criterion)
            bucket = agg.setdefault(
                key,
                {
                    "manager_name": manager,
                    "checklist_block_name": block_name,
                    "checklist_criterion": criterion,
                    "criterion_calls": 0,
                    "_score_sum": 0.0,
                    "criterion_failed_count": 0,
                    "criterion_partial_count": 0,
                    "training_recommendation": _training_recommendation(block_name, criterion),
                },
            )
            score = float(item.get("checklist_score") or 0.0)
            bucket["criterion_calls"] += 1
            bucket["_score_sum"] += score
            if score <= 0:
                bucket["criterion_failed_count"] += 1
            elif score < 1:
                bucket["criterion_partial_count"] += 1

    out: List[Dict[str, Any]] = []
    for bucket in agg.values():
        calls = max(1, int(bucket["criterion_calls"]))
        avg_score = float(bucket["_score_sum"]) / calls
        failed = int(bucket["criterion_failed_count"])
        partial = int(bucket["criterion_partial_count"])
        out.append(
            {
                "manager_name": bucket["manager_name"],
                "checklist_block_name": bucket["checklist_block_name"],
                "checklist_criterion": bucket["checklist_criterion"],
                "criterion_calls": calls,
                "criterion_avg_score": round(avg_score, 2),
                "criterion_completion_percent": round(avg_score * 100.0, 2),
                "criterion_failed_count": failed,
                "criterion_partial_count": partial,
                "criterion_fail_rate": round((failed + partial) * 100.0 / calls, 2),
                "training_recommendation": bucket["training_recommendation"],
            }
        )
    return sorted(
        out,
        key=lambda r: (
            float(r.get("criterion_completion_percent") or 0.0),
            -float(r.get("criterion_fail_rate") or 0.0),
            str(r.get("manager_name") or ""),
        ),
    )


def _aggregate_deal_crm_row(deal_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    sorted_rows = sorted(deal_rows, key=_call_sort_key)
    first = dict(sorted_rows[0])
    call_rows = [r for r in sorted_rows if not r.get("no_calls")]
    ok_calls = [r for r in call_rows if not r.get("error")]
    first["no_calls"] = bool(not call_rows or any(r.get("no_calls") for r in sorted_rows))
    first["calls_count"] = max(int(first.get("calls_count") or 0), len(call_rows))
    first["has_next_step_phrase"] = any(r.get("has_next_step_phrase") for r in ok_calls)
    first["next_step_synced"] = any(r.get("next_step_synced") for r in ok_calls)
    first["amount_mentioned"] = any(r.get("amount_mentioned") for r in ok_calls)
    first["crm_checklist_items"] = evaluate_crm_checklist(first, include_stage=True).get("crm_checklist_items") or []
    return first


def build_crm_checklist_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_row = _aggregate_deal_crm_row(deal_rows)
        for item in deal_row.get("crm_checklist_items") or []:
            if not isinstance(item, dict):
                continue
            out.append(
                {
                    "deal_url": deal_row.get("deal_url"),
                    "stage_name": deal_row.get("stage_name") or stage_display_name(deal_row.get("stage_id")),
                    "manager_name": deal_row.get("manager_name"),
                    "crm_checklist_block_name": item.get("crm_checklist_block_name"),
                    "crm_checklist_criterion": item.get("crm_checklist_criterion"),
                    "crm_checklist_score": item.get("crm_checklist_score"),
                    "crm_checklist_max_score": item.get("crm_checklist_max_score"),
                    "crm_checklist_comment": item.get("crm_checklist_comment"),
                    "crm_checklist_code": item.get("crm_checklist_code"),
                }
            )
    return out


def build_manager_crm_gap_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    agg: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for _, deal_rows in grouped.items():
        deal_row = _aggregate_deal_crm_row(deal_rows)
        manager = str(deal_row.get("manager_name") or deal_row.get("manager_id") or "Без менеджера")
        for item in deal_row.get("crm_checklist_items") or []:
            if not isinstance(item, dict):
                continue
            block_name = str(item.get("crm_checklist_block_name") or "")
            criterion = str(item.get("crm_checklist_criterion") or "")
            if not criterion:
                continue
            key = (manager, block_name, criterion)
            bucket = agg.setdefault(
                key,
                {
                    "manager_name": manager,
                    "crm_checklist_block_name": block_name,
                    "crm_checklist_criterion": criterion,
                    "crm_criterion_deals": 0,
                    "_score_sum": 0.0,
                    "crm_criterion_failed_count": 0,
                    "crm_criterion_partial_count": 0,
                    "training_recommendation": _training_recommendation(block_name, criterion),
                },
            )
            score = float(item.get("crm_checklist_score") or 0.0)
            bucket["crm_criterion_deals"] += 1
            bucket["_score_sum"] += score
            if score <= 0:
                bucket["crm_criterion_failed_count"] += 1
            elif score < 1:
                bucket["crm_criterion_partial_count"] += 1

    out: List[Dict[str, Any]] = []
    for bucket in agg.values():
        deals = max(1, int(bucket["crm_criterion_deals"]))
        avg_score = float(bucket["_score_sum"]) / deals
        failed = int(bucket["crm_criterion_failed_count"])
        partial = int(bucket["crm_criterion_partial_count"])
        out.append(
            {
                "manager_name": bucket["manager_name"],
                "crm_checklist_block_name": bucket["crm_checklist_block_name"],
                "crm_checklist_criterion": bucket["crm_checklist_criterion"],
                "crm_criterion_deals": deals,
                "crm_criterion_avg_score": round(avg_score, 2),
                "crm_criterion_completion_percent": round(avg_score * 100.0, 2),
                "crm_criterion_failed_count": failed,
                "crm_criterion_partial_count": partial,
                "crm_criterion_fail_rate": round((failed + partial) * 100.0 / deals, 2),
                "training_recommendation": bucket["training_recommendation"],
            }
        )
    return sorted(
        out,
        key=lambda r: (
            float(r.get("crm_criterion_completion_percent") or 0.0),
            -float(r.get("crm_criterion_fail_rate") or 0.0),
            str(r.get("manager_name") or ""),
        ),
    )


def _control_priority(score: float, reasons: List[str]) -> Tuple[int, str]:
    reason_text = " ".join(reasons).lower()
    if score >= 80 or "нет звонков" in reason_text or "тревога" in reason_text:
        return 1, "Критично"
    if score >= 55:
        return 2, "Высокий"
    if score >= 30:
        return 3, "Средний"
    return 4, "Наблюдать"


def _control_next_action(row: Dict[str, Any], reasons: List[str]) -> str:
    text = " ".join(reasons).lower()
    if "нет звонков" in text:
        return "Проверить, почему по сделке нет звонка менеджера; назначить контакт с клиентом и зафиксировать следующий шаг в CRM."
    if "ошиб" in text:
        return "Запустить режим «Повторить только ошибки», затем проверить, доступна ли запись звонка и корректно ли работает источник аудио."
    if "crm" in text or "следующий шаг" in text:
        return "Открыть карточку сделки, заполнить недостающие поля, зафиксировать договоренности и следующий шаг после разговора."
    if "разговор" in text or "возраж" in text:
        return "Разобрать расшифровку звонка с менеджером и закрепить конкретный сценарий: потребность, аргумент, следующий шаг."
    if "стади" in text or "воронк" in text:
        return "Проверить актуальность стадии, причину зависания и необходимость возврата/перевода сделки."
    return "Проверить сделку выборочно и оставить в мониторинге до следующего запуска отчета."


def build_quality_control_rows(deal_report_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in deal_report_rows:
        overall = float(row.get("avg_overall_score") or 0.0)
        call_score = float(row.get("avg_call_quality_score") or 0.0)
        crm_score = float(row.get("crm_checklist_percent") or row.get("crm_work_score") or 0.0)
        calls_total = int(row.get("calls_total") or 0)
        calls_failed = int(row.get("calls_failed") or 0)
        no_calls = bool(row.get("no_calls"))
        risk = str(row.get("stage_movement_risk") or "").strip()
        unhandled = str(row.get("unhandled_objections") or "").strip()
        reasons: List[str] = []
        risk_score = 0.0
        if no_calls or calls_total == 0:
            reasons.append("нет звонков менеджера")
            risk_score += 35
        if calls_failed:
            reasons.append(f"ошибки обработки звонков: {calls_failed}")
            risk_score += min(25, 8 * calls_failed)
        if overall < 40:
            reasons.append(f"низкая итоговая оценка: {overall:g}")
            risk_score += 25
        elif overall < 60:
            reasons.append(f"итоговая оценка требует контроля: {overall:g}")
            risk_score += 12
        if call_score < 50:
            reasons.append(f"слабое качество разговора: {call_score:g}")
            risk_score += 18
        if crm_score < 50:
            reasons.append(f"слабое ведение CRM: {crm_score:g}")
            risk_score += 18
        if risk and risk not in {"OK", "Финал"}:
            reasons.append(f"риск движения по воронке: {risk}")
            risk_score += 20 if "Тревога" in risk else 10
        if unhandled:
            reasons.append("есть неотработанные возражения")
            risk_score += 10
        if not reasons:
            reasons.append("критичных отклонений не найдено")

        order, priority = _control_priority(risk_score, reasons)
        out.append(
            {
                "control_priority_order": order,
                "control_priority": priority,
                "quality_control_score": round(min(100.0, risk_score), 2),
                "deal_url": row.get("deal_url"),
                "stage_name": row.get("stage_name"),
                "manager_name": row.get("manager_name"),
                "avg_overall_score": overall,
                "avg_call_quality_score": call_score,
                "crm_checklist_percent": crm_score,
                "calls_total": calls_total,
                "calls_failed": calls_failed,
                "stage_movement_risk": risk,
                "control_reason": "; ".join(reasons),
                "control_next_action": _control_next_action(row, reasons),
                "recommendations": row.get("recommendations"),
            }
        )
    return sorted(out, key=lambda r: (int(r.get("control_priority_order") or 9), -float(r.get("quality_control_score") or 0), str(r.get("manager_name") or "")))


def _coaching_priority(completion: float, fail_rate: float) -> str:
    if completion < 50 or fail_rate >= 60:
        return "Критично"
    if completion < 70 or fail_rate >= 35:
        return "Высокий"
    return "Средний"


def build_coaching_plan_rows(
    manager_criterion_gap_rows: List[Dict[str, Any]],
    manager_crm_gap_rows: List[Dict[str, Any]],
    manager_stage_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []

    for row in manager_criterion_gap_rows:
        completion = float(row.get("criterion_completion_percent") or 0.0)
        fail_rate = float(row.get("criterion_fail_rate") or 0.0)
        if completion >= 75 and fail_rate < 30:
            continue
        out.append(
            {
                "manager_name": row.get("manager_name"),
                "training_source": "Звонки",
                "training_topic": f"{row.get('checklist_block_name')}: {row.get('checklist_criterion')}",
                "coaching_priority": _coaching_priority(completion, fail_rate),
                "coaching_metric": f"выполнение {completion:g}%, проблем {fail_rate:g}%",
                "coaching_affected_count": row.get("criterion_calls"),
                "training_recommendation": row.get("training_recommendation"),
            }
        )

    for row in manager_crm_gap_rows:
        completion = float(row.get("crm_criterion_completion_percent") or 0.0)
        fail_rate = float(row.get("crm_criterion_fail_rate") or 0.0)
        if completion >= 80 and fail_rate < 25:
            continue
        out.append(
            {
                "manager_name": row.get("manager_name"),
                "training_source": "CRM",
                "training_topic": f"{row.get('crm_checklist_block_name')}: {row.get('crm_checklist_criterion')}",
                "coaching_priority": _coaching_priority(completion, fail_rate),
                "coaching_metric": f"выполнение {completion:g}%, проблем {fail_rate:g}%",
                "coaching_affected_count": row.get("crm_criterion_deals"),
                "training_recommendation": row.get("training_recommendation"),
            }
        )

    for row in manager_stage_rows:
        completion = float(row.get("manager_stage_avg_percent") or 0.0)
        weak_rate = float(row.get("manager_stage_weak_rate") or 0.0)
        if completion >= 75 and weak_rate < 30:
            continue
        stage = row.get("sales_stage_block_name")
        out.append(
            {
                "manager_name": row.get("manager_name"),
                "training_source": "Этап продаж",
                "training_topic": f"Этап: {stage}",
                "coaching_priority": _coaching_priority(completion, weak_rate),
                "coaching_metric": f"среднее выполнение {completion:g}%, слабых звонков {weak_rate:g}%",
                "coaching_affected_count": row.get("manager_stage_calls"),
                "training_recommendation": f"Провести разбор этапа «{stage}» на примерах звонков менеджера и закрепить короткий рабочий сценарий.",
            }
        )

    priority_order = {"Критично": 1, "Высокий": 2, "Средний": 3}
    return sorted(
        out,
        key=lambda r: (
            priority_order.get(str(r.get("coaching_priority") or ""), 9),
            str(r.get("manager_name") or ""),
            str(r.get("training_source") or ""),
        ),
    )


def _avg_numeric(rows: List[Dict[str, Any]], key: str) -> float:
    values = [float(row.get(key) or 0.0) for row in rows if row.get(key) is not None]
    return round(sum(values) / max(1, len(values)), 2)


def _count_by(rows: List[Dict[str, Any]], key: str, value: str) -> int:
    return sum(1 for row in rows if str(row.get(key) or "") == value)


def build_executive_summary_rows(
    deal_report_rows: List[Dict[str, Any]],
    manager_summary: List[Dict[str, Any]],
    quality_control_rows: List[Dict[str, Any]],
    manager_criterion_gap_rows: List[Dict[str, Any]],
    manager_crm_gap_rows: List[Dict[str, Any]],
    coaching_plan_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    def add(section: str, metric: str, value: Any, comment: str = "") -> None:
        rows.append(
            {
                "summary_section": section,
                "summary_metric": metric,
                "summary_value": value,
                "summary_comment": comment,
            }
        )

    deals_total = len(deal_report_rows)
    calls_total = sum(int(row.get("calls_total") or 0) for row in deal_report_rows)
    calls_ok = sum(int(row.get("calls_ok") or 0) for row in deal_report_rows)
    calls_failed = sum(int(row.get("calls_failed") or 0) for row in deal_report_rows)
    deals_without_calls = sum(1 for row in deal_report_rows if row.get("no_calls"))
    add("Общее", "Сделок в отчете", deals_total)
    add("Общее", "Звонков в сделках", calls_total)
    add("Общее", "Успешно обработано звонков", calls_ok)
    add("Общее", "Ошибок обработки звонков", calls_failed, "Используйте режим «Повторить только ошибки».")
    add("Общее", "Сделок без звонков менеджера", deals_without_calls, "Call-центр исключен из оценки менеджера.")
    add("Оценки", "Средняя итоговая оценка", _avg_numeric(deal_report_rows, "avg_overall_score"))
    add("Оценки", "Средняя оценка разговора", _avg_numeric(deal_report_rows, "avg_call_quality_score"))
    add("Оценки", "Средняя оценка CRM", _avg_numeric(deal_report_rows, "crm_checklist_percent"))
    add("Контроль", "Критичных сделок", _count_by(quality_control_rows, "control_priority", "Критично"))
    add("Контроль", "Сделок высокого приоритета", _count_by(quality_control_rows, "control_priority", "Высокий"))
    add("Контроль", "Сделок среднего приоритета", _count_by(quality_control_rows, "control_priority", "Средний"))

    for index, row in enumerate(quality_control_rows[:5], start=1):
        add(
            "Топ контроля",
            f"{index}. {row.get('manager_name') or ''}",
            row.get("deal_url"),
            str(row.get("control_reason") or ""),
        )

    for index, row in enumerate(manager_criterion_gap_rows[:5], start=1):
        add(
            "Провалы звонков",
            f"{index}. {row.get('manager_name') or ''}",
            f"{row.get('criterion_completion_percent')}%",
            f"{row.get('checklist_block_name')}: {row.get('checklist_criterion')}",
        )

    for index, row in enumerate(manager_crm_gap_rows[:5], start=1):
        add(
            "Провалы CRM",
            f"{index}. {row.get('manager_name') or ''}",
            f"{row.get('crm_criterion_completion_percent')}%",
            f"{row.get('crm_checklist_block_name')}: {row.get('crm_checklist_criterion')}",
        )

    for index, row in enumerate(coaching_plan_rows[:5], start=1):
        add(
            "Обучение",
            f"{index}. {row.get('manager_name') or ''}",
            row.get("coaching_priority"),
            f"{row.get('training_source')}: {row.get('training_topic')}",
        )

    return rows


def build_manager_scorecard_rows(
    manager_summary: List[Dict[str, Any]],
    quality_control_rows: List[Dict[str, Any]],
    coaching_plan_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    control_by_manager: Dict[str, List[Dict[str, Any]]] = {}
    for row in quality_control_rows:
        control_by_manager.setdefault(str(row.get("manager_name") or "Без менеджера"), []).append(row)

    coaching_by_manager: Dict[str, List[Dict[str, Any]]] = {}
    for row in coaching_plan_rows:
        coaching_by_manager.setdefault(str(row.get("manager_name") or "Без менеджера"), []).append(row)

    out: List[Dict[str, Any]] = []
    sorted_managers = sorted(manager_summary, key=lambda r: float(r.get("avg_overall_score") or 0.0))
    for rank, row in enumerate(sorted_managers, start=1):
        manager = str(row.get("manager_name") or "Без менеджера")
        control = control_by_manager.get(manager, [])
        coaching = coaching_by_manager.get(manager, [])
        critical = _count_by(control, "control_priority", "Критично")
        high = _count_by(control, "control_priority", "Высокий")
        top_training = coaching[0] if coaching else {}
        score = float(row.get("avg_overall_score") or 0.0)
        if critical:
            focus = "Срочный контроль сделок"
            next_action = "Начать с критичных сделок из листа «Контроль качества», затем провести точечный разбор звонков."
        elif high:
            focus = "Плановый контроль сделок"
            next_action = "Проверить сделки высокого приоритета и закрепить правила фиксации следующего шага."
        elif coaching:
            focus = "Обучение по стабильности качества"
            next_action = "Провести короткое обучение по главной теме из плана обучения."
        elif score < 70:
            focus = "Наблюдение"
            next_action = "Проверить несколько свежих сделок и сравнить динамику на следующем отчете."
        else:
            focus = "Поддерживать уровень"
            next_action = "Использовать лучшие звонки менеджера как примеры для команды."
        out.append(
            {
                "manager_rank": rank,
                "manager_name": manager,
                "avg_overall_score": row.get("avg_overall_score"),
                "deals_total": row.get("deals_total"),
                "calls_total": row.get("calls_total"),
                "calls_ok": row.get("calls_ok"),
                "deals_without_calls": row.get("deals_without_calls"),
                "manager_critical_deals": critical,
                "manager_high_deals": high,
                "manager_training_topics": len(coaching),
                "manager_top_training_topic": top_training.get("training_topic") or "",
                "top_growth_zones": row.get("top_growth_zones"),
                "top_checklist_gaps": row.get("top_checklist_gaps"),
                "manager_focus": focus,
                "manager_next_action": next_action,
            }
        )
    return out


def _unique_deal_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)
    return [sorted(deal_rows, key=_call_sort_key)[0] for deal_rows in grouped.values()]


def build_stage_history_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for deal_row in _unique_deal_rows(rows):
        for item in deal_row.get("stage_history_items") or []:
            if not isinstance(item, dict):
                continue
            out.append(
                {
                    "deal_url": deal_row.get("deal_url"),
                    "manager_name": deal_row.get("manager_name"),
                    "stage_history_event_id": item.get("stage_history_event_id"),
                    "stage_history_event_type": item.get("stage_history_event_type"),
                    "stage_history_created_at": item.get("stage_history_created_at"),
                    "stage_id": item.get("stage_id"),
                    "stage_name": item.get("stage_name"),
                    "stage_order": item.get("stage_order"),
                    "stage_duration_minutes": item.get("stage_duration_minutes"),
                    "stage_duration_hours": item.get("stage_duration_hours"),
                    "stage_is_current": item.get("stage_is_current"),
                }
            )
    return sorted(out, key=lambda r: (str(r.get("deal_url") or ""), str(r.get("stage_history_created_at") or "")))


def build_stage_sr_rows(rows: List[Dict[str, Any]], stage_map: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    deal_rows = _unique_deal_rows(rows)
    deals_total = len(deal_rows)
    stage_deals: Dict[str, set[str]] = {}
    ranks = stage_order_map(stage_map)

    for deal_row in deal_rows:
        deal_key = _deal_group_key(deal_row)
        seen_stages: set[str] = set()
        for item in deal_row.get("stage_history_items") or []:
            if not isinstance(item, dict):
                continue
            stage_id = str(item.get("stage_id") or "").strip()
            if stage_id:
                seen_stages.add(stage_id)
        if not seen_stages and deal_row.get("stage_id"):
            seen_stages.add(str(deal_row.get("stage_id")))
        for stage_id in seen_stages:
            stage_deals.setdefault(stage_id, set()).add(deal_key)

    out: List[Dict[str, Any]] = []
    for stage_id, deal_keys in stage_deals.items():
        passed = len(deal_keys)
        out.append(
            {
                "stage_id": stage_id,
                "stage_name": stage_display_name(stage_id, stage_map=stage_map),
                "stage_order": ranks.get(stage_id),
                "stage_deals_total": deals_total,
                "stage_deals_passed": passed,
                "stage_sr": round((passed / max(1, deals_total)) * 100.0, 2),
            }
        )
    return sorted(out, key=lambda r: (safe_int(r.get("stage_order")) or 9999, str(r.get("stage_name") or "")))


def build_stage_movement_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for deal_row in _unique_deal_rows(rows):
        out.append(
            {
                "deal_url": deal_row.get("deal_url"),
                "stage_name": deal_row.get("stage_name") or stage_display_name(deal_row.get("stage_id")),
                "manager_name": deal_row.get("manager_name"),
                "stage_history_count": deal_row.get("stage_history_count"),
                "stage_history_path": deal_row.get("stage_history_path"),
                "stage_last_change_at": deal_row.get("stage_last_change_at"),
                "stage_current_age_minutes": deal_row.get("stage_current_age_minutes"),
                "stage_current_age_days": deal_row.get("stage_current_age_days"),
                "stage_warning_threshold": deal_row.get("stage_warning_threshold"),
                "stage_critical_threshold": deal_row.get("stage_critical_threshold"),
                "stage_current_age_status": deal_row.get("stage_current_age_status"),
                "deal_total_work_minutes": deal_row.get("deal_total_work_minutes"),
                "deal_total_work_days": deal_row.get("deal_total_work_days"),
                "deal_total_work_warning_threshold": deal_row.get("deal_total_work_warning_threshold"),
                "deal_total_work_critical_threshold": deal_row.get("deal_total_work_critical_threshold"),
                "deal_total_work_status": deal_row.get("deal_total_work_status"),
                "stage_return_count": deal_row.get("stage_return_count"),
                "stage_reached_final": deal_row.get("stage_reached_final"),
                "stage_movement_risk": deal_row.get("stage_movement_risk"),
                "stage_movement_recommendation": deal_row.get("stage_movement_recommendation"),
            }
        )
    risk_order = {"Зависла": 0, "Нет продвижения": 1, "Возвраты": 2, "Нет истории": 3, "OK": 4, "Финал": 5}
    return sorted(out, key=lambda r: (risk_order.get(str(r.get("stage_movement_risk") or ""), 9), str(r.get("manager_name") or "")))


def build_manager_summary(rows: List[Dict[str, Any]], score_key: str = "overall_score") -> List[Dict[str, Any]]:
    agg: Dict[str, Dict[str, Any]] = {}
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    for _, deal_rows in grouped.items():
        first = deal_rows[0]
        mid = str(first.get("manager_id") or "unknown")
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        ok_calls = [r for r in call_rows if not r.get("error")]
        scored_calls = [r for r in ok_calls if r.get(score_key) is not None]
        deal_score = (
            sum(float(r.get(score_key) or 0.0) for r in scored_calls) / len(scored_calls)
            if scored_calls
            else 0.0
        )
        a = agg.setdefault(
            mid,
            {
                "manager_id": mid,
                "manager_name": first.get("manager_name") or mid,
                "deals_total": 0,
                "calls_total": 0,
                "calls_ok": 0,
                "deals_without_calls": 0,
                "scored_deals": 0,
                "overall_score_sum": 0.0,
                "growth_deal_data": 0,
                "growth_call_structure": 0,
                "growth_alignment": 0,
                "_checklist_gaps": {},
            },
        )
        a["deals_total"] += 1
        a["calls_total"] += len(call_rows)
        a["calls_ok"] += len(ok_calls)
        if any(r.get("no_calls") for r in deal_rows) or not call_rows:
            a["deals_without_calls"] += 1
        a["scored_deals"] += 1
        a["overall_score_sum"] += deal_score
        if any(not r.get("has_contact") or not r.get("has_amount") or not r.get("has_comments") for r in deal_rows):
            a["growth_deal_data"] += 1
        if not ok_calls or any(float(r.get("call_checklist_percent") or r.get("call_quality_score") or 0.0) < 70 for r in ok_calls):
            a["growth_call_structure"] += 1
        if not ok_calls or not any(r.get("next_step_synced") for r in ok_calls):
            a["growth_alignment"] += 1
        gaps = a.setdefault("_checklist_gaps", {})
        if isinstance(gaps, dict):
            for call in ok_calls:
                for item in call.get("call_checklist_items") or []:
                    if not isinstance(item, dict):
                        continue
                    score = float(item.get("checklist_score") or 0.0)
                    if score >= 1:
                        continue
                    criterion = str(item.get("checklist_criterion") or "").strip()
                    if criterion:
                        gaps[criterion] = int(gaps.get(criterion, 0)) + 1

    out: List[Dict[str, Any]] = []
    for _, a in agg.items():
        total = max(1, int(a["scored_deals"]))
        a["avg_overall_score"] = round(float(a["overall_score_sum"]) / total, 2)
        a.pop("scored_deals", None)
        gaps = a.pop("_checklist_gaps", {}) or {}
        if isinstance(gaps, dict):
            a["top_checklist_gaps"] = "; ".join(
                criterion for criterion, _ in sorted(gaps.items(), key=lambda x: int(x[1]), reverse=True)[:5]
            )
        else:
            a["top_checklist_gaps"] = ""
        zones = sorted(
            [("Ведение CRM", a["growth_deal_data"]), ("Структура разговора", a["growth_call_structure"]), ("Синхронизация звонок↔CRM", a["growth_alignment"])],
            key=lambda x: x[1],
            reverse=True,
        )
        a["top_growth_zones"] = ", ".join([z[0] for z in zones[:3]])
        out.append(a)
    return sorted(out, key=lambda x: x["avg_overall_score"])


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, str) and len(value) > 32000:
        return value[:31900] + "\n\n[Текст обрезан для Excel. Полная версия сохранена в txt-файле.]"
    return value


def _excel_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.map(_excel_safe_value)


def _ru_df(df: pd.DataFrame) -> pd.DataFrame:
    return _excel_df(df).rename(columns=RU_COLUMNS)


def kpi_profile_display(profile: Any) -> Tuple[str, str]:
    key = str(profile or "").strip()
    if not key:
        return "", ""
    title, explanation = KPI_PROFILE_DESCRIPTIONS.get(key, (key, "Пользовательский профиль KPI."))
    return title, explanation


def prepare_report_rows(rows: List[Dict[str, Any]], stage_map: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        r = dict(row)
        r["stage_name"] = stage_display_name(r.get("stage_id"), stage_map=stage_map) or r.get("stage_name")
        profile_title, profile_explanation = kpi_profile_display(r.get("kpi_profile"))
        r["kpi_profile_ru"] = profile_title
        r["kpi_explanation"] = profile_explanation
        if r.get("kpi_profile_cmp"):
            cmp_title, cmp_explanation = kpi_profile_display(r.get("kpi_profile_cmp"))
            r["kpi_profile_cmp_ru"] = cmp_title
            r["kpi_explanation_cmp"] = cmp_explanation
        out.append(r)
    return out


def _format_excel_writer(writer: pd.ExcelWriter) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return

    issue_fill = PatternFill("solid", fgColor="FFF2CC")
    unhandled_fill = PatternFill("solid", fgColor="F4CCCC")
    handled_fill = PatternFill("solid", fgColor="D9EAD3")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    critical_fill = PatternFill("solid", fgColor="F4CCCC")

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_cells in ws.columns:
            header = str(col_cells[0].value or "")
            max_len = len(header)
            for cell in col_cells[1:80]:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, min(80, len(str(value))))
                if isinstance(value, str) and ("\n" in value or len(value) > 90):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            width = min(70, max(12, max_len + 2))
            if any(word in header.lower() for word in ["расшифров", "вывод", "рекомендац", "диагност", "попытки", "ошибки"]):
                width = min(90, max(width, 45))
            ws.column_dimensions[col_cells[0].column_letter].width = width

        headers = {str(cell.value or ""): cell.column for cell in ws[1]}
        issue_headers = [
            "Моменты для улучшения",
            "Неотработанные возражения",
            "Варианты отработки возражений",
            "Расшифровка с пометками",
            "Что усилить по этапу",
        ]
        for row_idx in range(2, ws.max_row + 1):
            for header in issue_headers:
                col_idx = headers.get(header)
                if not col_idx:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell.fill = issue_fill

            status_col = headers.get("Статус отработки")
            if status_col:
                status = str(ws.cell(row=row_idx, column=status_col).value or "").lower()
                if "не отработано" in status:
                    fill = unhandled_fill
                elif "отработано" in status:
                    fill = handled_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            risk_headers = ["Статус срока в стадии", "Статус общего срока сделки", "Риск движения по воронке"]
            risk_values = " ".join(
                str(ws.cell(row=row_idx, column=headers[h]).value or "").lower()
                for h in risk_headers
                if h in headers
            )
            if "тревога" in risk_values:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = critical_fill
            elif "предупреждение" in risk_values:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = warning_fill

            checklist_score_col = headers.get("Оценка критерия")
            if not checklist_score_col:
                checklist_score_col = headers.get("Оценка CRM-критерия")
            if checklist_score_col:
                try:
                    checklist_score = float(ws.cell(row=row_idx, column=checklist_score_col).value or 0)
                    if checklist_score <= 0:
                        fill = critical_fill
                    elif checklist_score < 1:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            stage_percent_col = headers.get("Выполнение этапа, %")
            if stage_percent_col:
                try:
                    stage_percent = float(ws.cell(row=row_idx, column=stage_percent_col).value or 0)
                    if stage_percent < 50:
                        fill = critical_fill
                    elif stage_percent < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            avg_stage_percent_col = headers.get("Среднее выполнение этапа, %")
            if avg_stage_percent_col:
                try:
                    avg_stage_percent = float(ws.cell(row=row_idx, column=avg_stage_percent_col).value or 0)
                    if avg_stage_percent < 50:
                        fill = critical_fill
                    elif avg_stage_percent < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            criterion_completion_col = headers.get("Выполнение критерия, %")
            if not criterion_completion_col:
                criterion_completion_col = headers.get("Выполнение CRM-критерия, %")
            if criterion_completion_col:
                try:
                    completion = float(ws.cell(row=row_idx, column=criterion_completion_col).value or 0)
                    if completion < 50:
                        fill = critical_fill
                    elif completion < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            priority_col = headers.get("Приоритет контроля") or headers.get("Приоритет обучения")
            if priority_col:
                priority = str(ws.cell(row=row_idx, column=priority_col).value or "").lower()
                if "критично" in priority:
                    fill = critical_fill
                elif "высок" in priority:
                    fill = warning_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill


def flatten_results(
    rows: List[Dict[str, Any]],
    manager_summary: List[Dict[str, Any]],
    manager_summary_cmp: Optional[List[Dict[str, Any]]] = None,
    stage_map: Optional[Dict[str, str]] = None,
) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = REPORTS_DIR / f"bitnewton_sync_report_{ts}.xlsx"
    report_rows = prepare_report_rows(rows, stage_map=stage_map)
    df = pd.DataFrame(report_rows)
    deal_report_rows = build_deal_report_rows(report_rows)
    call_detail_rows = build_call_detail_rows(report_rows)
    objection_rows = build_objection_rows(report_rows)
    checklist_rows = build_call_checklist_rows(report_rows)
    sales_stage_score_rows = build_sales_stage_score_rows(report_rows)
    manager_stage_rows = build_manager_stage_summary_rows(report_rows)
    manager_criterion_gap_rows = build_manager_criterion_gap_rows(report_rows)
    crm_checklist_rows = build_crm_checklist_rows(report_rows)
    manager_crm_gap_rows = build_manager_crm_gap_rows(report_rows)
    quality_control_rows = build_quality_control_rows(deal_report_rows)
    coaching_plan_rows = build_coaching_plan_rows(manager_criterion_gap_rows, manager_crm_gap_rows, manager_stage_rows)
    executive_summary_rows = build_executive_summary_rows(
        deal_report_rows,
        manager_summary,
        quality_control_rows,
        manager_criterion_gap_rows,
        manager_crm_gap_rows,
        coaching_plan_rows,
    )
    manager_scorecard_rows = build_manager_scorecard_rows(manager_summary, quality_control_rows, coaching_plan_rows)
    stage_history_rows = build_stage_history_rows(report_rows)
    stage_sr_rows = build_stage_sr_rows(report_rows, stage_map=stage_map)
    stage_movement_rows = build_stage_movement_rows(report_rows)
    deal_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "kpi_profile_ru",
        "kpi_explanation",
        "calls_total",
        "calls_ok",
        "calls_failed",
        "ignored_call_center_calls",
        "no_calls",
        "transcripts_count",
        "avg_overall_score",
        "overall_score_details",
        "avg_call_quality_score",
        "call_quality_details",
        "call_checklist_percent",
        "call_checklist_block_details",
        "deal_quality_score",
        "deal_quality_details",
        "alignment_score",
        "crm_work_score",
        "crm_checklist_percent",
        "crm_checklist_details",
        "alignment_details",
        "stage_history_count",
        "stage_history_path",
        "stage_last_change_at",
        "stage_current_age_minutes",
        "stage_current_age_days",
        "stage_warning_threshold",
        "stage_critical_threshold",
        "stage_current_age_status",
        "deal_total_work_minutes",
        "deal_total_work_days",
        "deal_total_work_warning_threshold",
        "deal_total_work_critical_threshold",
        "deal_total_work_status",
        "stage_return_count",
        "stage_reached_final",
        "stage_movement_risk",
        "stage_movement_recommendation",
        "client_work_score",
        "client_work_quality",
        "conversation_quality",
        "client_work_conclusion",
        "call_quality_conclusion",
        "conversation_meaning",
        "recommendations",
        "calls_breakdown",
        "improvement_moments_combined",
        "objections_combined",
        "unhandled_objections",
        "objection_recommendations",
        "transcript_paths",
        "error",
    ]
    call_detail_cols = [
        "deal_url",
        "call_number",
        "subject",
        "duration_minutes",
        "call_has_error",
        "call_quality_score",
        "call_quality_details",
        "call_checklist_total_score",
        "call_checklist_max_score",
        "call_checklist_percent",
        "call_checklist_block_details",
        "has_greeting",
        "has_needs_discovery",
        "has_objection_work",
        "has_next_step_phrase",
        "alignment_score",
        "crm_work_score",
        "crm_checklist_percent",
        "crm_checklist_details",
        "alignment_details",
        "next_step_synced",
        "next_step_synced_details",
        "call_quality_conclusion",
        "conversation_meaning",
        "improvement_moments",
        "unhandled_objections",
        "objection_recommendations",
        "transcript_path",
        "transcript_marked",
        "transcript_text",
        "bitrix_card_transcript_status",
        "transcript_match_score",
        "bitrix_card_transcript",
        "error",
    ]
    objection_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "subject",
        "objection_fragment",
        "objection_status",
        "objection_recommendations",
    ]
    checklist_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "call_number",
        "subject",
        "duration_minutes",
        "checklist_block_name",
        "checklist_criterion",
        "checklist_score",
        "checklist_max_score",
        "checklist_evidence",
        "checklist_comment",
        "checklist_code",
    ]
    sales_stage_score_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "call_number",
        "subject",
        "duration_minutes",
        "sales_stage_block_name",
        "sales_stage_score",
        "sales_stage_max_score",
        "sales_stage_percent",
        "sales_stage_missing",
    ]
    manager_stage_cols = [
        "manager_name",
        "sales_stage_block_name",
        "manager_stage_calls",
        "manager_stage_deals",
        "manager_stage_avg_percent",
        "manager_stage_weak_calls",
        "manager_stage_weak_rate",
    ]
    manager_criterion_gap_cols = [
        "manager_name",
        "checklist_block_name",
        "checklist_criterion",
        "criterion_calls",
        "criterion_avg_score",
        "criterion_completion_percent",
        "criterion_failed_count",
        "criterion_partial_count",
        "criterion_fail_rate",
        "training_recommendation",
    ]
    crm_checklist_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "crm_checklist_block_name",
        "crm_checklist_criterion",
        "crm_checklist_score",
        "crm_checklist_max_score",
        "crm_checklist_comment",
        "crm_checklist_code",
    ]
    manager_crm_gap_cols = [
        "manager_name",
        "crm_checklist_block_name",
        "crm_checklist_criterion",
        "crm_criterion_deals",
        "crm_criterion_avg_score",
        "crm_criterion_completion_percent",
        "crm_criterion_failed_count",
        "crm_criterion_partial_count",
        "crm_criterion_fail_rate",
        "training_recommendation",
    ]
    quality_control_cols = [
        "control_priority",
        "quality_control_score",
        "deal_url",
        "stage_name",
        "manager_name",
        "avg_overall_score",
        "avg_call_quality_score",
        "crm_checklist_percent",
        "calls_total",
        "calls_failed",
        "stage_movement_risk",
        "control_reason",
        "control_next_action",
        "recommendations",
    ]
    coaching_plan_cols = [
        "manager_name",
        "coaching_priority",
        "training_source",
        "training_topic",
        "coaching_metric",
        "coaching_affected_count",
        "training_recommendation",
    ]
    executive_summary_cols = [
        "summary_section",
        "summary_metric",
        "summary_value",
        "summary_comment",
    ]
    manager_scorecard_cols = [
        "manager_rank",
        "manager_name",
        "avg_overall_score",
        "deals_total",
        "calls_total",
        "calls_ok",
        "deals_without_calls",
        "manager_critical_deals",
        "manager_high_deals",
        "manager_training_topics",
        "manager_top_training_topic",
        "top_growth_zones",
        "top_checklist_gaps",
        "manager_focus",
        "manager_next_action",
    ]
    stage_history_cols = [
        "deal_url",
        "manager_name",
        "stage_history_event_id",
        "stage_history_event_type",
        "stage_history_created_at",
        "stage_id",
        "stage_name",
        "stage_order",
        "stage_duration_minutes",
        "stage_duration_hours",
        "stage_is_current",
    ]
    stage_sr_cols = [
        "stage_order",
        "stage_id",
        "stage_name",
        "stage_deals_total",
        "stage_deals_passed",
        "stage_sr",
    ]
    stage_movement_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "stage_history_count",
        "stage_history_path",
        "stage_last_change_at",
        "stage_current_age_minutes",
        "stage_current_age_days",
        "stage_warning_threshold",
        "stage_critical_threshold",
        "stage_current_age_status",
        "deal_total_work_minutes",
        "deal_total_work_days",
        "deal_total_work_warning_threshold",
        "deal_total_work_critical_threshold",
        "deal_total_work_status",
        "stage_return_count",
        "stage_reached_final",
        "stage_movement_risk",
        "stage_movement_recommendation",
    ]
    manager_cols = [
        "manager_name",
        "deals_total",
        "calls_total",
        "calls_ok",
        "deals_without_calls",
        "growth_deal_data",
        "growth_call_structure",
        "growth_alignment",
        "avg_overall_score",
        "top_growth_zones",
        "top_checklist_gaps",
    ]
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        executive_summary_df = pd.DataFrame(executive_summary_rows, columns=executive_summary_cols)
        _ru_df(executive_summary_df).to_excel(writer, sheet_name="Итоги", index=False)
        manager_scorecard_df = pd.DataFrame(manager_scorecard_rows, columns=manager_scorecard_cols)
        _ru_df(manager_scorecard_df).to_excel(writer, sheet_name="Карточки менеджеров", index=False)
        deal_df = pd.DataFrame(deal_report_rows)
        _ru_df(deal_df[[c for c in deal_cols if c in deal_df.columns]]).to_excel(writer, sheet_name="Сделки", index=False)
        if not df.empty:
            call_df = pd.DataFrame(call_detail_rows)
            _ru_df(call_df[[c for c in call_detail_cols if c in call_df.columns]]).to_excel(writer, sheet_name="Звонки внутри сделок", index=False)
            objection_df = pd.DataFrame(objection_rows, columns=objection_cols)
            _ru_df(objection_df).to_excel(writer, sheet_name="Разбор возражений", index=False)
            checklist_df = pd.DataFrame(checklist_rows, columns=checklist_cols)
            _ru_df(checklist_df).to_excel(writer, sheet_name="Чек-лист звонков", index=False)
            sales_stage_score_df = pd.DataFrame(sales_stage_score_rows, columns=sales_stage_score_cols)
            _ru_df(sales_stage_score_df).to_excel(writer, sheet_name="Этапы продаж", index=False)
            manager_stage_df = pd.DataFrame(manager_stage_rows, columns=manager_stage_cols)
            _ru_df(manager_stage_df).to_excel(writer, sheet_name="Слабые этапы", index=False)
            manager_criterion_gap_df = pd.DataFrame(manager_criterion_gap_rows, columns=manager_criterion_gap_cols)
            _ru_df(manager_criterion_gap_df).to_excel(writer, sheet_name="Проблемные критерии", index=False)
            crm_checklist_df = pd.DataFrame(crm_checklist_rows, columns=crm_checklist_cols)
            _ru_df(crm_checklist_df).to_excel(writer, sheet_name="Чек-лист CRM", index=False)
            manager_crm_gap_df = pd.DataFrame(manager_crm_gap_rows, columns=manager_crm_gap_cols)
            _ru_df(manager_crm_gap_df).to_excel(writer, sheet_name="Проблемы CRM", index=False)
            quality_control_df = pd.DataFrame(quality_control_rows, columns=quality_control_cols)
            _ru_df(quality_control_df).to_excel(writer, sheet_name="Контроль качества", index=False)
            coaching_plan_df = pd.DataFrame(coaching_plan_rows, columns=coaching_plan_cols)
            _ru_df(coaching_plan_df).to_excel(writer, sheet_name="План обучения", index=False)
            stage_history_df = pd.DataFrame(stage_history_rows, columns=stage_history_cols)
            _ru_df(stage_history_df).to_excel(writer, sheet_name="История стадий", index=False)
            stage_sr_df = pd.DataFrame(stage_sr_rows, columns=stage_sr_cols)
            _ru_df(stage_sr_df).to_excel(writer, sheet_name="SR по стадиям", index=False)
            stage_movement_df = pd.DataFrame(stage_movement_rows, columns=stage_movement_cols)
            _ru_df(stage_movement_df).to_excel(writer, sheet_name="Риски движения", index=False)
        manager_df = pd.DataFrame(manager_summary)
        _ru_df(manager_df[[c for c in manager_cols if c in manager_df.columns]]).to_excel(writer, sheet_name="Сводка менеджеров", index=False)
        if manager_summary_cmp is not None:
            manager_cmp_df = pd.DataFrame(manager_summary_cmp)
            _ru_df(manager_cmp_df[[c for c in manager_cols if c in manager_cmp_df.columns]]).to_excel(writer, sheet_name="Сводка менеджеров cmp", index=False)
        _format_excel_writer(writer)
    return out


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(base)
    for k, v in (override or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def validate_kpi_config(kpi: Dict[str, Any]) -> None:
    if not isinstance(kpi, dict):
        raise RuntimeError("KPI config: должен быть объектом")
    prof = kpi.get("profile")
    if not isinstance(prof, dict) or not str(prof.get("name") or "").strip():
        raise RuntimeError("KPI config: profile.name обязателен (строка)")
    sla = kpi.get("sla")
    if not isinstance(sla, dict):
        raise RuntimeError("KPI config: sla должен быть объектом")
    fr = float(sla.get("first_response_hours", FIRST_RESPONSE_SLA_HOURS))
    mg = float(sla.get("max_gap_between_calls_hours", MAX_GAP_BETWEEN_CALLS_HOURS))
    if fr <= 0 or mg <= 0:
        raise RuntimeError("KPI config: sla значения должны быть > 0")
    w = kpi.get("weights")
    if not isinstance(w, dict):
        raise RuntimeError("KPI config: weights должен быть объектом")
    overall = w.get("overall")
    if not isinstance(overall, dict):
        raise RuntimeError("KPI config: weights.overall должен быть объектом")
    s = float(overall.get("call_quality", 0)) + float(overall.get("discipline", 0)) + float(overall.get("crm_alignment", 0))
    if not (0.95 <= s <= 1.05):
        raise RuntimeError(f"KPI config: weights.overall должны суммироваться примерно в 1.0 (сейчас {s})")
    patterns = kpi.get("patterns")
    if patterns is not None:
        if not isinstance(patterns, dict):
            raise RuntimeError("KPI config: patterns должен быть объектом")
        for key, arr in patterns.items():
            if not isinstance(arr, list):
                raise RuntimeError(f"KPI config: patterns.{key} должен быть списком regex-строк")
            for p in arr:
                if not isinstance(p, str):
                    raise RuntimeError(f"KPI config: patterns.{key} содержит не строку")


def enforce_reaction_kpi(kpi: Dict[str, Any]) -> Dict[str, Any]:
    """
    Единая итоговая оценка: разговор + ведение CRM. Скорость реакции не влияет на итог.
    """
    sla = kpi.setdefault("sla", {})
    if isinstance(sla, dict):
        sla["first_response_hours"] = FIRST_RESPONSE_SLA_HOURS
        sla.setdefault("max_gap_between_calls_hours", MAX_GAP_BETWEEN_CALLS_HOURS)
    weights = kpi.setdefault("weights", {})
    if isinstance(weights, dict):
        weights["overall"] = {"call_quality": 0.50, "discipline": 0.0, "crm_alignment": 0.50}
        weights["discipline_split"] = {"first_response": 1.0, "cadence": 0.0}
    return kpi


def load_kpi_config(path: Optional[str]) -> Dict[str, Any]:
    cfg = dict(DEFAULT_KPI_CONFIG)
    if not path:
        cfg = enforce_reaction_kpi(cfg)
        validate_kpi_config(cfg)
        return cfg
    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise RuntimeError("--kpi-config должен содержать JSON-объект")
    merged = _deep_merge(cfg, raw)
    if isinstance(merged.get("profile"), dict):
        merged["profile"].setdefault("name", Path(path).name)
        merged["profile"].setdefault("version", "1")
    else:
        merged["profile"] = {"name": Path(path).name, "version": "1"}
    merged = enforce_reaction_kpi(merged)
    validate_kpi_config(merged)
    return merged


def _load_state_cache() -> Dict[str, Any]:
    try:
        if STATE_CACHE_PATH.exists():
            raw = json.loads(STATE_CACHE_PATH.read_text(encoding="utf-8"))
            return raw if isinstance(raw, dict) else {}
    except Exception:
        return {}
    return {}


def _save_state_cache(cache: Dict[str, Any]) -> None:
    try:
        STATE_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        STATE_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def load_report_json(path: str | Path) -> List[Dict[str, Any]]:
    report_path = Path(path)
    if not report_path.exists():
        raise SystemExit(f"Отчет не найден: {report_path}")
    raw = json.loads(report_path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise SystemExit(f"Ожидался JSON-список строк отчета: {report_path}")
    return [r for r in raw if isinstance(r, dict)]


def deal_id_from_report_row(row: Dict[str, Any]) -> Optional[str]:
    deal_id = row.get("deal_id")
    if deal_id:
        deal_str = str(deal_id).strip()
        if deal_str.isdigit():
            return deal_str
    url = str(row.get("deal_url") or "").strip()
    m = re.search(r"/crm/deal/details/(\d+)/?", url)
    return m.group(1) if m else None


def load_retry_scope(path: str | Path) -> Dict[str, Any]:
    rows = load_report_json(path)
    deal_ids: List[str] = []
    activity_ids_by_deal: Dict[str, set[int]] = {}
    full_deals: set[str] = set()
    errors = 0
    seen_deals: set[str] = set()

    for row in rows:
        if not row.get("error"):
            continue
        deal_id = deal_id_from_report_row(row)
        if not deal_id:
            continue
        errors += 1
        if deal_id not in seen_deals:
            seen_deals.add(deal_id)
            deal_ids.append(deal_id)
        activity_id = safe_int(row.get("activity_id"))
        if activity_id is None:
            full_deals.add(deal_id)
        else:
            activity_ids_by_deal.setdefault(deal_id, set()).add(activity_id)

    return {
        "deal_ids": deal_ids,
        "activity_ids_by_deal": activity_ids_by_deal,
        "full_deals": full_deals,
        "errors": errors,
        "source_rows": rows,
        "source_path": str(path),
    }


def _report_row_identity(row: Dict[str, Any]) -> Tuple[Optional[str], Optional[int]]:
    return deal_id_from_report_row(row), safe_int(row.get("activity_id"))


def merge_retry_results(
    original_rows: List[Dict[str, Any]],
    retry_rows: List[Dict[str, Any]],
    retry_scope: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Повтор ошибок должен обновлять полный отчет, а не заменять его маленьким отчетом
    только по ошибочным строкам.
    """
    full_deals = set(str(x) for x in (retry_scope.get("full_deals") or set()))
    retry_activity_ids_by_deal = retry_scope.get("activity_ids_by_deal") or {}

    retry_by_deal: Dict[str, List[Dict[str, Any]]] = {}
    retry_by_activity: Dict[Tuple[str, int], List[Dict[str, Any]]] = {}
    for row in retry_rows:
        deal_id, activity_id = _report_row_identity(row)
        if not deal_id:
            continue
        retry_by_deal.setdefault(deal_id, []).append(row)
        if activity_id is not None:
            retry_by_activity.setdefault((deal_id, activity_id), []).append(row)

    merged: List[Dict[str, Any]] = []
    inserted_full_deals: set[str] = set()
    inserted_activities: set[Tuple[str, int]] = set()
    inserted_fallback_deals: set[str] = set()

    for row in original_rows:
        deal_id, activity_id = _report_row_identity(row)
        if not deal_id:
            merged.append(row)
            continue

        if deal_id in full_deals:
            if deal_id not in inserted_full_deals:
                replacements = retry_by_deal.get(deal_id)
                merged.extend(replacements if replacements else [row])
                inserted_full_deals.add(deal_id)
            continue

        target_activities = retry_activity_ids_by_deal.get(deal_id, set())
        if activity_id is not None and activity_id in target_activities:
            key = (deal_id, activity_id)
            if key not in inserted_activities:
                replacements = retry_by_activity.get(key)
                if replacements:
                    merged.extend(replacements)
                elif deal_id not in inserted_fallback_deals:
                    fallback = [r for r in retry_by_deal.get(deal_id, []) if safe_int(r.get("activity_id")) is None]
                    merged.extend(fallback if fallback else [row])
                    inserted_fallback_deals.add(deal_id)
                else:
                    merged.append(row)
                inserted_activities.add(key)
            continue

        merged.append(row)

    existing_object_ids = {id(row) for row in merged}
    for row in retry_rows:
        if id(row) in existing_object_ids:
            continue
        deal_id, activity_id = _report_row_identity(row)
        if not deal_id:
            merged.append(row)
            continue
        if deal_id in full_deals and deal_id in inserted_full_deals:
            continue
        if activity_id is not None and (deal_id, activity_id) in inserted_activities:
            continue
        merged.append(row)

    return merged


def _sha256_text(s: str) -> str:
    return hashlib.sha256((s or "").encode("utf-8", errors="ignore")).hexdigest()


def cleanup_old_chrome_tmp_profiles(base_dir: Path, keep_days: int = 7) -> int:
    try:
        if keep_days <= 0:
            return 0
        now = time.time()
        removed = 0
        for p in base_dir.glob("chrome_profile_tmp_*"):
            try:
                age_days = (now - p.stat().st_mtime) / (3600 * 24)
                if age_days < keep_days:
                    continue
                for sub in sorted(p.rglob("*"), reverse=True):
                    try:
                        if sub.is_file() or sub.is_symlink():
                            sub.unlink(missing_ok=True)  # type: ignore[call-arg]
                        else:
                            sub.rmdir()
                    except Exception:
                        pass
                try:
                    p.rmdir()
                except Exception:
                    pass
                removed += 1
            except Exception:
                continue
        return removed
    except Exception:
        return 0


def cleanup_old_outputs(base_dir: Path, keep_days: int = 30, extra_audio_dirs: Optional[List[Path]] = None) -> Dict[str, int]:
    counts = {"reports": 0, "audio": 0, "transcripts": 0, "total": 0}
    try:
        if keep_days <= 0:
            return counts
        base_dir = Path(base_dir)
        cutoff = time.time() - (float(keep_days) * 24 * 3600)

        def remove_file(path: Path, bucket: str) -> None:
            try:
                if not path.exists() or path.is_dir():
                    return
                if path.stat().st_mtime >= cutoff:
                    return
                path.unlink(missing_ok=True)  # type: ignore[call-arg]
                counts[bucket] = counts.get(bucket, 0) + 1
                counts["total"] += 1
            except Exception:
                return

        for pattern in ("bitnewton_sync_report_*.json", "bitnewton_sync_report_*.xlsx"):
            for path in base_dir.glob(pattern):
                remove_file(path, "reports")

        cleanup_dirs: List[Tuple[Path, str]] = [
            (base_dir / "audio", "audio"),
            (base_dir / "audio_ui", "audio"),
            (base_dir / "transcripts", "transcripts"),
        ]
        for extra in extra_audio_dirs or []:
            extra_path = Path(extra)
            if extra_path not in [p for p, _ in cleanup_dirs]:
                cleanup_dirs.append((extra_path, "audio"))

        for folder, bucket in cleanup_dirs:
            if not folder.exists():
                continue
            for path in folder.rglob("*"):
                remove_file(path, bucket)
            for path in sorted(folder.rglob("*"), reverse=True):
                try:
                    if path.is_dir() and not any(path.iterdir()):
                        path.rmdir()
                except Exception:
                    continue
        return counts
    except Exception:
        return counts


def apply_scores(row: Dict[str, Any], deal: Dict[str, Any], comments: List[str], text: str, kpi: Dict[str, Any], suffix: str = "") -> None:
    first_h = row.get("first_response_hours")
    first_m = row.get("first_response_minutes")
    sla_cfg = kpi.get("sla", {})
    first_response_sla = float(sla_cfg.get("first_response_hours", FIRST_RESPONSE_SLA_HOURS))
    if first_h is None and first_m is not None:
        first_h = float(first_m) / 60.0
    row[f"first_response_sla_ok{suffix}"] = (first_h is not None and float(first_h) <= first_response_sla)

    deal_q = compute_deal_quality(deal, comments, kpi)
    call_q = evaluate_call_text(text, kpi)
    align = crm_call_alignment(deal, text, comments, kpi)
    for k, v in deal_q.items():
        row[f"{k}{suffix}"] = v
    for k, v in call_q.items():
        row[f"{k}{suffix}"] = v
    for k, v in align.items():
        row[f"{k}{suffix}"] = v

    crm_q = evaluate_crm_checklist(row, suffix=suffix, include_stage=False)
    for k, v in crm_q.items():
        if suffix and k in {"crm_checklist_items"}:
            continue
        row[f"{k}{suffix}"] = v

    recalculate_overall_score(row, kpi, suffix=suffix)


def refresh_crm_scores_after_stage_metrics(rows: List[Dict[str, Any]], kpi: Dict[str, Any], kpi_cmp: Optional[Dict[str, Any]] = None) -> None:
    for row in rows:
        crm_q = evaluate_crm_checklist(row, include_stage=True)
        row.update(crm_q)
        recalculate_overall_score(row, kpi, suffix="")
        if kpi_cmp is not None:
            row["crm_work_score_cmp"] = crm_q.get("crm_work_score")
            row["crm_checklist_percent_cmp"] = crm_q.get("crm_checklist_percent")
            recalculate_overall_score(row, kpi_cmp, suffix="_cmp")
            row["overall_score_delta"] = round(float(row.get("overall_score_cmp") or 0) - float(row.get("overall_score") or 0), 2)


def recompute_existing_row(row: Dict[str, Any], kpi: Dict[str, Any], kpi_cmp: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    r = dict(row)
    r["kpi_profile"] = (kpi.get("profile") or {}).get("name")
    r["kpi_version"] = (kpi.get("profile") or {}).get("version")
    if kpi_cmp is not None:
        r["kpi_profile_cmp"] = (kpi_cmp.get("profile") or {}).get("name")
        r["kpi_version_cmp"] = (kpi_cmp.get("profile") or {}).get("version")

    text = str(r.get("combined_transcript_text") or r.get("transcript_text") or "").strip()
    if not text and r.get("transcript_path"):
        try:
            p = Path(str(r.get("transcript_path")))
            if p.exists():
                text = p.read_text(encoding="utf-8", errors="ignore").strip()
        except Exception:
            text = ""
    bitrix_text = str(r.get("bitrix_card_transcript") or "").strip()
    analysis_text = merged_transcript_text(text, bitrix_text)
    r["combined_transcript_text"] = analysis_text
    if text and not r.get("transcript_text"):
        r["transcript_text"] = text

    if not analysis_text:
        r["call_quality_score"] = 0.0
        r["call_quality_details"] = "Качество разговора не рассчитано: нет сохраненной расшифровки."
        r.update(evaluate_crm_checklist(r, include_stage=True))
        recalculate_overall_score(r, kpi)
        return r

    for k, v in evaluate_call_text(analysis_text, kpi).items():
        r[k] = v

    r.update(evaluate_crm_checklist(r, include_stage=True))
    recalculate_overall_score(r, kpi)

    r.update(analyze_transcript_improvements(analysis_text, r))
    call_conclusion, call_recommendations = call_quality_conclusion(r)
    r["call_quality_conclusion"] = call_conclusion
    r["recommendations"] = call_recommendations
    r["conversation_meaning"] = conversation_meaning(analysis_text, r)

    if kpi_cmp is not None:
        tmp = recompute_existing_row({**r, "overall_score": None}, kpi_cmp, None)
        for key in ["call_quality_score", "deal_quality_score", "alignment_score", "crm_work_score", "overall_score"]:
            if key in tmp:
                r[f"{key}_cmp"] = tmp.get(key)
        r["overall_score_delta"] = round(float(r.get("overall_score_cmp") or 0.0) - float(r.get("overall_score") or 0.0), 2)

    return r


def reevaluate_report(args: argparse.Namespace, kpi: Dict[str, Any], kpi_cmp: Optional[Dict[str, Any]]) -> Tuple[int, Path, Path]:
    rows = load_report_json(args.reevaluate_from)
    recalculated = [recompute_existing_row(row, kpi, kpi_cmp) for row in rows]
    manager_summary = build_manager_summary(recalculated)
    manager_summary_cmp = build_manager_summary(recalculated, score_key="overall_score_cmp") if kpi_cmp is not None else None

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_out = REPORTS_DIR / f"bitnewton_reevaluated_report_{ts}.json"
    json_out.write_text(json.dumps(prepare_report_rows(recalculated), ensure_ascii=False, indent=2), encoding="utf-8")
    xlsx_out = flatten_results(recalculated, manager_summary, manager_summary_cmp=manager_summary_cmp)
    publish_latest_report(json_out, xlsx_out)
    print(f"[OK] Переоценено строк без повторной расшифровки: {len(recalculated)}")
    print(f"\nОтчет JSON: {json_out}")
    print(f"Отчет Excel: {xlsx_out}")
    print(f"Последний JSON: {LATEST_JSON_REPORT}")
    print(f"Последний Excel: {LATEST_XLSX_REPORT}")
    print(f"ИТОГО: OK={len(recalculated)} ERR=0")
    return len(recalculated), json_out, xlsx_out


def attach_transcription_to_bitrix(api: Bitrix24API, call_id: str, transcript_text: str, duration: int) -> Dict[str, Any]:
    text = (transcript_text or "").strip()
    if not text:
        raise RuntimeError("Пустая расшифровка (transcript_text)")
    duration = int(duration or 60)
    duration = max(1, duration)
    msg = {"SIDE": "User", "MESSAGE": text, "START_TIME": 0, "STOP_TIME": duration}

    variants = [str(call_id or "").strip()]
    if variants[0].startswith("VI_"):
        variants.append(variants[0][3:])

    seen = set()
    errors: List[str] = []
    for cid in variants:
        if not cid or cid in seen:
            continue
        seen.add(cid)
        try:
            res = api.call("telephony.call.attachTranscription", {"CALL_ID": cid, "MESSAGES": [msg]}).get("result", {}) or {}
            return {"call_id_used": cid, "result": res}
        except Exception as e:
            errors.append(f"{cid}: {e}")

    raise RuntimeError("Не удалось прикрепить расшифровку в Bitrix: " + " | ".join(errors))


def resolve_deal_ids(args: argparse.Namespace, api: Bitrix24API) -> List[str]:
    if args.mode == "single":
        if args.deal_id:
            deal_id = str(args.deal_id).strip()
            if not deal_id.isdigit():
                raise SystemExit(f"--deal-id должен быть числом. Сейчас: {deal_id!r}")
            return [deal_id]
        if args.deal_url:
            u = str(args.deal_url).strip()
            # Пытаемся вытащить числовой ID из URL вида .../crm/deal/details/83735/
            m = re.search(r"/crm/deal/details/(\d+)/?", u)
            if m:
                return [m.group(1)]
            # fallback: последний сегмент
            tail = u.rstrip("/").split("/")[-1]
            if tail.isdigit():
                return [tail]
            raise SystemExit(
                "Не смог распознать ID сделки из --deal-url. "
                "Ожидаю ссылку вида https://<domain>/crm/deal/details/12345/ "
                f"(получил: {u!r})."
            )
        raise SystemExit("Для --mode single нужен --deal-id или --deal-url")

    if not args.mode:
        raise SystemExit("Нужен --mode single/filter, --retry-errors-from или --reevaluate-from")

    if not args.filter_json:
        raise SystemExit("Для --mode filter нужен --filter-json")
    flt = normalize_deal_filter_dates(json.loads(Path(args.filter_json).read_text(encoding="utf-8")))
    deals = fetch_deals_by_filter(api, flt, limit=args.limit)
    print(f"Найдено сделок: {len(deals)}")
    return [str(d.get("ID")) for d in deals if d.get("ID")]


def run_sync(args: argparse.Namespace) -> Tuple[int, Path, Path]:
    kpi = load_kpi_config(args.kpi_config)
    kpi_cmp = load_kpi_config(args.kpi_config_compare) if args.kpi_config_compare else None

    if args.reevaluate_from:
        return reevaluate_report(args, kpi, kpi_cmp)

    api = Bitrix24API()
    if not api.test_connection():
        raise SystemExit(1)

    if args.use_bitnewton or args.bitnewton_flow:
        asr = env_bitnewton_asr()
        if not asr:
            raise SystemExit("Не задан BITNEWTON_TOKEN в .env")
    else:
        raise SystemExit("Нужен флаг --use-bitnewton (или --bitnewton-flow)")

    retry_scope = load_retry_scope(args.retry_errors_from) if args.retry_errors_from else None
    if retry_scope is not None:
        deal_ids = list(retry_scope.get("deal_ids") or [])
        if not deal_ids:
            raise SystemExit("В выбранном отчете нет строк с ошибками для повторного запуска")
        print(
            f"[RETRY] Повторяю только ошибки из отчета: "
            f"строк с ошибками={retry_scope.get('errors', 0)}, сделок={len(deal_ids)}",
            flush=True,
        )
    else:
        deal_ids = resolve_deal_ids(args, api)
    results: List[Dict[str, Any]] = []
    state_cache = _load_state_cache()

    audio_dir = REPORTS_DIR / "audio"
    audio_dir.mkdir(parents=True, exist_ok=True)
    ui_audio_dir = Path(args.ui_download_dir) if args.ui_download_dir else (REPORTS_DIR / "audio_ui")
    ui_audio_dir.mkdir(parents=True, exist_ok=True)
    audio_source_dirs = parse_audio_source_dirs(getattr(args, "audio_source_dir", []))
    audio_source_index = build_audio_source_index(audio_source_dirs)
    if audio_source_dirs:
        print(
            f"[AUDIO] Локальных аудиофайлов в индексе: {len(audio_source_index)}; "
            f"папки: {', '.join(str(p) for p in audio_source_dirs)}",
            flush=True,
        )
    cleanup_days = int(args.cleanup_output_days or 0)
    if cleanup_days > 0:
        removed = cleanup_old_outputs(REPORTS_DIR, keep_days=cleanup_days, extra_audio_dirs=[ui_audio_dir])
        if removed.get("total"):
            print(
                f"[OK] Автоочистка старше {cleanup_days} дней: "
                f"отчеты={removed.get('reports', 0)}, "
                f"аудио={removed.get('audio', 0)}, "
                f"расшифровки={removed.get('transcripts', 0)}",
                flush=True,
            )

    ok = 0
    err = 0
    ui_browser_session = None
    user_cache: Dict[int, Dict[str, Any]] = {}
    department_cache: Dict[int, Dict[str, Any]] = {}
    for di, deal_id in enumerate(deal_ids, 1):
        print(f"\nDEAL {di}/{len(deal_ids)}: {deal_url_from_id(args.domain, deal_id)}")
        acts_raw = list_deal_call_activities(api, deal_id)
        if args.include_call_center:
            acts = acts_raw
            call_center_acts: List[Dict[str, Any]] = []
        else:
            acts, call_center_acts = split_call_center_operator_activities(api, acts_raw, user_cache, department_cache)
        print(f"Звонков (crm.activity): {len(acts_raw)}")
        if call_center_acts:
            print(f"[SKIP] Звонков операторов Call-центра: {len(call_center_acts)}; к анализу: {len(acts)}", flush=True)

        if retry_scope is not None:
            retry_ids = retry_scope.get("activity_ids_by_deal", {}).get(str(deal_id), set())
            full_deal_retry = str(deal_id) in retry_scope.get("full_deals", set())
            if full_deal_retry:
                print("[RETRY] Ошибка была на уровне сделки, перепроверяю все звонки этой сделки", flush=True)
            else:
                before_retry = len(acts)
                acts = [a for a in acts if safe_int(a.get("ID")) in retry_ids]
                print(f"[RETRY] К повторной обработке звонков: {len(acts)} из {before_retry}", flush=True)

        max_calls_per_deal = max(0, int(getattr(args, "max_calls_per_deal", 0) or 0))
        if max_calls_per_deal and len(acts) > max_calls_per_deal:
            skipped_by_limit = len(acts) - max_calls_per_deal
            acts = acts[-max_calls_per_deal:]
            print(f"[FAST] Ограничение звонков по сделке: анализирую последние {len(acts)}, пропущено {skipped_by_limit}", flush=True)
        deal = deal_get(api, deal_id)
        comments = fetch_timeline_comments(api, deal_id)
        discipline = compute_discipline_metrics(deal, acts, kpi)
        deal_quality = compute_deal_quality(deal, comments, kpi)
        manager_id = safe_int(deal.get("ASSIGNED_BY_ID"))

        if not acts:
            row: Dict[str, Any] = {
                "deal_id": deal_id,
                "deal_url": deal_url_from_id(args.domain, deal_id),
                "stage_id": deal.get("STAGE_ID"),
                "manager_id": manager_id,
                "manager_name": None,
                "kpi_profile": (kpi.get("profile") or {}).get("name"),
                "kpi_version": (kpi.get("profile") or {}).get("version"),
                "kpi_profile_cmp": (kpi_cmp.get("profile") or {}).get("name") if kpi_cmp else None,
                "kpi_version_cmp": (kpi_cmp.get("profile") or {}).get("version") if kpi_cmp else None,
                "activity_id": None,
                "origin_id": None,
                "subject": "Звонков не найдено",
                "start_time": None,
                "end_time": None,
                "duration_minutes": None,
                "disk_file_id": None,
                "download_url": None,
                "audio_path": None,
                "bitnewton_task_id": None,
                "attach_result": None,
                "error": "По сделке не найдено звонков менеджера после исключения Call-центра" if call_center_acts else "По сделке не найдено звонков",
                "no_calls": True,
                "ignored_call_center_calls": len(call_center_acts),
            }
            row.update(discipline)
            row.update(deal_quality)
            apply_scores(row, deal, comments, "", kpi, suffix="")
            if kpi_cmp is not None:
                apply_scores(row, deal, comments, "", kpi_cmp, suffix="_cmp")
                row["overall_score_delta"] = round(float(row.get("overall_score_cmp") or 0) - float(row.get("overall_score") or 0), 2)
            row["call_quality_conclusion"] = "Оценить разговор невозможно: по сделке не найдено звонков."
            row["recommendations"] = "Проверить, был ли контакт с клиентом вне телефонии Bitrix. Если звонка не было — запланировать касание и зафиксировать следующий шаг в CRM."
            results.append(row)
            err += 1
            print(f"[NO CALLS] OK={ok} ERR={err}", flush=True)
            continue

        for ai, act in enumerate(acts, 1):
            row: Dict[str, Any] = {
                "deal_id": deal_id,
                "deal_url": deal_url_from_id(args.domain, deal_id),
                "stage_id": deal.get("STAGE_ID"),
                "manager_id": manager_id,
                "manager_name": None,
                "kpi_profile": (kpi.get("profile") or {}).get("name"),
                "kpi_version": (kpi.get("profile") or {}).get("version"),
                "kpi_profile_cmp": (kpi_cmp.get("profile") or {}).get("name") if kpi_cmp else None,
                "kpi_version_cmp": (kpi_cmp.get("profile") or {}).get("version") if kpi_cmp else None,
                "activity_id": act.get("ID"),
                "origin_id": act.get("ORIGIN_ID"),
                "subject": act.get("SUBJECT"),
                "start_time": act.get("START_TIME"),
                "end_time": act.get("END_TIME"),
                "duration_minutes": guess_duration_minutes(act),
                "disk_file_id": None,
                "download_url": None,
                "audio_path": None,
                "bitnewton_task_id": None,
                "attach_result": None,
                "error": None,
                "ignored_call_center_calls": len(call_center_acts),
            }
            row.update(discipline)
            row.update(deal_quality)

            try:
                call_id = str(row["origin_id"] or "")
                if not call_id:
                    raise RuntimeError("Нет ORIGIN_ID (CALL_ID) для резолва записи")

                if not bool(getattr(args, "no_reuse_transcripts", False)):
                    cached_text, cached_path = load_cached_transcript(state_cache, call_id, deal_id, row.get("activity_id"))
                    if cached_text and cached_path:
                        row["bitnewton_task_id"] = "cache"
                        row["transcript_path"] = str(cached_path)
                        row["transcript_text"] = cached_text
                        row["transcript_excerpt"] = cached_text[:1200]
                        row["transcript_hash"] = _sha256_text(cached_text)
                        row["bitrix_card_transcript_status"] = "Не запрашивалась: использована сохранённая расшифровка"
                        finalize_transcript_analysis(row, deal, comments, cached_text, "", kpi, kpi_cmp)
                        if args.force_attach:
                            act_full = activity_get(api, int(act.get("ID"))) if act.get("ID") else act
                            attach = attach_transcription_to_bitrix(api, call_id=call_id, transcript_text=cached_text, duration=guess_duration_sec(act_full))
                            row["attach_result"] = attach
                        else:
                            row["attach_result"] = {"skipped": True, "reason": "cached_transcript_reused"}
                        state_cache[call_id] = {
                            "hash": row["transcript_hash"],
                            "transcript_path": str(cached_path),
                            "updated_at": datetime.now().isoformat(timespec="seconds"),
                            "deal_id": deal_id,
                            "activity_id": row.get("activity_id"),
                            "source": "cache",
                        }
                        _save_state_cache(state_cache)
                        ok += 1
                        print(f"[CACHE] Использую сохранённую расшифровку: activity_id={row.get('activity_id')}", flush=True)
                        continue

                rr = resolve_call_recording(api, call_id=call_id, activity_id=safe_int(act.get("ID")))
                row["disk_file_id"] = rr.disk_file_id
                row["recording_diagnostics"] = rr.diagnostics
                candidates = rr.candidates
                row["download_url"] = candidates[0] if candidates else None

                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                local_source = find_local_audio_source(
                    audio_source_index,
                    deal_id=deal_id,
                    activity_id=act.get("ID"),
                    disk_file_id=row.get("disk_file_id"),
                    call_id=call_id,
                    disk_file_name=(rr.diagnostics or {}).get("disk_file_name"),
                )
                out_suffix = local_source.suffix if local_source else ".mp3"
                out_path = audio_dir / f"deal{deal_id}_act{act.get('ID')}_{row.get('disk_file_id')}_{ts}{out_suffix}"
                if local_source:
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(local_source, out_path)
                    row["local_audio_source_used"] = True
                    row["local_audio_source_path"] = str(local_source)
                    print(f"[AUDIO] Использую локальный файл: activity_id={row.get('activity_id')} file={local_source}", flush=True)
                else:
                    row["local_audio_source_used"] = False
                    if not candidates:
                        raise RuntimeError("Не нашёл URL кандидаты для скачивания (resolver не смог), и локальный аудиофайл тоже не найден.")
                    rest_timeout_sec = max(5, int(getattr(args, "rest_timeout_sec", 20) or 20))
                    dl = download_best_effort(candidates=candidates, out_path=out_path, timeout_sec=rest_timeout_sec, retries=0)
                    if not dl.ok or not dl.path:
                        row["download_attempts"] = [a.__dict__ for a in dl.attempts]
                        if args.ui_download:
                            try:
                                from ui_audio_downloader import (
                                    UiBrowserSession,
                                )

                                ui_res = None
                                ui_errors: List[str] = []
                                mode = str(getattr(args, "ui_download_mode", "auto") or "auto")
                                browser = str(getattr(args, "ui_browser", "chrome"))
                                ui_timeout_sec = max(5, int(getattr(args, "ui_timeout_sec", 20) or 20))
                                browser_profile_directory = str(getattr(args, "browser_profile_directory", "Default") or "Default")
                                if ui_browser_session is None:
                                    ui_browser_session = UiBrowserSession(
                                        downloads_dir=ui_audio_dir,
                                        chrome_profile_dir=args.chrome_profile_dir,
                                        browser=browser,
                                        browser_profile_directory=browser_profile_directory,
                                    )

                                if mode in {"direct", "auto"}:
                                    for candidate in [html.unescape(c).strip() for c in candidates if isinstance(c, str) and c.startswith("http")]:
                                        print(f"[UI] Пробую ссылку активности через {browser}, таймаут {ui_timeout_sec} сек.: {candidate}", flush=True)
                                        ui_res = ui_browser_session.download_url(
                                            candidate,
                                            timeout_sec=ui_timeout_sec,
                                            referer_url=row["deal_url"],
                                        )
                                        if ui_res.ok and ui_res.path:
                                            break
                                        ui_errors.append(f"url={candidate}: {ui_res.error if ui_res else 'no result'}")

                                if (ui_res is None or not ui_res.ok) and mode in {"timeline", "auto"}:
                                    print(f"[UI] Пробую таймлайн сделки через {browser}, таймаут {ui_timeout_sec} сек.: activity_id={row.get('activity_id')}", flush=True)
                                    ui_res = ui_browser_session.download_call_from_deal_timeline(
                                        row["deal_url"],
                                        int(row.get("activity_id") or 0),
                                        timeout_sec=ui_timeout_sec,
                                    )
                                    if not ui_res.ok:
                                        ui_errors.append(f"timeline activity_id={row.get('activity_id')}: {ui_res.error}")

                                row["ui_download_errors"] = ui_errors
                                if ui_res is None or not ui_res.ok or not ui_res.path:
                                    raise RuntimeError("; ".join(ui_errors) or (ui_res.error if ui_res else "UI download failed"))
                                out_path.parent.mkdir(parents=True, exist_ok=True)
                                out_path.write_bytes(Path(ui_res.path).read_bytes())
                                row["ui_download_used"] = True
                                row["ui_download_path"] = str(ui_res.path)
                            except Exception as e:
                                raise RuntimeError(f"Не удалось скачать аудио (REST+UI): {dl.error}; UI: {e}")
                        else:
                            raise RuntimeError(f"Не удалось скачать аудио: {dl.error}")
                row["audio_path"] = str(out_path)
                row["audio_size_bytes"] = int(out_path.stat().st_size) if out_path.exists() else None
                bad = validate_audio_file(out_path)
                if bad:
                    raise RuntimeError(f"Скачанный файл не похож на аудио: {bad}")

                # ASR
                task = asr.start_transcribing(str(out_path), diarize=bool(args.diarize), remove_timestamps=True)
                row["bitnewton_task_id"] = task.task_id
                text = asr.wait_and_get_text(task.task_id, timeout_sec=1800)
                transcript_path = save_transcript_file(
                    deal_id=str(deal_id),
                    activity_id=row.get("activity_id"),
                    task_id=task.task_id,
                    text=text or "",
                )
                row["transcript_path"] = str(transcript_path)
                row["transcript_text"] = text or ""
                row["transcript_excerpt"] = (text or "")[:1200]

                bitrix_text = ""
                if args.fetch_bitrix_card_transcript and args.ui_download:
                    try:
                        from ui_audio_downloader import UiBrowserSession

                        browser = str(getattr(args, "ui_browser", "chrome"))
                        ui_timeout_sec = max(5, int(getattr(args, "ui_timeout_sec", 20) or 20))
                        browser_profile_directory = str(getattr(args, "browser_profile_directory", "Default") or "Default")
                        if ui_browser_session is None:
                            ui_browser_session = UiBrowserSession(
                                downloads_dir=ui_audio_dir,
                                chrome_profile_dir=args.chrome_profile_dir,
                                browser=browser,
                                browser_profile_directory=browser_profile_directory,
                            )
                        print(f"[UI] Пробую прочитать расшифровку из карточки Bitrix: activity_id={row.get('activity_id')}", flush=True)
                        tr_res = ui_browser_session.fetch_transcript_from_deal_timeline(
                            row["deal_url"],
                            int(row.get("activity_id") or 0),
                            timeout_sec=ui_timeout_sec,
                        )
                        if tr_res.ok and tr_res.text:
                            bitrix_text = tr_res.text
                            row["bitrix_card_transcript"] = bitrix_text
                            row["bitrix_card_transcript_status"] = "Получена"
                            row["transcript_match_score"] = transcript_match_score(text or "", bitrix_text)
                        else:
                            row["bitrix_card_transcript_status"] = tr_res.error or "Расшифровка Bitrix не найдена"
                    except Exception as e:
                        row["bitrix_card_transcript_status"] = f"Не удалось прочитать расшифровку Bitrix: {e}"
                else:
                    row["bitrix_card_transcript_status"] = "Не запрашивалась"

                finalize_transcript_analysis(row, deal, comments, text or "", bitrix_text, kpi, kpi_cmp)

                # attach (idempotent via cache)
                txt_hash = _sha256_text(text or "")
                row["transcript_hash"] = txt_hash
                cached = (state_cache.get(call_id) or {}) if isinstance(state_cache.get(call_id), dict) else None
                if (not args.force_attach) and cached and cached.get("hash") == txt_hash:
                    row["attach_result"] = {"skipped": True, "reason": "state_cache_same_hash"}
                else:
                    act_full = activity_get(api, int(act.get("ID"))) if act.get("ID") else act
                    attach = attach_transcription_to_bitrix(api, call_id=call_id, transcript_text=text, duration=guess_duration_sec(act_full))
                    row["attach_result"] = attach
                state_cache[call_id] = {
                    "hash": txt_hash,
                    "transcript_path": str(transcript_path),
                    "bitnewton_task_id": task.task_id,
                    "updated_at": datetime.now().isoformat(timespec="seconds"),
                    "deal_id": deal_id,
                    "activity_id": row.get("activity_id"),
                    "source": "bitnewton",
                }
                _save_state_cache(state_cache)
                ok += 1
            except (BitNewtonError, requests.RequestException, Exception) as e:
                row["error"] = str(e)
                err += 1
            finally:
                if row.get("audio_path") and not args.download_audio:
                    try:
                        Path(str(row["audio_path"])).unlink(missing_ok=True)  # type: ignore[call-arg]
                        row["audio_path"] = None
                    except Exception:
                        pass
                results.append(row)
                print(f"[{ai}/{len(acts)}] OK={ok} ERR={err}", flush=True)

    if ui_browser_session is not None:
        ui_browser_session.close()

    _save_state_cache(state_cache)

    manager_ids = [int(r["manager_id"]) for r in results if isinstance(r.get("manager_id"), int)]
    names = user_name_map(api, manager_ids) if manager_ids else {}
    for r in results:
        mid = r.get("manager_id")
        if isinstance(mid, int):
            r["manager_name"] = names.get(mid, str(mid))
        if r.get("overall_score") is None:
            r["overall_score"] = 0.0

    stage_ids_for_map = [str(r.get("stage_id") or "") for r in results]
    stage_history_by_deal: Dict[str, List[Dict[str, Any]]] = {}
    try:
        unique_deal_ids = sorted({str(deal_id_from_report_row(r) or "") for r in results if deal_id_from_report_row(r)})
        if unique_deal_ids:
            print("[STAGE] Загружаю историю перемещений сделок по стадиям", flush=True)
            stage_history_by_deal = fetch_stage_history_by_deals(api, unique_deal_ids)
            for items in stage_history_by_deal.values():
                stage_ids_for_map.extend(str(item.get("STAGE_ID") or "") for item in items if isinstance(item, dict))
    except Exception as e:
        print(f"[WARN] Не удалось загрузить историю стадий: {e}", flush=True)

    stage_map = fetch_stage_name_map(api, stage_ids_for_map)
    if stage_history_by_deal:
        attach_stage_history_metrics(results, stage_history_by_deal, stage_map=stage_map)
    refresh_crm_scores_after_stage_metrics(results, kpi, kpi_cmp)

    final_results = results
    if retry_scope is not None:
        source_rows = list(retry_scope.get("source_rows") or [])
        final_results = merge_retry_results(source_rows, results, retry_scope)
        print(
            f"[RETRY] Пересобираю полный отчет: исходных строк={len(source_rows)}, "
            f"повторно обработано={len(results)}, итоговых строк={len(final_results)}",
            flush=True,
        )

    manager_summary = build_manager_summary(final_results)
    manager_summary_cmp = build_manager_summary(final_results, score_key="overall_score_cmp") if kpi_cmp is not None else None

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_results = prepare_report_rows(final_results, stage_map=stage_map)
    json_out = REPORTS_DIR / f"bitnewton_sync_report_{ts}.json"
    json_out.write_text(json.dumps(report_results, ensure_ascii=False, indent=2), encoding="utf-8")
    xlsx_out = flatten_results(report_results, manager_summary, manager_summary_cmp=manager_summary_cmp, stage_map=stage_map)
    publish_latest_report(json_out, xlsx_out)
    print(f"\nОтчет JSON: {json_out}")
    print(f"Отчет Excel: {xlsx_out}")
    print(f"Последний JSON: {LATEST_JSON_REPORT}")
    print(f"Последний Excel: {LATEST_XLSX_REPORT}")
    if kpi_cmp is not None:
        ranked = sorted(
            [r for r in final_results if r.get("overall_score_delta") is not None],
            key=lambda x: abs(float(x.get("overall_score_delta") or 0.0)),
            reverse=True,
        )[:5]
        if ranked:
            print("\nТоп-5 кейсов с максимальной разницей KPI:")
            for i, r in enumerate(ranked, 1):
                print(
                    f"{i}. deal={r.get('deal_id')} act={r.get('activity_id')} "
                    f"manager={r.get('manager_name') or r.get('manager_id')} "
                    f"base={r.get('overall_score')} cmp={r.get('overall_score_cmp')} "
                    f"delta={r.get('overall_score_delta')}"
                )
    print(f"ИТОГО: OK={ok} ERR={err}")

    if int(args.cleanup_chrome_tmp_days or 0) > 0:
        removed = cleanup_old_chrome_tmp_profiles(REPORTS_DIR, keep_days=int(args.cleanup_chrome_tmp_days))
        if removed:
            print(f"[OK] Удалено старых chrome_profile_tmp_*: {removed}")

    return ok, json_out, xlsx_out
