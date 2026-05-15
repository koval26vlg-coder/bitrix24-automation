from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from pipelines.paths import LATEST_JSON_REPORT, LATEST_XLSX_REPORT, REPORTS_DIR
from pipelines.scoring import _clean_text_for_report, evaluate_crm_checklist, quality_label
from pipelines.stages import safe_int, stage_display_name, stage_order_map

from logging_setup import get_logger

logger = get_logger(__name__)


def publish_latest_report(json_path: Path, xlsx_path: Path) -> None:
    try:
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        if json_path.exists():
            shutil.copy2(json_path, LATEST_JSON_REPORT)
        if xlsx_path.exists():
            shutil.copy2(xlsx_path, LATEST_XLSX_REPORT)
    except Exception as e:
        logger.warning(f"[WARN] Не удалось обновить latest-отчет: {e}")


RU_COLUMNS: Dict[str, str] = {
    "deal_id": "ID сделки",
    "deal_url": "Ссылка на сделку",
    "stage_id": "Стадия",
    "stage_name": "Стадия",
    "stage_history_event_id": "ID события стадии",
    "stage_history_event_type": "Тип события стадии",
    "stage_history_created_at": "Дата смены стадии",
    "stage_history_count": "Кол-во переходов по стадиям",
    "stage_history_path": "Маршрут стадий",
    "stage_duration_minutes": "Время до следующей стадии, мин.",
    "stage_duration_hours": "Время до следующей стадии, ч.",
    "stage_is_current": "Текущая стадия",
    "stage_last_change_at": "Последняя смена стадии",
    "stage_current_age_minutes": "Время в текущей стадии, мин.",
    "stage_current_age_days": "Время в текущей стадии, дн.",
    "stage_warning_threshold": "Порог предупреждения по стадии",
    "stage_critical_threshold": "Порог тревоги по стадии",
    "stage_current_age_status": "Статус срока в стадии",
    "deal_total_work_minutes": "Общее время в работе, мин.",
    "deal_total_work_days": "Общее время в работе, дн.",
    "deal_total_work_warning_threshold": "Порог предупреждения по сделке",
    "deal_total_work_critical_threshold": "Порог тревоги по сделке",
    "deal_total_work_status": "Статус общего срока сделки",
    "stage_return_count": "Возвраты на предыдущие стадии",
    "stage_reached_final": "Достигнута финальная стадия",
    "stage_movement_risk": "Риск движения по воронке",
    "stage_movement_recommendation": "Рекомендация по движению сделки",
    "stage_deals_total": "Всего сделок",
    "stage_deals_passed": "Сделок прошло стадию",
    "stage_sr": "SR стадии, %",
    "stage_order": "Порядок стадии",
    "manager_id": "ID менеджера",
    "manager_name": "Менеджер",
    "kpi_profile": "Профиль KPI",
    "kpi_profile_ru": "Профиль KPI",
    "kpi_explanation": "Расшифровка KPI",
    "kpi_version": "Версия KPI",
    "kpi_profile_cmp": "Профиль KPI сравнения",
    "kpi_version_cmp": "Версия KPI сравнения",
    "activity_id": "ID активности",
    "origin_id": "ID звонка Bitrix",
    "subject": "Тема звонка",
    "duration_minutes": "Длительность звонка, мин.",
    "disk_file_id": "ID файла записи",
    "download_url": "Ссылка на запись",
    "audio_path": "Путь к аудио",
    "local_audio_source_used": "Использован локальный источник аудио",
    "local_audio_source_path": "Локальный источник аудио",
    "bitnewton_task_id": "ID задачи Bit.Newton",
    "asr_status": "Статус Bit.Newton ASR",
    "asr_skipped": "ASR пропущена",
    "asr_skipped_calls": "Звонков без новой ASR",
    "attach_result": "Результат прикрепления в Bitrix",
    "error": "Ошибка",
    "calls_count": "Кол-во звонков по сделке",
    "first_response_minutes": "Время до первого звонка, мин.",
    "reaction_speed_label": "Скорость реакции менеджера",
    "first_response_sla_ok": "Первый контакт в срок",
    "first_response_explanation": "Что значит первый контакт",
    "deal_quality_score": "Оценка заполнения сделки",
    "deal_quality_details": "Критерии заполнения сделки",
    "has_contact": "Есть контакт/компания",
    "has_amount": "Указана сумма",
    "has_title": "Есть название",
    "has_comments": "Есть комментарии/следующие шаги",
    "recording_diagnostics": "Диагностика записи",
    "download_attempts": "Попытки REST-скачивания",
    "ui_download_errors": "Ошибки UI-скачивания",
    "ui_download_used": "Использовано UI-скачивание",
    "ui_download_path": "Путь UI-скачивания",
    "audio_size_bytes": "Размер аудио, байт",
    "transcript_excerpt": "Фрагмент расшифровки",
    "transcript_text": "Полная расшифровка",
    "transcript_marked": "Расшифровка с пометками",
    "transcript_path": "Файл расшифровки",
    "transcript_hash": "Хеш расшифровки",
    "call_quality_score": "Оценка качества разговора",
    "call_quality_details": "Критерии качества разговора",
    "call_checklist_total_score": "Баллы по чек-листу звонка",
    "call_checklist_max_score": "Максимум чек-листа звонка",
    "call_checklist_percent": "Чек-лист звонка, %",
    "call_checklist_block_details": "Оценка этапов продаж",
    "checklist_block_name": "Этап продаж",
    "checklist_criterion": "Критерий чек-листа",
    "checklist_score": "Оценка критерия",
    "checklist_max_score": "Максимум критерия",
    "checklist_evidence": "Доказательство из расшифровки",
    "checklist_comment": "Комментарий оценки",
    "checklist_code": "Код критерия",
    "sales_stage_block_name": "Этап продаж",
    "sales_stage_score": "Баллы этапа",
    "sales_stage_max_score": "Максимум этапа",
    "sales_stage_percent": "Выполнение этапа, %",
    "sales_stage_missing": "Что усилить по этапу",
    "manager_stage_calls": "Звонков с оценкой этапа",
    "manager_stage_deals": "Сделок с оценкой этапа",
    "manager_stage_avg_percent": "Среднее выполнение этапа, %",
    "manager_stage_weak_calls": "Звонков ниже нормы",
    "manager_stage_weak_rate": "Доля слабых звонков, %",
    "criterion_calls": "Звонков с критерием",
    "criterion_avg_score": "Средняя оценка критерия",
    "criterion_completion_percent": "Выполнение критерия, %",
    "criterion_failed_count": "Провалов критерия",
    "criterion_partial_count": "Частично выполнено",
    "criterion_fail_rate": "Доля проблем, %",
    "training_recommendation": "Рекомендация для обучения",
    "top_checklist_gaps": "Главные провалы чек-листа",
    "has_greeting": "Есть приветствие",
    "has_needs_discovery": "Выявлены потребности",
    "has_objection_work": "Есть работа с возражениями",
    "has_next_step_phrase": "Озвучен следующий шаг",
    "alignment_score": "Связь звонка с CRM",
    "crm_work_score": "Оценка ведения CRM",
    "alignment_details": "Что значит связь звонка с CRM",
    "title_mentions": "Совпадения с названием сделки",
    "amount_mentioned": "Сумма упомянута в разговоре",
    "next_step_synced": "Следующий шаг синхронизирован с CRM",
    "next_step_synced_details": "Что дает синхронизация следующего шага",
    "crm_checklist_total_score": "Баллы по CRM-чек-листу",
    "crm_checklist_max_score": "Максимум CRM-чек-листа",
    "crm_checklist_percent": "CRM-чек-лист, %",
    "crm_checklist_details": "Критерии ведения CRM",
    "crm_checklist_block_name": "Блок CRM",
    "crm_checklist_criterion": "Критерий CRM",
    "crm_checklist_score": "Оценка CRM-критерия",
    "crm_checklist_max_score": "Максимум CRM-критерия",
    "crm_checklist_comment": "Комментарий CRM-оценки",
    "crm_checklist_code": "Код CRM-критерия",
    "crm_criterion_deals": "Сделок с критерием",
    "crm_criterion_avg_score": "Средняя оценка CRM-критерия",
    "crm_criterion_completion_percent": "Выполнение CRM-критерия, %",
    "crm_criterion_failed_count": "Провалов CRM-критерия",
    "crm_criterion_partial_count": "Частично выполнено CRM",
    "crm_criterion_fail_rate": "Доля CRM-проблем, %",
    "control_priority": "Приоритет контроля",
    "control_priority_order": "Порядок приоритета",
    "control_reason": "Причина контроля",
    "control_next_action": "Что сделать дальше",
    "quality_control_score": "Оценка риска контроля",
    "training_source": "Источник обучения",
    "training_topic": "Тема обучения",
    "coaching_priority": "Приоритет обучения",
    "coaching_metric": "Метрика провала",
    "coaching_affected_count": "Сколько кейсов затронуто",
    "summary_section": "Раздел",
    "summary_metric": "Показатель",
    "summary_value": "Значение",
    "summary_comment": "Комментарий",
    "manager_rank": "Место",
    "manager_focus": "Фокус руководителя",
    "manager_critical_deals": "Критичных сделок",
    "manager_high_deals": "Высокий приоритет сделок",
    "manager_training_topics": "Тем обучения",
    "manager_top_training_topic": "Главная тема обучения",
    "manager_next_action": "Следующее действие руководителя",
    "overall_score": "Итоговая оценка",
    "overall_score_details": "Из чего складывается итоговая оценка",
    "overall_score_cmp": "Итоговая оценка сравнения",
    "overall_score_delta": "Разница итоговой оценки",
    "call_quality_score_cmp": "Качество разговора сравнение",
    "deal_quality_score_cmp": "Заполнение сделки сравнение",
    "alignment_score_cmp": "Связь звонка с CRM сравнение",
    "call_quality_conclusion": "Вывод по разговору",
    "client_work_conclusion": "Вывод по проработке клиента",
    "improvement_moments": "Моменты для улучшения",
    "objections_count": "Кол-во возражений",
    "unhandled_objections_count": "Кол-во неотработанных возражений",
    "objections_handled": "Возражения отработаны",
    "objections_found": "Найденные возражения",
    "unhandled_objections": "Неотработанные возражения",
    "objection_recommendations": "Варианты отработки возражений",
    "objection_status": "Статус отработки",
    "objection_fragment": "Фрагмент с возражением",
    "recommendations": "Рекомендации",
    "calls_total": "Всего звонков",
    "calls_ok": "Успешно обработано",
    "overall_score_sum": "Сумма итоговых оценок",
    "growth_deal_data": "Зона роста: данные CRM",
    "growth_call_structure": "Зона роста: структура разговора",
    "growth_alignment": "Зона роста: связь с CRM",
    "avg_overall_score": "Средняя итоговая оценка",
    "top_growth_zones": "Главные зоны роста",
    "avg_call_quality_score": "Средняя оценка разговоров",
    "client_work_score": "Оценка проработки клиента",
    "client_work_quality": "Качество проработки клиента",
    "conversation_quality": "Качество разговоров",
    "transcripts_count": "Кол-во расшифровок",
    "transcript_paths": "Файлы расшифровок",
    "no_calls": "Нет звонков по сделке",
    "deals_without_calls": "Сделок без звонков",
    "deals_total": "Всего сделок",
    "calls_failed": "Ошибок по звонкам",
    "ignored_call_center_calls": "Исключено звонков Call-центра",
    "skipped_short_calls": "Исключено коротких звонков",
    "calls_breakdown": "Звонки внутри сделки",
    "transcripts_combined": "Все расшифровки по сделке",
    "bitrix_card_transcript": "Расшифровка из карточки Bitrix",
    "bitrix_card_transcript_status": "Статус расшифровки Bitrix",
    "transcript_match_score": "Совпадение Bit.Newton/Bitrix, %",
    "conversation_meaning": "Смысл разговора",
    "improvement_moments_combined": "Общие моменты для улучшения",
    "objections_combined": "Возражения по сделке",
    "call_number": "Номер звонка в сделке",
    "call_summary": "Кратко по звонку",
    "call_has_error": "Ошибка по звонку",
    "deal_row_type": "Тип строки",
    "lost_deal_url": "Ссылка на проигранную сделку",
    "lost_deal_title": "Название проигранной сделки",
    "lost_stage_name": "Стадия проигрыша",
    "lost_manager_name": "Менеджер",
    "lost_amount": "Сумма потери",
    "lost_date_create": "Дата создания",
    "lost_close_date": "Дата проигрыша",
    "lost_lifetime_days": "Срок жизни сделки, дн.",
    "lost_source": "Источник",
    "lost_analysis_basis": "Основа анализа отказа",
    "loss_reason_category": "Типичная причина отказа",
    "loss_reason_confidence": "Уверенность причины",
    "loss_reason_evidence": "Фрагмент/основание причины",
    "lost_deals_count": "Кол-во проигранных сделок",
    "lost_deals_share": "Доля проигрышей, %",
    "lost_avg_lifetime_days": "Средний срок жизни, дн.",
    "lost_top_managers": "Менеджеры с частыми потерями",
    "conversion_tools": "Инструменты для роста конверсии",
    "conversion_next_action": "Что внедрить дальше",
    "conversion_priority": "Приоритет внедрения",
    "conversion_rank": "Порядок внедрения",
    "conversion_expected_effect": "Ожидаемый эффект",
}


KPI_PROFILE_DESCRIPTIONS: Dict[str, Tuple[str, str]] = {
    "example": ("Базовый контроль звонка и CRM", "Сбалансированная оценка разговора, заполнения сделки, дисциплины касаний и связи звонка с CRM."),
    "default": ("Базовый контроль звонка и CRM", "Сбалансированная оценка разговора, заполнения сделки, дисциплины касаний и связи звонка с CRM."),
    "sales": ("Продажи: стандартный контроль менеджера", "Основной рабочий профиль для продаж: выявление потребностей, следующий шаг, работа с возражениями, SLA и связь с CRM."),
    "sales_soft": ("Продажи: мягкая оценка диалога", "Больше веса качеству разговора и следующему шагу, мягче требования к скорости первого контакта и паузам между касаниями."),
    "sales_strict": ("Продажи: строгая дисциплина и результат", "Жесткий контроль скорости реакции, регулярности касаний, выявления потребностей, возражений и фиксации следующего шага."),
}


def build_deal_conclusions(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("deal_id") or "unknown"), []).append(row)

    out: List[Dict[str, Any]] = []
    for deal_id, deal_rows in grouped.items():
        ok_rows = [r for r in deal_rows if not r.get("error")]
        scored = ok_rows or deal_rows
        calls_total = len(deal_rows)
        calls_ok = len(ok_rows)
        avg_overall = round(sum(float(r.get("overall_score") or 0.0) for r in scored) / max(1, len(scored)), 2)
        avg_call = round(sum(float(r.get("call_quality_score") or 0.0) for r in scored) / max(1, len(scored)), 2)
        deal_quality = max(float(r.get("deal_quality_score") or 0.0) for r in scored) if scored else 0.0
        needs_ratio = sum(1 for r in scored if r.get("has_needs_discovery")) / max(1, len(scored))
        next_step_ratio = sum(1 for r in scored if r.get("has_next_step_phrase")) / max(1, len(scored))
        objection_ratio = sum(1 for r in scored if r.get("has_objection_work")) / max(1, len(scored))
        synced_ratio = sum(1 for r in scored if r.get("next_step_synced")) / max(1, len(scored))
        avg_crm_work = round(sum(float(r.get("crm_work_score") or 0.0) for r in scored) / max(1, len(scored)), 2)

        client_work_score = round(
            avg_crm_work if avg_crm_work else (
                0.60 * deal_quality
                + 0.40 * synced_ratio * 100
            ),
            2,
        )

        issues: List[str] = []
        if deal_quality < 100:
            issues.append("не полностью заполнена сделка")
        if needs_ratio < 0.5:
            issues.append("слабо выявлены потребности")
        if next_step_ratio < 0.5:
            issues.append("не фиксируется следующий шаг")
        if synced_ratio < 0.5:
            issues.append("следующий шаг плохо синхронизирован с CRM")
        if objection_ratio == 0 and calls_total > 0:
            issues.append("работа с возражениями не прослеживается")

        first = deal_rows[0]
        if client_work_score >= 80:
            client_conclusion = "Клиент проработан качественно: CRM и разговоры дают понятную картину дальнейших действий."
        elif client_work_score >= 60:
            client_conclusion = "Проработка клиента нормальная, но есть зоны для усиления."
        elif client_work_score >= 40:
            client_conclusion = "Проработка клиента слабая: часть важных элементов не закреплена в разговоре или CRM."
        else:
            client_conclusion = "Проработка клиента критически слабая: не хватает структуры, фиксации потребностей и следующего шага."

        if avg_call >= 80:
            conversation_conclusion = "Качество разговоров сильное."
        elif avg_call >= 60:
            conversation_conclusion = "Качество разговоров нормальное, но нестабильное."
        elif avg_call >= 40:
            conversation_conclusion = "Качество разговоров слабое: менеджеру нужна более четкая структура диалога."
        else:
            conversation_conclusion = "Качество разговоров критически слабое."

        recommendations = "Усилить: " + ", ".join(issues) + "." if issues else "Сохранять текущий подход и фиксировать следующий шаг после каждого контакта."
        transcript_paths = [str(r.get("transcript_path")) for r in deal_rows if r.get("transcript_path")]

        out.append(
            {
                "deal_url": first.get("deal_url"),
                "stage_name": first.get("stage_name") or stage_display_name(first.get("stage_id")),
                "manager_name": first.get("manager_name"),
                "calls_total": calls_total,
                "calls_ok": calls_ok,
                "transcripts_count": len(transcript_paths),
                "avg_overall_score": avg_overall,
                "avg_call_quality_score": avg_call,
                "client_work_score": client_work_score,
                "client_work_quality": quality_label(client_work_score),
                "conversation_quality": quality_label(avg_call),
                "client_work_conclusion": client_conclusion,
                "call_quality_conclusion": conversation_conclusion,
                "recommendations": recommendations,
                "transcript_paths": "\n".join(transcript_paths),
            }
        )

    return sorted(out, key=lambda x: float(x.get("client_work_score") or 0.0))


def _deal_group_key(row: Dict[str, Any]) -> str:
    return str(row.get("deal_id") or row.get("deal_url") or "unknown")


def _join_unique(values: List[Any], sep: str = "\n") -> str:
    seen: set[str] = set()
    out: List[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return sep.join(out)


def _call_sort_key(row: Dict[str, Any]) -> Tuple[str, str]:
    return (str(row.get("start_time") or ""), str(row.get("activity_id") or ""))


def _call_heading(row: Dict[str, Any], number: int) -> str:
    subject = str(row.get("subject") or "Звонок").strip()
    duration = row.get("duration_minutes")
    duration_text = f"{duration:g} мин." if isinstance(duration, (int, float)) else ""
    if duration_text:
        return f"Звонок {number}: {subject} ({duration_text})"
    return f"Звонок {number}: {subject}"


def _short_error(text: Any, limit: int = 550) -> str:
    raw = _clean_text_for_report(str(text or ""))
    if not raw:
        return ""
    if "не дождался входа в Bitrix" in raw or "oauth/authorize" in raw:
        return "Не удалось открыть Bitrix через выбранный профиль браузера: требуется вход в Bitrix или выбран не тот профиль Edge/Chrome."
    if "UI download timeout" in raw:
        return "Не удалось скачать запись через интерфейс Bitrix: файл не появился в папке загрузок."
    if "download failed" in raw:
        return "Не удалось скачать запись звонка через REST/UI."
    return raw[:limit] + ("..." if len(raw) > limit else "")


def build_deal_report_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        first = deal_rows[0]
        no_call_rows = [r for r in deal_rows if r.get("no_calls")]
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        ok_calls = [r for r in call_rows if not r.get("error") and not r.get("asr_skipped")]
        failed_calls = [r for r in call_rows if r.get("error")]
        asr_skipped_calls = len([r for r in call_rows if r.get("asr_skipped")])
        scored = ok_calls

        calls_total = len(call_rows)
        calls_ok = len(ok_calls)
        calls_failed = len(failed_calls)
        skipped_short_calls = max(int(r.get("skipped_short_calls") or 0) for r in deal_rows) if deal_rows else 0
        deal_quality = max(float(r.get("deal_quality_score") or 0.0) for r in deal_rows) if deal_rows else 0.0
        if scored:
            avg_overall = round(sum(float(r.get("overall_score") or 0.0) for r in scored) / len(scored), 2)
            avg_call = round(sum(float(r.get("call_quality_score") or 0.0) for r in scored) / len(scored), 2)
            avg_checklist = round(sum(float(r.get("call_checklist_percent") or r.get("call_quality_score") or 0.0) for r in scored) / len(scored), 2)
            avg_alignment = round(sum(float(r.get("alignment_score") or 0.0) for r in scored) / len(scored), 2)
            avg_crm_work = round(sum(float(r.get("crm_work_score") or 0.0) for r in scored) / len(scored), 2)
            avg_crm_checklist = round(sum(float(r.get("crm_checklist_percent") or r.get("crm_work_score") or 0.0) for r in scored) / len(scored), 2)
            needs_ratio = sum(1 for r in scored if r.get("has_needs_discovery")) / len(scored)
            next_step_ratio = sum(1 for r in scored if r.get("has_next_step_phrase")) / len(scored)
            objection_ratio = sum(1 for r in scored if r.get("has_objection_work")) / len(scored)
            synced_ratio = sum(1 for r in scored if r.get("next_step_synced")) / len(scored)
        else:
            avg_overall = 0.0
            avg_call = 0.0
            avg_checklist = 0.0
            avg_alignment = 0.0
            avg_crm_work = 0.0
            avg_crm_checklist = 0.0
            needs_ratio = 0.0
            next_step_ratio = 0.0
            objection_ratio = 0.0
            synced_ratio = 0.0

        client_work_score = round(
            avg_crm_work if avg_crm_work else (
                0.60 * deal_quality
                + 0.40 * synced_ratio * 100
            ),
            2,
        ) if scored else 0.0

        issues: List[str] = []
        if no_call_rows or calls_total == 0:
            issues.append("по сделке не найдено звонков")
        if calls_failed:
            issues.append(f"не обработано звонков: {calls_failed}")
        if asr_skipped_calls:
            issues.append(f"без новой расшифровки Bit.Newton: {asr_skipped_calls}")
        if deal_quality < 100:
            issues.append("не полностью заполнена сделка")
        if scored and needs_ratio < 0.5:
            issues.append("слабо выявлены потребности")
        if scored and next_step_ratio < 0.5:
            issues.append("не фиксируется следующий шаг")
        if scored and synced_ratio < 0.5:
            issues.append("следующий шаг плохо синхронизирован с CRM")
        if scored and objection_ratio == 0:
            issues.append("работа с возражениями не прослеживается")

        if not scored:
            client_conclusion = "По сделке нет обработанных звонков: оценка клиента невозможна до появления записи или успешной расшифровки."
            conversation_conclusion = "Нет обработанных разговоров для оценки."
        elif client_work_score >= 80:
            client_conclusion = "Клиент проработан качественно: CRM и разговоры дают понятную картину дальнейших действий."
            conversation_conclusion = "Качество разговоров сильное." if avg_call >= 80 else "Разговоры в целом рабочие, но есть отдельные зоны для усиления."
        elif client_work_score >= 60:
            client_conclusion = "Проработка клиента нормальная, но есть зоны для усиления."
            conversation_conclusion = "Качество разговоров нормальное, но нестабильное." if avg_call >= 60 else "Качество разговоров слабое: менеджеру нужна более четкая структура диалога."
        elif client_work_score >= 40:
            client_conclusion = "Проработка клиента слабая: часть важных элементов не закреплена в разговоре или CRM."
            conversation_conclusion = "Качество разговоров слабое: менеджеру нужна более четкая структура диалога."
        else:
            client_conclusion = "Проработка клиента критически слабая: не хватает структуры, фиксации потребностей и следующего шага."
            conversation_conclusion = "Качество разговоров критически слабое."

        call_blocks: List[str] = []
        transcript_blocks: List[str] = []
        for index, row in enumerate(call_rows, start=1):
            heading = _call_heading(row, index)
            if row.get("error"):
                call_blocks.append(f"{heading}\nСтатус: ошибка обработки — {row.get('error')}")
                continue
            call_blocks.append(
                "\n".join(
                    [
                        heading,
                        f"Качество разговора: {row.get('call_quality_score') or 0}",
                        f"Потребности: {'да' if row.get('has_needs_discovery') else 'нет'}; "
                        f"возражения: {'да' if row.get('has_objection_work') else 'нет'}; "
                        f"следующий шаг: {'да' if row.get('has_next_step_phrase') else 'нет'}",
                        str(row.get("improvement_moments") or "").strip(),
                    ]
                ).strip()
            )
            transcript = str(row.get("transcript_marked") or row.get("transcript_text") or "").strip()
            if transcript:
                transcript_blocks.append(f"{heading}\n{transcript}")

        if no_call_rows and not call_blocks:
            call_blocks.append("Звонков по сделке не найдено.")

        out.append(
            {
                "deal_url": first.get("deal_url"),
                "stage_name": first.get("stage_name") or stage_display_name(first.get("stage_id")),
                "manager_name": first.get("manager_name"),
                "kpi_profile_ru": first.get("kpi_profile_ru"),
                "kpi_explanation": first.get("kpi_explanation"),
                "calls_total": calls_total,
                "calls_ok": calls_ok,
                "calls_failed": calls_failed,
                "asr_skipped_calls": asr_skipped_calls,
                "ignored_call_center_calls": int(first.get("ignored_call_center_calls") or 0),
                "skipped_short_calls": skipped_short_calls,
                "no_calls": bool(no_call_rows or calls_total == 0),
                "transcripts_count": len([r for r in ok_calls if r.get("transcript_path")]),
                "avg_overall_score": avg_overall,
                "overall_score_details": _join_unique([r.get("overall_score_details") for r in ok_calls]) or first.get("overall_score_details"),
                "avg_call_quality_score": avg_call,
                "call_quality_details": _join_unique([r.get("call_quality_details") for r in ok_calls]),
                "call_checklist_percent": avg_checklist,
                "call_checklist_block_details": _join_unique([r.get("call_checklist_block_details") for r in ok_calls]),
                "deal_quality_score": deal_quality,
                "deal_quality_details": first.get("deal_quality_details"),
                "alignment_score": avg_alignment,
                "crm_work_score": client_work_score,
                "crm_checklist_percent": avg_crm_checklist,
                "crm_checklist_details": _join_unique([r.get("crm_checklist_details") for r in ok_calls]),
                "alignment_details": _join_unique([r.get("alignment_details") for r in ok_calls]),
                "stage_history_count": first.get("stage_history_count"),
                "stage_history_path": first.get("stage_history_path"),
                "stage_last_change_at": first.get("stage_last_change_at"),
                "stage_current_age_minutes": first.get("stage_current_age_minutes"),
                "stage_current_age_days": first.get("stage_current_age_days"),
                "stage_warning_threshold": first.get("stage_warning_threshold"),
                "stage_critical_threshold": first.get("stage_critical_threshold"),
                "stage_current_age_status": first.get("stage_current_age_status"),
                "deal_total_work_minutes": first.get("deal_total_work_minutes"),
                "deal_total_work_days": first.get("deal_total_work_days"),
                "deal_total_work_warning_threshold": first.get("deal_total_work_warning_threshold"),
                "deal_total_work_critical_threshold": first.get("deal_total_work_critical_threshold"),
                "deal_total_work_status": first.get("deal_total_work_status"),
                "stage_return_count": first.get("stage_return_count"),
                "stage_reached_final": first.get("stage_reached_final"),
                "stage_movement_risk": first.get("stage_movement_risk"),
                "stage_movement_recommendation": first.get("stage_movement_recommendation"),
                "client_work_score": client_work_score,
                "client_work_quality": quality_label(client_work_score) if scored else "Нет данных",
                "conversation_quality": quality_label(avg_call) if scored else "Нет данных",
                "client_work_conclusion": client_conclusion,
                "call_quality_conclusion": conversation_conclusion,
                "recommendations": "Усилить: " + ", ".join(issues) + "." if issues else "Сохранять текущий подход и фиксировать следующий шаг после каждого контакта.",
                "calls_breakdown": "\n\n".join(call_blocks),
                "transcripts_combined": "\n\n---\n\n".join(transcript_blocks),
                "conversation_meaning": _join_unique([r.get("conversation_meaning") for r in ok_calls], sep="\n\n"),
                "improvement_moments_combined": _join_unique([r.get("improvement_moments") for r in ok_calls]),
                "objections_combined": _join_unique([r.get("objections_found") for r in ok_calls]),
                "unhandled_objections": _join_unique([r.get("unhandled_objections") for r in ok_calls]),
                "objection_recommendations": _join_unique([r.get("objection_recommendations") for r in ok_calls]),
                "transcript_paths": _join_unique([r.get("transcript_path") for r in ok_calls]),
                "error": _short_error(_join_unique([r.get("error") for r in deal_rows if r.get("error")])),
            }
        )

    return sorted(out, key=lambda x: (str(x.get("manager_name") or ""), str(x.get("deal_url") or "")))


def build_call_detail_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        no_call_rows = [r for r in deal_rows if r.get("no_calls")]
        first = deal_rows[0]
        if not call_rows and no_call_rows:
            out.append(
                {
                    "deal_url": first.get("deal_url"),
                    "call_number": "",
                    "subject": "Звонков не найдено",
                    "duration_minutes": "",
                    "call_has_error": True,
                    "error": _short_error(first.get("error") or "По сделке не найдено звонков"),
                }
            )
            continue
        for index, row in enumerate(call_rows, start=1):
            out.append(
                {
                    "deal_url": row.get("deal_url"),
                    "call_number": index,
                    "subject": row.get("subject"),
                    "duration_minutes": row.get("duration_minutes"),
                    "call_quality_score": row.get("call_quality_score"),
                    "call_quality_details": row.get("call_quality_details"),
                    "call_checklist_total_score": row.get("call_checklist_total_score"),
                    "call_checklist_max_score": row.get("call_checklist_max_score"),
                    "call_checklist_percent": row.get("call_checklist_percent"),
                    "call_checklist_block_details": row.get("call_checklist_block_details"),
                    "has_greeting": row.get("has_greeting"),
                    "has_needs_discovery": row.get("has_needs_discovery"),
                    "has_objection_work": row.get("has_objection_work"),
                    "has_next_step_phrase": row.get("has_next_step_phrase"),
                    "alignment_score": row.get("alignment_score"),
                    "alignment_details": row.get("alignment_details"),
                    "next_step_synced": row.get("next_step_synced"),
                    "next_step_synced_details": row.get("next_step_synced_details"),
                    "call_quality_conclusion": row.get("call_quality_conclusion"),
                    "conversation_meaning": row.get("conversation_meaning"),
                    "improvement_moments": row.get("improvement_moments"),
                    "unhandled_objections": row.get("unhandled_objections"),
                    "objection_recommendations": row.get("objection_recommendations"),
                    "transcript_path": row.get("transcript_path"),
                    "transcript_marked": row.get("transcript_marked"),
                    "transcript_text": row.get("transcript_text"),
                    "bitrix_card_transcript_status": row.get("bitrix_card_transcript_status"),
                    "transcript_match_score": row.get("transcript_match_score"),
                    "bitrix_card_transcript": row.get("bitrix_card_transcript"),
                    "call_has_error": bool(row.get("error")),
                    "error": _short_error(row.get("error")),
                }
            )
    return out


def build_objection_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        for objection in row.get("objection_rows") or []:
            if not isinstance(objection, dict):
                continue
            out.append(
                {
                    "deal_url": row.get("deal_url"),
                    "stage_name": row.get("stage_name") or stage_display_name(row.get("stage_id")),
                    "manager_name": row.get("manager_name"),
                    "subject": row.get("subject"),
                    "objection_fragment": objection.get("objection_fragment"),
                    "objection_status": objection.get("objection_status"),
                    "objection_recommendations": objection.get("objection_recommendation"),
                }
            )
    return out


def build_call_checklist_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        for index, row in enumerate(call_rows, start=1):
            for item in row.get("call_checklist_items") or []:
                if not isinstance(item, dict):
                    continue
                out.append(
                    {
                        "deal_url": row.get("deal_url"),
                        "stage_name": row.get("stage_name") or stage_display_name(row.get("stage_id")),
                        "manager_name": row.get("manager_name"),
                        "call_number": index,
                        "subject": row.get("subject"),
                        "duration_minutes": row.get("duration_minutes"),
                        "checklist_block_name": item.get("checklist_block_name"),
                        "checklist_criterion": item.get("checklist_criterion"),
                        "checklist_score": item.get("checklist_score"),
                        "checklist_max_score": item.get("checklist_max_score"),
                        "checklist_evidence": item.get("checklist_evidence"),
                        "checklist_comment": item.get("checklist_comment"),
                        "checklist_code": item.get("checklist_code"),
                    }
                )
    return out


def build_sales_stage_score_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_rows = sorted(deal_rows, key=_call_sort_key)
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        for index, row in enumerate(call_rows, start=1):
            for block in row.get("call_checklist_blocks") or []:
                if not isinstance(block, dict):
                    continue
                out.append(
                    {
                        "deal_url": row.get("deal_url"),
                        "stage_name": row.get("stage_name") or stage_display_name(row.get("stage_id")),
                        "manager_name": row.get("manager_name"),
                        "call_number": index,
                        "subject": row.get("subject"),
                        "duration_minutes": row.get("duration_minutes"),
                        "sales_stage_block_name": block.get("sales_stage_block_name"),
                        "sales_stage_score": block.get("sales_stage_score"),
                        "sales_stage_max_score": block.get("sales_stage_max_score"),
                        "sales_stage_percent": block.get("sales_stage_percent"),
                        "sales_stage_missing": block.get("sales_stage_missing"),
                    }
                )
    return out


def _training_recommendation(block_name: Any, criterion: Any) -> str:
    block = str(block_name or "").lower()
    crit = str(criterion or "").strip()
    if "контакт" in block:
        return f"Отработать старт звонка: {crit}. Сделать короткий обязательный скрипт первых 20 секунд."
    if "потребност" in block:
        return f"Отработать выявление потребности: {crit}. Добавить 3-5 обязательных уточняющих вопросов до презентации."
    if "презентац" in block:
        return f"Усилить презентацию: {crit}. Привязывать предложение к задаче клиента и говорить языком выгоды."
    if "возраж" in block:
        return f"Отработать возражения: {crit}. Использовать схему: признать, уточнить причину, дать аргумент, закрепить следующий шаг."
    if "закрытие" in block:
        return f"Усилить закрытие звонка: {crit}. Фиксировать договоренность, срок и следующий контакт в разговоре и CRM."
    if "заполнение" in block:
        return f"Навести порядок в карточке сделки: {crit}. Сделать поле обязательным или добавить контроль перед переводом стадии."
    if "активность" in block:
        return f"Проверить дисциплину касаний: {crit}. Звонок менеджера должен быть виден в CRM после исключения Call-центра."
    if "связь" in block:
        return f"Синхронизировать разговор и CRM: {crit}. После звонка фиксировать следующий шаг и ключевые договоренности в карточке."
    if "движение" in block:
        return f"Контролировать движение по воронке: {crit}. Проверить зависшие стадии, возвраты и актуальность текущей стадии."
    return f"Разобрать на обучении: {crit}. Проверить фрагменты звонков с низкой оценкой."


def build_manager_stage_summary_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        if row.get("no_calls") or row.get("error"):
            continue
        manager = str(row.get("manager_name") or row.get("manager_id") or "Без менеджера")
        deal_key = _deal_group_key(row)
        for block in row.get("call_checklist_blocks") or []:
            if not isinstance(block, dict):
                continue
            block_name = str(block.get("sales_stage_block_name") or "")
            if not block_name:
                continue
            key = (manager, block_name)
            item = agg.setdefault(
                key,
                {
                    "manager_name": manager,
                    "sales_stage_block_name": block_name,
                    "manager_stage_calls": 0,
                    "_deals": set(),
                    "_percent_sum": 0.0,
                    "manager_stage_weak_calls": 0,
                },
            )
            percent = float(block.get("sales_stage_percent") or 0.0)
            item["manager_stage_calls"] += 1
            item["_percent_sum"] += percent
            item["_deals"].add(deal_key)
            if percent < 70:
                item["manager_stage_weak_calls"] += 1

    out: List[Dict[str, Any]] = []
    for item in agg.values():
        calls = max(1, int(item["manager_stage_calls"]))
        weak = int(item["manager_stage_weak_calls"])
        out.append(
            {
                "manager_name": item["manager_name"],
                "sales_stage_block_name": item["sales_stage_block_name"],
                "manager_stage_calls": calls,
                "manager_stage_deals": len(item["_deals"]),
                "manager_stage_avg_percent": round(float(item["_percent_sum"]) / calls, 2),
                "manager_stage_weak_calls": weak,
                "manager_stage_weak_rate": round(weak * 100.0 / calls, 2),
            }
        )
    return sorted(out, key=lambda r: (float(r.get("manager_stage_avg_percent") or 0.0), str(r.get("manager_name") or "")))


def build_manager_criterion_gap_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    agg: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for row in rows:
        if row.get("no_calls") or row.get("error"):
            continue
        manager = str(row.get("manager_name") or row.get("manager_id") or "Без менеджера")
        for item in row.get("call_checklist_items") or []:
            if not isinstance(item, dict):
                continue
            block_name = str(item.get("checklist_block_name") or "")
            criterion = str(item.get("checklist_criterion") or "")
            if not criterion:
                continue
            key = (manager, block_name, criterion)
            bucket = agg.setdefault(
                key,
                {
                    "manager_name": manager,
                    "checklist_block_name": block_name,
                    "checklist_criterion": criterion,
                    "criterion_calls": 0,
                    "_score_sum": 0.0,
                    "criterion_failed_count": 0,
                    "criterion_partial_count": 0,
                    "training_recommendation": _training_recommendation(block_name, criterion),
                },
            )
            score = float(item.get("checklist_score") or 0.0)
            bucket["criterion_calls"] += 1
            bucket["_score_sum"] += score
            if score <= 0:
                bucket["criterion_failed_count"] += 1
            elif score < 1:
                bucket["criterion_partial_count"] += 1

    out: List[Dict[str, Any]] = []
    for bucket in agg.values():
        calls = max(1, int(bucket["criterion_calls"]))
        avg_score = float(bucket["_score_sum"]) / calls
        failed = int(bucket["criterion_failed_count"])
        partial = int(bucket["criterion_partial_count"])
        out.append(
            {
                "manager_name": bucket["manager_name"],
                "checklist_block_name": bucket["checklist_block_name"],
                "checklist_criterion": bucket["checklist_criterion"],
                "criterion_calls": calls,
                "criterion_avg_score": round(avg_score, 2),
                "criterion_completion_percent": round(avg_score * 100.0, 2),
                "criterion_failed_count": failed,
                "criterion_partial_count": partial,
                "criterion_fail_rate": round((failed + partial) * 100.0 / calls, 2),
                "training_recommendation": bucket["training_recommendation"],
            }
        )
    return sorted(
        out,
        key=lambda r: (
            float(r.get("criterion_completion_percent") or 0.0),
            -float(r.get("criterion_fail_rate") or 0.0),
            str(r.get("manager_name") or ""),
        ),
    )


def _aggregate_deal_crm_row(deal_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    sorted_rows = sorted(deal_rows, key=_call_sort_key)
    first = dict(sorted_rows[0])
    call_rows = [r for r in sorted_rows if not r.get("no_calls")]
    ok_calls = [r for r in call_rows if not r.get("error")]
    first["no_calls"] = bool(not call_rows or any(r.get("no_calls") for r in sorted_rows))
    first["calls_count"] = max(int(first.get("calls_count") or 0), len(call_rows))
    first["has_next_step_phrase"] = any(r.get("has_next_step_phrase") for r in ok_calls)
    first["next_step_synced"] = any(r.get("next_step_synced") for r in ok_calls)
    first["amount_mentioned"] = any(r.get("amount_mentioned") for r in ok_calls)
    first["crm_checklist_items"] = evaluate_crm_checklist(first, include_stage=True).get("crm_checklist_items") or []
    return first


def build_crm_checklist_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    out: List[Dict[str, Any]] = []
    for _, deal_rows in grouped.items():
        deal_row = _aggregate_deal_crm_row(deal_rows)
        for item in deal_row.get("crm_checklist_items") or []:
            if not isinstance(item, dict):
                continue
            out.append(
                {
                    "deal_url": deal_row.get("deal_url"),
                    "stage_name": deal_row.get("stage_name") or stage_display_name(deal_row.get("stage_id")),
                    "manager_name": deal_row.get("manager_name"),
                    "crm_checklist_block_name": item.get("crm_checklist_block_name"),
                    "crm_checklist_criterion": item.get("crm_checklist_criterion"),
                    "crm_checklist_score": item.get("crm_checklist_score"),
                    "crm_checklist_max_score": item.get("crm_checklist_max_score"),
                    "crm_checklist_comment": item.get("crm_checklist_comment"),
                    "crm_checklist_code": item.get("crm_checklist_code"),
                }
            )
    return out


def build_manager_crm_gap_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    agg: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for _, deal_rows in grouped.items():
        deal_row = _aggregate_deal_crm_row(deal_rows)
        manager = str(deal_row.get("manager_name") or deal_row.get("manager_id") or "Без менеджера")
        for item in deal_row.get("crm_checklist_items") or []:
            if not isinstance(item, dict):
                continue
            block_name = str(item.get("crm_checklist_block_name") or "")
            criterion = str(item.get("crm_checklist_criterion") or "")
            if not criterion:
                continue
            key = (manager, block_name, criterion)
            bucket = agg.setdefault(
                key,
                {
                    "manager_name": manager,
                    "crm_checklist_block_name": block_name,
                    "crm_checklist_criterion": criterion,
                    "crm_criterion_deals": 0,
                    "_score_sum": 0.0,
                    "crm_criterion_failed_count": 0,
                    "crm_criterion_partial_count": 0,
                    "training_recommendation": _training_recommendation(block_name, criterion),
                },
            )
            score = float(item.get("crm_checklist_score") or 0.0)
            bucket["crm_criterion_deals"] += 1
            bucket["_score_sum"] += score
            if score <= 0:
                bucket["crm_criterion_failed_count"] += 1
            elif score < 1:
                bucket["crm_criterion_partial_count"] += 1

    out: List[Dict[str, Any]] = []
    for bucket in agg.values():
        deals = max(1, int(bucket["crm_criterion_deals"]))
        avg_score = float(bucket["_score_sum"]) / deals
        failed = int(bucket["crm_criterion_failed_count"])
        partial = int(bucket["crm_criterion_partial_count"])
        out.append(
            {
                "manager_name": bucket["manager_name"],
                "crm_checklist_block_name": bucket["crm_checklist_block_name"],
                "crm_checklist_criterion": bucket["crm_checklist_criterion"],
                "crm_criterion_deals": deals,
                "crm_criterion_avg_score": round(avg_score, 2),
                "crm_criterion_completion_percent": round(avg_score * 100.0, 2),
                "crm_criterion_failed_count": failed,
                "crm_criterion_partial_count": partial,
                "crm_criterion_fail_rate": round((failed + partial) * 100.0 / deals, 2),
                "training_recommendation": bucket["training_recommendation"],
            }
        )
    return sorted(
        out,
        key=lambda r: (
            float(r.get("crm_criterion_completion_percent") or 0.0),
            -float(r.get("crm_criterion_fail_rate") or 0.0),
            str(r.get("manager_name") or ""),
        ),
    )


def _control_priority(score: float, reasons: List[str]) -> Tuple[int, str]:
    reason_text = " ".join(reasons).lower()
    if score >= 80 or "нет звонков" in reason_text or "тревога" in reason_text:
        return 1, "Критично"
    if score >= 55:
        return 2, "Высокий"
    if score >= 30:
        return 3, "Средний"
    return 4, "Наблюдать"


def _control_next_action(row: Dict[str, Any], reasons: List[str]) -> str:
    text = " ".join(reasons).lower()
    if "нет звонков" in text:
        return "Проверить, почему по сделке нет звонка менеджера; назначить контакт с клиентом и зафиксировать следующий шаг в CRM."
    if "ошиб" in text:
        return "Запустить режим «Повторить только ошибки», затем проверить, доступна ли запись звонка и корректно ли работает источник аудио."
    if "crm" in text or "следующий шаг" in text:
        return "Открыть карточку сделки, заполнить недостающие поля, зафиксировать договоренности и следующий шаг после разговора."
    if "разговор" in text or "возраж" in text:
        return "Разобрать расшифровку звонка с менеджером и закрепить конкретный сценарий: потребность, аргумент, следующий шаг."
    if "стади" in text or "воронк" in text:
        return "Проверить актуальность стадии, причину зависания и необходимость возврата/перевода сделки."
    return "Проверить сделку выборочно и оставить в мониторинге до следующего запуска отчета."


def build_quality_control_rows(deal_report_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in deal_report_rows:
        overall = float(row.get("avg_overall_score") or 0.0)
        call_score = float(row.get("avg_call_quality_score") or 0.0)
        crm_score = float(row.get("crm_checklist_percent") or row.get("crm_work_score") or 0.0)
        calls_total = int(row.get("calls_total") or 0)
        calls_failed = int(row.get("calls_failed") or 0)
        no_calls = bool(row.get("no_calls"))
        risk = str(row.get("stage_movement_risk") or "").strip()
        unhandled = str(row.get("unhandled_objections") or "").strip()
        reasons: List[str] = []
        risk_score = 0.0
        if no_calls or calls_total == 0:
            reasons.append("нет звонков менеджера")
            risk_score += 35
        if calls_failed:
            reasons.append(f"ошибки обработки звонков: {calls_failed}")
            risk_score += min(25, 8 * calls_failed)
        if overall < 40:
            reasons.append(f"низкая итоговая оценка: {overall:g}")
            risk_score += 25
        elif overall < 60:
            reasons.append(f"итоговая оценка требует контроля: {overall:g}")
            risk_score += 12
        if call_score < 50:
            reasons.append(f"слабое качество разговора: {call_score:g}")
            risk_score += 18
        if crm_score < 50:
            reasons.append(f"слабое ведение CRM: {crm_score:g}")
            risk_score += 18
        if risk and risk not in {"OK", "Финал"}:
            reasons.append(f"риск движения по воронке: {risk}")
            risk_score += 20 if "Тревога" in risk else 10
        if unhandled:
            reasons.append("есть неотработанные возражения")
            risk_score += 10
        if not reasons:
            reasons.append("критичных отклонений не найдено")

        order, priority = _control_priority(risk_score, reasons)
        out.append(
            {
                "control_priority_order": order,
                "control_priority": priority,
                "quality_control_score": round(min(100.0, risk_score), 2),
                "deal_url": row.get("deal_url"),
                "stage_name": row.get("stage_name"),
                "manager_name": row.get("manager_name"),
                "avg_overall_score": overall,
                "avg_call_quality_score": call_score,
                "crm_checklist_percent": crm_score,
                "calls_total": calls_total,
                "calls_failed": calls_failed,
                "stage_movement_risk": risk,
                "control_reason": "; ".join(reasons),
                "control_next_action": _control_next_action(row, reasons),
                "recommendations": row.get("recommendations"),
            }
        )
    return sorted(out, key=lambda r: (int(r.get("control_priority_order") or 9), -float(r.get("quality_control_score") or 0), str(r.get("manager_name") or "")))


def _coaching_priority(completion: float, fail_rate: float) -> str:
    if completion < 50 or fail_rate >= 60:
        return "Критично"
    if completion < 70 or fail_rate >= 35:
        return "Высокий"
    return "Средний"


def build_coaching_plan_rows(
    manager_criterion_gap_rows: List[Dict[str, Any]],
    manager_crm_gap_rows: List[Dict[str, Any]],
    manager_stage_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []

    for row in manager_criterion_gap_rows:
        completion = float(row.get("criterion_completion_percent") or 0.0)
        fail_rate = float(row.get("criterion_fail_rate") or 0.0)
        if completion >= 75 and fail_rate < 30:
            continue
        out.append(
            {
                "manager_name": row.get("manager_name"),
                "training_source": "Звонки",
                "training_topic": f"{row.get('checklist_block_name')}: {row.get('checklist_criterion')}",
                "coaching_priority": _coaching_priority(completion, fail_rate),
                "coaching_metric": f"выполнение {completion:g}%, проблем {fail_rate:g}%",
                "coaching_affected_count": row.get("criterion_calls"),
                "training_recommendation": row.get("training_recommendation"),
            }
        )

    for row in manager_crm_gap_rows:
        completion = float(row.get("crm_criterion_completion_percent") or 0.0)
        fail_rate = float(row.get("crm_criterion_fail_rate") or 0.0)
        if completion >= 80 and fail_rate < 25:
            continue
        out.append(
            {
                "manager_name": row.get("manager_name"),
                "training_source": "CRM",
                "training_topic": f"{row.get('crm_checklist_block_name')}: {row.get('crm_checklist_criterion')}",
                "coaching_priority": _coaching_priority(completion, fail_rate),
                "coaching_metric": f"выполнение {completion:g}%, проблем {fail_rate:g}%",
                "coaching_affected_count": row.get("crm_criterion_deals"),
                "training_recommendation": row.get("training_recommendation"),
            }
        )

    for row in manager_stage_rows:
        completion = float(row.get("manager_stage_avg_percent") or 0.0)
        weak_rate = float(row.get("manager_stage_weak_rate") or 0.0)
        if completion >= 75 and weak_rate < 30:
            continue
        stage = row.get("sales_stage_block_name")
        out.append(
            {
                "manager_name": row.get("manager_name"),
                "training_source": "Этап продаж",
                "training_topic": f"Этап: {stage}",
                "coaching_priority": _coaching_priority(completion, weak_rate),
                "coaching_metric": f"среднее выполнение {completion:g}%, слабых звонков {weak_rate:g}%",
                "coaching_affected_count": row.get("manager_stage_calls"),
                "training_recommendation": f"Провести разбор этапа «{stage}» на примерах звонков менеджера и закрепить короткий рабочий сценарий.",
            }
        )

    priority_order = {"Критично": 1, "Высокий": 2, "Средний": 3}
    return sorted(
        out,
        key=lambda r: (
            priority_order.get(str(r.get("coaching_priority") or ""), 9),
            str(r.get("manager_name") or ""),
            str(r.get("training_source") or ""),
        ),
    )


def _avg_numeric(rows: List[Dict[str, Any]], key: str) -> float:
    values = [float(row.get(key) or 0.0) for row in rows if row.get(key) is not None]
    return round(sum(values) / max(1, len(values)), 2)


def _count_by(rows: List[Dict[str, Any]], key: str, value: str) -> int:
    return sum(1 for row in rows if str(row.get(key) or "") == value)


def build_executive_summary_rows(
    deal_report_rows: List[Dict[str, Any]],
    manager_summary: List[Dict[str, Any]],
    quality_control_rows: List[Dict[str, Any]],
    manager_criterion_gap_rows: List[Dict[str, Any]],
    manager_crm_gap_rows: List[Dict[str, Any]],
    coaching_plan_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    def add(section: str, metric: str, value: Any, comment: str = "") -> None:
        rows.append(
            {
                "summary_section": section,
                "summary_metric": metric,
                "summary_value": value,
                "summary_comment": comment,
            }
        )

    deals_total = len(deal_report_rows)
    calls_total = sum(int(row.get("calls_total") or 0) for row in deal_report_rows)
    calls_ok = sum(int(row.get("calls_ok") or 0) for row in deal_report_rows)
    calls_failed = sum(int(row.get("calls_failed") or 0) for row in deal_report_rows)
    asr_skipped_calls = sum(int(row.get("asr_skipped_calls") or 0) for row in deal_report_rows)
    skipped_short_calls = sum(int(row.get("skipped_short_calls") or 0) for row in deal_report_rows)
    deals_without_calls = sum(1 for row in deal_report_rows if row.get("no_calls"))
    add("Общее", "Сделок в отчете", deals_total)
    add("Общее", "Звонков в сделках", calls_total)
    add("Общее", "Успешно обработано звонков", calls_ok)
    add("Общее", "Ошибок обработки звонков", calls_failed, "Используйте режим «Повторить только ошибки».")
    add("Общее", "Звонков без новой ASR", asr_skipped_calls, "Bit.Newton недоступен или токен не принят; кэш расшифровок используется, где он есть.")
    add("Общее", "Исключено коротких звонков", skipped_short_calls, "Технические дозвоны короче минимальной длительности не отправляются в Bit.Newton.")
    add("Общее", "Сделок без звонков менеджера", deals_without_calls, "Call-центр исключен из оценки менеджера.")
    add("Оценки", "Средняя итоговая оценка", _avg_numeric(deal_report_rows, "avg_overall_score"))
    add("Оценки", "Средняя оценка разговора", _avg_numeric(deal_report_rows, "avg_call_quality_score"))
    add("Оценки", "Средняя оценка CRM", _avg_numeric(deal_report_rows, "crm_checklist_percent"))
    add("Контроль", "Критичных сделок", _count_by(quality_control_rows, "control_priority", "Критично"))
    add("Контроль", "Сделок высокого приоритета", _count_by(quality_control_rows, "control_priority", "Высокий"))
    add("Контроль", "Сделок среднего приоритета", _count_by(quality_control_rows, "control_priority", "Средний"))

    for index, row in enumerate(quality_control_rows[:5], start=1):
        add(
            "Топ контроля",
            f"{index}. {row.get('manager_name') or ''}",
            row.get("deal_url"),
            str(row.get("control_reason") or ""),
        )

    for index, row in enumerate(manager_criterion_gap_rows[:5], start=1):
        add(
            "Провалы звонков",
            f"{index}. {row.get('manager_name') or ''}",
            f"{row.get('criterion_completion_percent')}%",
            f"{row.get('checklist_block_name')}: {row.get('checklist_criterion')}",
        )

    for index, row in enumerate(manager_crm_gap_rows[:5], start=1):
        add(
            "Провалы CRM",
            f"{index}. {row.get('manager_name') or ''}",
            f"{row.get('crm_criterion_completion_percent')}%",
            f"{row.get('crm_checklist_block_name')}: {row.get('crm_checklist_criterion')}",
        )

    for index, row in enumerate(coaching_plan_rows[:5], start=1):
        add(
            "Обучение",
            f"{index}. {row.get('manager_name') or ''}",
            row.get("coaching_priority"),
            f"{row.get('training_source')}: {row.get('training_topic')}",
        )

    return rows


def build_manager_scorecard_rows(
    manager_summary: List[Dict[str, Any]],
    quality_control_rows: List[Dict[str, Any]],
    coaching_plan_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    control_by_manager: Dict[str, List[Dict[str, Any]]] = {}
    for row in quality_control_rows:
        control_by_manager.setdefault(str(row.get("manager_name") or "Без менеджера"), []).append(row)

    coaching_by_manager: Dict[str, List[Dict[str, Any]]] = {}
    for row in coaching_plan_rows:
        coaching_by_manager.setdefault(str(row.get("manager_name") or "Без менеджера"), []).append(row)

    out: List[Dict[str, Any]] = []
    sorted_managers = sorted(manager_summary, key=lambda r: float(r.get("avg_overall_score") or 0.0))
    for rank, row in enumerate(sorted_managers, start=1):
        manager = str(row.get("manager_name") or "Без менеджера")
        control = control_by_manager.get(manager, [])
        coaching = coaching_by_manager.get(manager, [])
        critical = _count_by(control, "control_priority", "Критично")
        high = _count_by(control, "control_priority", "Высокий")
        top_training = coaching[0] if coaching else {}
        score = float(row.get("avg_overall_score") or 0.0)
        if critical:
            focus = "Срочный контроль сделок"
            next_action = "Начать с критичных сделок из листа «Контроль качества», затем провести точечный разбор звонков."
        elif high:
            focus = "Плановый контроль сделок"
            next_action = "Проверить сделки высокого приоритета и закрепить правила фиксации следующего шага."
        elif coaching:
            focus = "Обучение по стабильности качества"
            next_action = "Провести короткое обучение по главной теме из плана обучения."
        elif score < 70:
            focus = "Наблюдение"
            next_action = "Проверить несколько свежих сделок и сравнить динамику на следующем отчете."
        else:
            focus = "Поддерживать уровень"
            next_action = "Использовать лучшие звонки менеджера как примеры для команды."
        out.append(
            {
                "manager_rank": rank,
                "manager_name": manager,
                "avg_overall_score": row.get("avg_overall_score"),
                "deals_total": row.get("deals_total"),
                "calls_total": row.get("calls_total"),
                "calls_ok": row.get("calls_ok"),
                "deals_without_calls": row.get("deals_without_calls"),
                "manager_critical_deals": critical,
                "manager_high_deals": high,
                "manager_training_topics": len(coaching),
                "manager_top_training_topic": top_training.get("training_topic") or "",
                "top_growth_zones": row.get("top_growth_zones"),
                "top_checklist_gaps": row.get("top_checklist_gaps"),
                "manager_focus": focus,
                "manager_next_action": next_action,
            }
        )
    return out


def _unique_deal_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)
    return [sorted(deal_rows, key=_call_sort_key)[0] for deal_rows in grouped.values()]


def build_stage_history_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for deal_row in _unique_deal_rows(rows):
        for item in deal_row.get("stage_history_items") or []:
            if not isinstance(item, dict):
                continue
            out.append(
                {
                    "deal_url": deal_row.get("deal_url"),
                    "manager_name": deal_row.get("manager_name"),
                    "stage_history_event_id": item.get("stage_history_event_id"),
                    "stage_history_event_type": item.get("stage_history_event_type"),
                    "stage_history_created_at": item.get("stage_history_created_at"),
                    "stage_id": item.get("stage_id"),
                    "stage_name": item.get("stage_name"),
                    "stage_order": item.get("stage_order"),
                    "stage_duration_minutes": item.get("stage_duration_minutes"),
                    "stage_duration_hours": item.get("stage_duration_hours"),
                    "stage_is_current": item.get("stage_is_current"),
                }
            )
    return sorted(out, key=lambda r: (str(r.get("deal_url") or ""), str(r.get("stage_history_created_at") or "")))


def build_stage_sr_rows(rows: List[Dict[str, Any]], stage_map: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    deal_rows = _unique_deal_rows(rows)
    deals_total = len(deal_rows)
    stage_deals: Dict[str, set[str]] = {}
    ranks = stage_order_map(stage_map)

    for deal_row in deal_rows:
        deal_key = _deal_group_key(deal_row)
        seen_stages: set[str] = set()
        for item in deal_row.get("stage_history_items") or []:
            if not isinstance(item, dict):
                continue
            stage_id = str(item.get("stage_id") or "").strip()
            if stage_id:
                seen_stages.add(stage_id)
        if not seen_stages and deal_row.get("stage_id"):
            seen_stages.add(str(deal_row.get("stage_id")))
        for stage_id in seen_stages:
            stage_deals.setdefault(stage_id, set()).add(deal_key)

    out: List[Dict[str, Any]] = []
    for stage_id, deal_keys in stage_deals.items():
        passed = len(deal_keys)
        out.append(
            {
                "stage_id": stage_id,
                "stage_name": stage_display_name(stage_id, stage_map=stage_map),
                "stage_order": ranks.get(stage_id),
                "stage_deals_total": deals_total,
                "stage_deals_passed": passed,
                "stage_sr": round((passed / max(1, deals_total)) * 100.0, 2),
            }
        )
    return sorted(out, key=lambda r: (safe_int(r.get("stage_order")) or 9999, str(r.get("stage_name") or "")))


def build_stage_movement_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for deal_row in _unique_deal_rows(rows):
        out.append(
            {
                "deal_url": deal_row.get("deal_url"),
                "stage_name": deal_row.get("stage_name") or stage_display_name(deal_row.get("stage_id")),
                "manager_name": deal_row.get("manager_name"),
                "stage_history_count": deal_row.get("stage_history_count"),
                "stage_history_path": deal_row.get("stage_history_path"),
                "stage_last_change_at": deal_row.get("stage_last_change_at"),
                "stage_current_age_minutes": deal_row.get("stage_current_age_minutes"),
                "stage_current_age_days": deal_row.get("stage_current_age_days"),
                "stage_warning_threshold": deal_row.get("stage_warning_threshold"),
                "stage_critical_threshold": deal_row.get("stage_critical_threshold"),
                "stage_current_age_status": deal_row.get("stage_current_age_status"),
                "deal_total_work_minutes": deal_row.get("deal_total_work_minutes"),
                "deal_total_work_days": deal_row.get("deal_total_work_days"),
                "deal_total_work_warning_threshold": deal_row.get("deal_total_work_warning_threshold"),
                "deal_total_work_critical_threshold": deal_row.get("deal_total_work_critical_threshold"),
                "deal_total_work_status": deal_row.get("deal_total_work_status"),
                "stage_return_count": deal_row.get("stage_return_count"),
                "stage_reached_final": deal_row.get("stage_reached_final"),
                "stage_movement_risk": deal_row.get("stage_movement_risk"),
                "stage_movement_recommendation": deal_row.get("stage_movement_recommendation"),
            }
        )
    risk_order = {"Зависла": 0, "Нет продвижения": 1, "Возвраты": 2, "Нет истории": 3, "OK": 4, "Финал": 5}
    return sorted(out, key=lambda r: (risk_order.get(str(r.get("stage_movement_risk") or ""), 9), str(r.get("manager_name") or "")))


def build_manager_summary(rows: List[Dict[str, Any]], score_key: str = "overall_score") -> List[Dict[str, Any]]:
    agg: Dict[str, Dict[str, Any]] = {}
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(_deal_group_key(row), []).append(row)

    for _, deal_rows in grouped.items():
        first = deal_rows[0]
        mid = str(first.get("manager_id") or "unknown")
        call_rows = [r for r in deal_rows if not r.get("no_calls")]
        ok_calls = [r for r in call_rows if not r.get("error")]
        scored_calls = [r for r in ok_calls if r.get(score_key) is not None]
        deal_score = (
            sum(float(r.get(score_key) or 0.0) for r in scored_calls) / len(scored_calls)
            if scored_calls
            else 0.0
        )
        a = agg.setdefault(
            mid,
            {
                "manager_id": mid,
                "manager_name": first.get("manager_name") or mid,
                "deals_total": 0,
                "calls_total": 0,
                "calls_ok": 0,
                "deals_without_calls": 0,
                "scored_deals": 0,
                "overall_score_sum": 0.0,
                "growth_deal_data": 0,
                "growth_call_structure": 0,
                "growth_alignment": 0,
                "_checklist_gaps": {},
            },
        )
        a["deals_total"] += 1
        a["calls_total"] += len(call_rows)
        a["calls_ok"] += len(ok_calls)
        if any(r.get("no_calls") for r in deal_rows) or not call_rows:
            a["deals_without_calls"] += 1
        a["scored_deals"] += 1
        a["overall_score_sum"] += deal_score
        if any(not r.get("has_contact") or not r.get("has_amount") or not r.get("has_comments") for r in deal_rows):
            a["growth_deal_data"] += 1
        if not ok_calls or any(float(r.get("call_checklist_percent") or r.get("call_quality_score") or 0.0) < 70 for r in ok_calls):
            a["growth_call_structure"] += 1
        if not ok_calls or not any(r.get("next_step_synced") for r in ok_calls):
            a["growth_alignment"] += 1
        gaps = a.setdefault("_checklist_gaps", {})
        if isinstance(gaps, dict):
            for call in ok_calls:
                for item in call.get("call_checklist_items") or []:
                    if not isinstance(item, dict):
                        continue
                    score = float(item.get("checklist_score") or 0.0)
                    if score >= 1:
                        continue
                    criterion = str(item.get("checklist_criterion") or "").strip()
                    if criterion:
                        gaps[criterion] = int(gaps.get(criterion, 0)) + 1

    out: List[Dict[str, Any]] = []
    for _, a in agg.items():
        total = max(1, int(a["scored_deals"]))
        a["avg_overall_score"] = round(float(a["overall_score_sum"]) / total, 2)
        a.pop("scored_deals", None)
        gaps = a.pop("_checklist_gaps", {}) or {}
        if isinstance(gaps, dict):
            a["top_checklist_gaps"] = "; ".join(
                criterion for criterion, _ in sorted(gaps.items(), key=lambda x: int(x[1]), reverse=True)[:5]
            )
        else:
            a["top_checklist_gaps"] = ""
        zones = sorted(
            [("Ведение CRM", a["growth_deal_data"]), ("Структура разговора", a["growth_call_structure"]), ("Синхронизация звонок↔CRM", a["growth_alignment"])],
            key=lambda x: x[1],
            reverse=True,
        )
        a["top_growth_zones"] = ", ".join([z[0] for z in zones[:3]])
        out.append(a)
    return sorted(out, key=lambda x: x["avg_overall_score"])


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, str) and len(value) > 32000:
        return value[:31900] + "\n\n[Текст обрезан для Excel. Полная версия сохранена в txt-файле.]"
    return value


def _excel_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.map(_excel_safe_value)


def _ru_df(df: pd.DataFrame) -> pd.DataFrame:
    return _excel_df(df).rename(columns=RU_COLUMNS)


def kpi_profile_display(profile: Any) -> Tuple[str, str]:
    key = str(profile or "").strip()
    if not key:
        return "", ""
    title, explanation = KPI_PROFILE_DESCRIPTIONS.get(key, (key, "Пользовательский профиль KPI."))
    return title, explanation


def prepare_report_rows(rows: List[Dict[str, Any]], stage_map: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        r = dict(row)
        r["stage_name"] = stage_display_name(r.get("stage_id"), stage_map=stage_map) or r.get("stage_name")
        profile_title, profile_explanation = kpi_profile_display(r.get("kpi_profile"))
        r["kpi_profile_ru"] = profile_title
        r["kpi_explanation"] = profile_explanation
        if r.get("kpi_profile_cmp"):
            cmp_title, cmp_explanation = kpi_profile_display(r.get("kpi_profile_cmp"))
            r["kpi_profile_cmp_ru"] = cmp_title
            r["kpi_explanation_cmp"] = cmp_explanation
        out.append(r)
    return out


def _format_excel_writer(writer: pd.ExcelWriter) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return

    issue_fill = PatternFill("solid", fgColor="FFF2CC")
    unhandled_fill = PatternFill("solid", fgColor="F4CCCC")
    handled_fill = PatternFill("solid", fgColor="D9EAD3")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    critical_fill = PatternFill("solid", fgColor="F4CCCC")

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_cells in ws.columns:
            header = str(col_cells[0].value or "")
            max_len = len(header)
            for cell in col_cells[1:80]:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, min(80, len(str(value))))
                if isinstance(value, str) and ("\n" in value or len(value) > 90):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            width = min(70, max(12, max_len + 2))
            if any(word in header.lower() for word in ["расшифров", "вывод", "рекомендац", "диагност", "попытки", "ошибки"]):
                width = min(90, max(width, 45))
            ws.column_dimensions[col_cells[0].column_letter].width = width

        headers = {str(cell.value or ""): cell.column for cell in ws[1]}
        issue_headers = [
            "Моменты для улучшения",
            "Неотработанные возражения",
            "Варианты отработки возражений",
            "Расшифровка с пометками",
            "Что усилить по этапу",
        ]
        for row_idx in range(2, ws.max_row + 1):
            for header in issue_headers:
                col_idx = headers.get(header)
                if not col_idx:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell.fill = issue_fill

            status_col = headers.get("Статус отработки")
            if status_col:
                status = str(ws.cell(row=row_idx, column=status_col).value or "").lower()
                if "не отработано" in status:
                    fill = unhandled_fill
                elif "отработано" in status:
                    fill = handled_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            risk_headers = ["Статус срока в стадии", "Статус общего срока сделки", "Риск движения по воронке"]
            risk_values = " ".join(
                str(ws.cell(row=row_idx, column=headers[h]).value or "").lower()
                for h in risk_headers
                if h in headers
            )
            if "тревога" in risk_values:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = critical_fill
            elif "предупреждение" in risk_values:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = warning_fill

            checklist_score_col = headers.get("Оценка критерия")
            if not checklist_score_col:
                checklist_score_col = headers.get("Оценка CRM-критерия")
            if checklist_score_col:
                try:
                    checklist_score = float(ws.cell(row=row_idx, column=checklist_score_col).value or 0)
                    if checklist_score <= 0:
                        fill = critical_fill
                    elif checklist_score < 1:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            stage_percent_col = headers.get("Выполнение этапа, %")
            if stage_percent_col:
                try:
                    stage_percent = float(ws.cell(row=row_idx, column=stage_percent_col).value or 0)
                    if stage_percent < 50:
                        fill = critical_fill
                    elif stage_percent < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            avg_stage_percent_col = headers.get("Среднее выполнение этапа, %")
            if avg_stage_percent_col:
                try:
                    avg_stage_percent = float(ws.cell(row=row_idx, column=avg_stage_percent_col).value or 0)
                    if avg_stage_percent < 50:
                        fill = critical_fill
                    elif avg_stage_percent < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            criterion_completion_col = headers.get("Выполнение критерия, %")
            if not criterion_completion_col:
                criterion_completion_col = headers.get("Выполнение CRM-критерия, %")
            if criterion_completion_col:
                try:
                    completion = float(ws.cell(row=row_idx, column=criterion_completion_col).value or 0)
                    if completion < 50:
                        fill = critical_fill
                    elif completion < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            priority_col = headers.get("Приоритет контроля") or headers.get("Приоритет обучения")
            if priority_col:
                priority = str(ws.cell(row=row_idx, column=priority_col).value or "").lower()
                if "критично" in priority:
                    fill = critical_fill
                elif "высок" in priority:
                    fill = warning_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill


def flatten_results(
    rows: List[Dict[str, Any]],
    manager_summary: List[Dict[str, Any]],
    manager_summary_cmp: Optional[List[Dict[str, Any]]] = None,
    stage_map: Optional[Dict[str, str]] = None,
    lost_deals_analysis: Optional[Dict[str, List[Dict[str, Any]]]] = None,
) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = REPORTS_DIR / f"bitnewton_sync_report_{ts}.xlsx"
    report_rows = prepare_report_rows(rows, stage_map=stage_map)
    df = pd.DataFrame(report_rows)
    deal_report_rows = build_deal_report_rows(report_rows)
    call_detail_rows = build_call_detail_rows(report_rows)
    objection_rows = build_objection_rows(report_rows)
    checklist_rows = build_call_checklist_rows(report_rows)
    sales_stage_score_rows = build_sales_stage_score_rows(report_rows)
    manager_stage_rows = build_manager_stage_summary_rows(report_rows)
    manager_criterion_gap_rows = build_manager_criterion_gap_rows(report_rows)
    crm_checklist_rows = build_crm_checklist_rows(report_rows)
    manager_crm_gap_rows = build_manager_crm_gap_rows(report_rows)
    quality_control_rows = build_quality_control_rows(deal_report_rows)
    coaching_plan_rows = build_coaching_plan_rows(manager_criterion_gap_rows, manager_crm_gap_rows, manager_stage_rows)
    executive_summary_rows = build_executive_summary_rows(
        deal_report_rows,
        manager_summary,
        quality_control_rows,
        manager_criterion_gap_rows,
        manager_crm_gap_rows,
        coaching_plan_rows,
    )
    manager_scorecard_rows = build_manager_scorecard_rows(manager_summary, quality_control_rows, coaching_plan_rows)
    stage_history_rows = build_stage_history_rows(report_rows)
    stage_sr_rows = build_stage_sr_rows(report_rows, stage_map=stage_map)
    stage_movement_rows = build_stage_movement_rows(report_rows)
    lost_deal_rows = (lost_deals_analysis or {}).get("rows") or []
    lost_reason_summary_rows = (lost_deals_analysis or {}).get("summary_rows") or []
    conversion_action_rows = (lost_deals_analysis or {}).get("action_rows") or []
    deal_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "kpi_profile_ru",
        "kpi_explanation",
        "calls_total",
        "calls_ok",
        "calls_failed",
        "asr_skipped_calls",
        "ignored_call_center_calls",
        "skipped_short_calls",
        "no_calls",
        "transcripts_count",
        "avg_overall_score",
        "overall_score_details",
        "avg_call_quality_score",
        "call_quality_details",
        "call_checklist_percent",
        "call_checklist_block_details",
        "deal_quality_score",
        "deal_quality_details",
        "alignment_score",
        "crm_work_score",
        "crm_checklist_percent",
        "crm_checklist_details",
        "alignment_details",
        "stage_history_count",
        "stage_history_path",
        "stage_last_change_at",
        "stage_current_age_minutes",
        "stage_current_age_days",
        "stage_warning_threshold",
        "stage_critical_threshold",
        "stage_current_age_status",
        "deal_total_work_minutes",
        "deal_total_work_days",
        "deal_total_work_warning_threshold",
        "deal_total_work_critical_threshold",
        "deal_total_work_status",
        "stage_return_count",
        "stage_reached_final",
        "stage_movement_risk",
        "stage_movement_recommendation",
        "client_work_score",
        "client_work_quality",
        "conversation_quality",
        "client_work_conclusion",
        "call_quality_conclusion",
        "conversation_meaning",
        "recommendations",
        "calls_breakdown",
        "improvement_moments_combined",
        "objections_combined",
        "unhandled_objections",
        "objection_recommendations",
        "transcript_paths",
        "error",
    ]
    call_detail_cols = [
        "deal_url",
        "call_number",
        "subject",
        "duration_minutes",
        "call_has_error",
        "asr_status",
        "skipped_short_calls",
        "call_quality_score",
        "call_quality_details",
        "call_checklist_total_score",
        "call_checklist_max_score",
        "call_checklist_percent",
        "call_checklist_block_details",
        "has_greeting",
        "has_needs_discovery",
        "has_objection_work",
        "has_next_step_phrase",
        "alignment_score",
        "crm_work_score",
        "crm_checklist_percent",
        "crm_checklist_details",
        "alignment_details",
        "next_step_synced",
        "next_step_synced_details",
        "call_quality_conclusion",
        "conversation_meaning",
        "improvement_moments",
        "unhandled_objections",
        "objection_recommendations",
        "transcript_path",
        "transcript_marked",
        "transcript_text",
        "bitrix_card_transcript_status",
        "transcript_match_score",
        "bitrix_card_transcript",
        "error",
    ]
    objection_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "subject",
        "objection_fragment",
        "objection_status",
        "objection_recommendations",
    ]
    checklist_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "call_number",
        "subject",
        "duration_minutes",
        "checklist_block_name",
        "checklist_criterion",
        "checklist_score",
        "checklist_max_score",
        "checklist_evidence",
        "checklist_comment",
        "checklist_code",
    ]
    sales_stage_score_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "call_number",
        "subject",
        "duration_minutes",
        "sales_stage_block_name",
        "sales_stage_score",
        "sales_stage_max_score",
        "sales_stage_percent",
        "sales_stage_missing",
    ]
    manager_stage_cols = [
        "manager_name",
        "sales_stage_block_name",
        "manager_stage_calls",
        "manager_stage_deals",
        "manager_stage_avg_percent",
        "manager_stage_weak_calls",
        "manager_stage_weak_rate",
    ]
    manager_criterion_gap_cols = [
        "manager_name",
        "checklist_block_name",
        "checklist_criterion",
        "criterion_calls",
        "criterion_avg_score",
        "criterion_completion_percent",
        "criterion_failed_count",
        "criterion_partial_count",
        "criterion_fail_rate",
        "training_recommendation",
    ]
    crm_checklist_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "crm_checklist_block_name",
        "crm_checklist_criterion",
        "crm_checklist_score",
        "crm_checklist_max_score",
        "crm_checklist_comment",
        "crm_checklist_code",
    ]
    manager_crm_gap_cols = [
        "manager_name",
        "crm_checklist_block_name",
        "crm_checklist_criterion",
        "crm_criterion_deals",
        "crm_criterion_avg_score",
        "crm_criterion_completion_percent",
        "crm_criterion_failed_count",
        "crm_criterion_partial_count",
        "crm_criterion_fail_rate",
        "training_recommendation",
    ]
    quality_control_cols = [
        "control_priority",
        "quality_control_score",
        "deal_url",
        "stage_name",
        "manager_name",
        "avg_overall_score",
        "avg_call_quality_score",
        "crm_checklist_percent",
        "calls_total",
        "calls_failed",
        "stage_movement_risk",
        "control_reason",
        "control_next_action",
        "recommendations",
    ]
    coaching_plan_cols = [
        "manager_name",
        "coaching_priority",
        "training_source",
        "training_topic",
        "coaching_metric",
        "coaching_affected_count",
        "training_recommendation",
    ]
    executive_summary_cols = [
        "summary_section",
        "summary_metric",
        "summary_value",
        "summary_comment",
    ]
    manager_scorecard_cols = [
        "manager_rank",
        "manager_name",
        "avg_overall_score",
        "deals_total",
        "calls_total",
        "calls_ok",
        "deals_without_calls",
        "manager_critical_deals",
        "manager_high_deals",
        "manager_training_topics",
        "manager_top_training_topic",
        "top_growth_zones",
        "top_checklist_gaps",
        "manager_focus",
        "manager_next_action",
    ]
    stage_history_cols = [
        "deal_url",
        "manager_name",
        "stage_history_event_id",
        "stage_history_event_type",
        "stage_history_created_at",
        "stage_id",
        "stage_name",
        "stage_order",
        "stage_duration_minutes",
        "stage_duration_hours",
        "stage_is_current",
    ]
    stage_sr_cols = [
        "stage_order",
        "stage_id",
        "stage_name",
        "stage_deals_total",
        "stage_deals_passed",
        "stage_sr",
    ]
    stage_movement_cols = [
        "deal_url",
        "stage_name",
        "manager_name",
        "stage_history_count",
        "stage_history_path",
        "stage_last_change_at",
        "stage_current_age_minutes",
        "stage_current_age_days",
        "stage_warning_threshold",
        "stage_critical_threshold",
        "stage_current_age_status",
        "deal_total_work_minutes",
        "deal_total_work_days",
        "deal_total_work_warning_threshold",
        "deal_total_work_critical_threshold",
        "deal_total_work_status",
        "stage_return_count",
        "stage_reached_final",
        "stage_movement_risk",
        "stage_movement_recommendation",
    ]
    lost_deal_cols = [
        "lost_deal_url",
        "lost_deal_title",
        "lost_stage_name",
        "lost_manager_name",
        "lost_amount",
        "lost_date_create",
        "lost_close_date",
        "lost_lifetime_days",
        "lost_source",
        "loss_reason_category",
        "loss_reason_confidence",
        "loss_reason_evidence",
        "conversion_tools",
        "conversion_next_action",
        "lost_analysis_basis",
    ]
    lost_reason_summary_cols = [
        "loss_reason_category",
        "lost_deals_count",
        "lost_deals_share",
        "lost_amount",
        "lost_avg_lifetime_days",
        "lost_top_managers",
        "conversion_tools",
        "conversion_next_action",
    ]
    conversion_action_cols = [
        "conversion_priority",
        "conversion_rank",
        "loss_reason_category",
        "lost_deals_count",
        "lost_deals_share",
        "conversion_tools",
        "conversion_next_action",
        "conversion_expected_effect",
    ]
    manager_cols = [
        "manager_name",
        "deals_total",
        "calls_total",
        "calls_ok",
        "deals_without_calls",
        "growth_deal_data",
        "growth_call_structure",
        "growth_alignment",
        "avg_overall_score",
        "top_growth_zones",
        "top_checklist_gaps",
    ]
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        executive_summary_df = pd.DataFrame(executive_summary_rows, columns=executive_summary_cols)
        _ru_df(executive_summary_df).to_excel(writer, sheet_name="Итоги", index=False)
        manager_scorecard_df = pd.DataFrame(manager_scorecard_rows, columns=manager_scorecard_cols)
        _ru_df(manager_scorecard_df).to_excel(writer, sheet_name="Карточки менеджеров", index=False)
        deal_df = pd.DataFrame(deal_report_rows)
        _ru_df(deal_df[[c for c in deal_cols if c in deal_df.columns]]).to_excel(writer, sheet_name="Сделки", index=False)
        if not df.empty:
            call_df = pd.DataFrame(call_detail_rows)
            _ru_df(call_df[[c for c in call_detail_cols if c in call_df.columns]]).to_excel(writer, sheet_name="Звонки внутри сделок", index=False)
            objection_df = pd.DataFrame(objection_rows, columns=objection_cols)
            _ru_df(objection_df).to_excel(writer, sheet_name="Разбор возражений", index=False)
            checklist_df = pd.DataFrame(checklist_rows, columns=checklist_cols)
            _ru_df(checklist_df).to_excel(writer, sheet_name="Чек-лист звонков", index=False)
            sales_stage_score_df = pd.DataFrame(sales_stage_score_rows, columns=sales_stage_score_cols)
            _ru_df(sales_stage_score_df).to_excel(writer, sheet_name="Этапы продаж", index=False)
            manager_stage_df = pd.DataFrame(manager_stage_rows, columns=manager_stage_cols)
            _ru_df(manager_stage_df).to_excel(writer, sheet_name="Слабые этапы", index=False)
            manager_criterion_gap_df = pd.DataFrame(manager_criterion_gap_rows, columns=manager_criterion_gap_cols)
            _ru_df(manager_criterion_gap_df).to_excel(writer, sheet_name="Проблемные критерии", index=False)
            crm_checklist_df = pd.DataFrame(crm_checklist_rows, columns=crm_checklist_cols)
            _ru_df(crm_checklist_df).to_excel(writer, sheet_name="Чек-лист CRM", index=False)
            manager_crm_gap_df = pd.DataFrame(manager_crm_gap_rows, columns=manager_crm_gap_cols)
            _ru_df(manager_crm_gap_df).to_excel(writer, sheet_name="Проблемы CRM", index=False)
            quality_control_df = pd.DataFrame(quality_control_rows, columns=quality_control_cols)
            _ru_df(quality_control_df).to_excel(writer, sheet_name="Контроль качества", index=False)
            coaching_plan_df = pd.DataFrame(coaching_plan_rows, columns=coaching_plan_cols)
            _ru_df(coaching_plan_df).to_excel(writer, sheet_name="План обучения", index=False)
            stage_history_df = pd.DataFrame(stage_history_rows, columns=stage_history_cols)
            _ru_df(stage_history_df).to_excel(writer, sheet_name="История стадий", index=False)
            stage_sr_df = pd.DataFrame(stage_sr_rows, columns=stage_sr_cols)
            _ru_df(stage_sr_df).to_excel(writer, sheet_name="SR по стадиям", index=False)
            stage_movement_df = pd.DataFrame(stage_movement_rows, columns=stage_movement_cols)
            _ru_df(stage_movement_df).to_excel(writer, sheet_name="Риски движения", index=False)
        if lost_deal_rows or lost_reason_summary_rows or conversion_action_rows:
            lost_df = pd.DataFrame(lost_deal_rows, columns=lost_deal_cols)
            _ru_df(lost_df).to_excel(writer, sheet_name="Проигранные сделки", index=False)
            lost_summary_df = pd.DataFrame(lost_reason_summary_rows, columns=lost_reason_summary_cols)
            _ru_df(lost_summary_df).to_excel(writer, sheet_name="Причины отказов", index=False)
            conversion_df = pd.DataFrame(conversion_action_rows, columns=conversion_action_cols)
            _ru_df(conversion_df).to_excel(writer, sheet_name="Рост конверсии", index=False)
        manager_df = pd.DataFrame(manager_summary)
        _ru_df(manager_df[[c for c in manager_cols if c in manager_df.columns]]).to_excel(writer, sheet_name="Сводка менеджеров", index=False)
        if manager_summary_cmp is not None:
            manager_cmp_df = pd.DataFrame(manager_summary_cmp)
            _ru_df(manager_cmp_df[[c for c in manager_cols if c in manager_cmp_df.columns]]).to_excel(writer, sheet_name="Сводка менеджеров cmp", index=False)
        _format_excel_writer(writer)
    return out
