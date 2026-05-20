from __future__ import annotations

from typing import Any

import pandas as pd

from pipelines.reporting.ru_columns import RU_COLUMNS


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, str) and len(value) > 32000:
        return value[:31900] + "\n\n[Текст обрезан для Excel. Полная версия сохранена в txt-файле.]"
    return value


def _excel_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.map(_excel_safe_value)


def _ru_df(df: pd.DataFrame) -> pd.DataFrame:
    return _excel_df(df).rename(columns=RU_COLUMNS)


def _format_excel_writer(writer: pd.ExcelWriter) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return

    issue_fill = PatternFill("solid", fgColor="FFF2CC")
    unhandled_fill = PatternFill("solid", fgColor="F4CCCC")
    handled_fill = PatternFill("solid", fgColor="D9EAD3")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    critical_fill = PatternFill("solid", fgColor="F4CCCC")

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_cells in ws.columns:
            header = str(col_cells[0].value or "")
            max_len = len(header)
            for cell in col_cells[1:80]:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, min(80, len(str(value))))
                if isinstance(value, str) and ("\n" in value or len(value) > 90):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            width = min(70, max(12, max_len + 2))
            if any(
                word in header.lower()
                for word in ["расшифров", "вывод", "рекомендац", "диагност", "попытки", "ошибки"]
            ):
                width = min(90, max(width, 45))
            ws.column_dimensions[col_cells[0].column_letter].width = width

        headers = {str(cell.value or ""): cell.column for cell in ws[1]}
        issue_headers = [
            "Моменты для улучшения",
            "Неотработанные возражения",
            "Варианты отработки возражений",
            "Расшифровка с пометками",
            "Что усилить по этапу",
            "Критичные ошибки",
            "Проваленные шаги",
        ]
        for row_idx in range(2, ws.max_row + 1):
            for header in issue_headers:
                col_idx = headers.get(header)
                if not col_idx:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell.fill = issue_fill

            status_col = headers.get("Статус отработки")
            if status_col:
                status = str(ws.cell(row=row_idx, column=status_col).value or "").lower()
                if "не отработано" in status:
                    fill = unhandled_fill
                elif "отработано" in status:
                    fill = handled_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            script_status_col = headers.get("Статус шага")
            if script_status_col:
                status = str(ws.cell(row=row_idx, column=script_status_col).value or "").lower()
                if "провал" in status:
                    fill = critical_fill
                elif "частично" in status:
                    fill = warning_fill
                elif "выполнено" in status:
                    fill = handled_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            script_profile_status_col = headers.get("Статус соответствия скрипту")
            if script_profile_status_col:
                status = str(
                    ws.cell(row=row_idx, column=script_profile_status_col).value or ""
                ).lower()
                if "не соответствует" in status:
                    fill = critical_fill
                elif "частично" in status or "замечания" in status:
                    fill = warning_fill
                elif "соответствует" in status:
                    fill = handled_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            risk_headers = [
                "Статус срока в стадии",
                "Статус общего срока сделки",
                "Риск движения по воронке",
            ]
            risk_values = " ".join(
                str(ws.cell(row=row_idx, column=headers[h]).value or "").lower()
                for h in risk_headers
                if h in headers
            )
            if "тревога" in risk_values:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = critical_fill
            elif "предупреждение" in risk_values:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = warning_fill

            checklist_score_col = headers.get("Оценка критерия")
            if not checklist_score_col:
                checklist_score_col = headers.get("Оценка CRM-критерия")
            if checklist_score_col:
                try:
                    checklist_score = float(
                        ws.cell(row=row_idx, column=checklist_score_col).value or 0
                    )
                    if checklist_score <= 0:
                        fill = critical_fill
                    elif checklist_score < 1:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            stage_percent_col = headers.get("Выполнение этапа, %")
            if stage_percent_col:
                try:
                    stage_percent = float(ws.cell(row=row_idx, column=stage_percent_col).value or 0)
                    if stage_percent < 50:
                        fill = critical_fill
                    elif stage_percent < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            avg_stage_percent_col = headers.get("Среднее выполнение этапа, %")
            if avg_stage_percent_col:
                try:
                    avg_stage_percent = float(
                        ws.cell(row=row_idx, column=avg_stage_percent_col).value or 0
                    )
                    if avg_stage_percent < 50:
                        fill = critical_fill
                    elif avg_stage_percent < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            criterion_completion_col = headers.get("Выполнение критерия, %")
            if not criterion_completion_col:
                criterion_completion_col = headers.get("Выполнение CRM-критерия, %")
            if criterion_completion_col:
                try:
                    completion = float(
                        ws.cell(row=row_idx, column=criterion_completion_col).value or 0
                    )
                    if completion < 50:
                        fill = critical_fill
                    elif completion < 70:
                        fill = warning_fill
                    else:
                        fill = None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                except Exception:
                    pass

            priority_col = (
                headers.get("Приоритет контроля")
                or headers.get("Приоритет обучения")
                or headers.get("Приоритет рекомендации")
                or headers.get("Приоритет фактора")
                or headers.get("Уровень эмоционального риска")
                or headers.get("Риск момента")
                or headers.get("Приоритет")
            )
            if priority_col:
                priority = str(ws.cell(row=row_idx, column=priority_col).value or "").lower()
                if "критично" in priority:
                    fill = critical_fill
                elif "высок" in priority:
                    fill = warning_fill
                else:
                    fill = None
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
