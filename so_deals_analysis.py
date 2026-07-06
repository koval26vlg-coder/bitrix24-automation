from __future__ import annotations

import argparse
import asyncio
import json
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bitrix24_api import Bitrix24API
from client_requests_analysis import (
    FIRED_DEPARTMENT_NAME,
    SALES_DEPARTMENT_NAME,
    _clean,
    _dedupe_texts,
    _extract_text_from_activity,
    _short,
)
from logging_setup import get_logger
from pipelines.calls import (
    extract_bitrix_gpt_summaries,
    fetch_deal_activities,
    fetch_open_next_step_activities,
    fetch_timeline_comments,
)
from pipelines.deals import deal_url_from_id
from pipelines.paths import REPORTS_DIR
from pipelines.stages import safe_int
from pipelines.transcription import _load_state_cache, load_cached_transcript

logger = get_logger(__name__)

SO_CATEGORY_ID = "55"
SO_CATEGORY_NAME = "Условный отказ"
DEFAULT_DOMAIN = "online-kassa.bitrix24.ru"

STAGE_NAMES = {
    "C55:NEW": "Сделка из ОП",
    "C55:PREPARATION": "Назначен ответственный",
    "C55:PREPAYMENT_INVOIC": "Потребность выявлена",
    "C55:UC_T1XILX": "Закрыто и не реализовано",
    "C55:WON": "Сделка успешна",
    "C55:LOSE": "Сделка провалена",
    "C55:APOLOGY": "Анализ причины провала",
}

FINAL_NOT_WON_STAGES = {"C55:UC_T1XILX", "C55:LOSE", "C55:APOLOGY"}


@dataclass(frozen=True)
class ReasonRule:
    name: str
    pattern: str
    recommendation: str


REASON_RULES = (
    ReasonRule(
        "Цена/бюджет",
        r"дорог|дороже|цена|стоим|прайс|бюджет|денег|скидк|оплат|сумм|дешев",
        "Разбирать цену через ценность, риски бездействия и пакетные варианты КП.",
    ),
    ReasonRule(
        "Нет связи / не дозвонились",
        r"не дозвон|не отвечает|не берет|не бер[её]т|недоступ|нет связи|молчит|автосекрет|трубк",
        "Ввести регламент касаний: звонок, сообщение, письмо и обязательная задача на повтор.",
    ),
    ReasonRule(
        "Нет потребности / не актуально",
        r"не актуаль|не нужно|не требуется|не интересно|нет потребн|отказ|передумал",
        "Усилить квалификацию на входе и переводить отложенный спрос в контролируемый прогрев.",
    ),
    ReasonRule(
        "Клиент отложил решение",
        r"позже|потом|отлож|не сейчас|верн[её]мся|следующ|месяц|квартал|август|июл|срок",
        "Фиксировать дату возврата и конкретное дело в CRM вместо закрытия без плана.",
    ),
    ReasonRule(
        "Конкурент / уже купили",
        r"конкур|друг(ой|ая|ие)|уже куп|купили|выбрали|поставщик|альтернатив|местн",
        "Собирать причину выбора конкурента и использовать battlecard по отличиям.",
    ),
    ReasonRule(
        "Техническое несоответствие",
        r"не подходит|невозмож|не можем|не умеет|нет функц|доработ|ошибк|тех|интеграц|совместим",
        "Подключать тех. пресейл до КП и вести реестр типовых технических ограничений.",
    ),
    ReasonRule(
        "КП/счет/документы не доведены",
        r"\bкп\b|коммерческ|счет|сч[её]т|договор|документ|предложени|защита",
        "Контролировать отправку КП, защиту КП и follow-up после счета.",
    ),
    ReasonRule(
        "Дубль / ошибочная сделка",
        r"дубл|дубликат|ошибоч|ошибка|спам|тест",
        "Очистить аналитику от дублей и настроить проверку карточки перед началом работы.",
    ),
)


def parse_dt(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except Exception:
        return None


def days_since(value: Any) -> float | None:
    dt = parse_dt(value)
    if not dt:
        return None
    now = datetime.now(dt.tzinfo or timezone.utc)
    return round(max(0.0, (now - dt).total_seconds() / 86400.0), 1)


def lifetime_days(deal: dict[str, Any]) -> float | None:
    created = parse_dt(deal.get("DATE_CREATE"))
    closed = parse_dt(deal.get("CLOSEDATE")) or parse_dt(deal.get("DATE_MODIFY"))
    if not created or not closed:
        return None
    return round(max(0.0, (closed - created).total_seconds() / 86400.0), 2)


def as_float(value: Any) -> float:
    try:
        return float(str(value or 0).replace(",", "."))
    except Exception:
        return 0.0


def loss_state(stage_id: str) -> str:
    if stage_id == "C55:WON":
        return "Выиграна"
    if stage_id in {"C55:LOSE", "C55:APOLOGY"}:
        return "Провалена"
    if stage_id == "C55:UC_T1XILX":
        return "Закрыто без реализации"
    return "Не выиграна, но не закрыта"


def has_meaningful_comment(*values: Any) -> bool:
    text = " ".join(_clean(v) for v in values if v)
    lowered = text.lower().replace("ё", "е")
    if len(lowered) < 25:
        return False
    if "изменен ответственный" in lowered or "изменён ответственный" in lowered:
        return False
    return True


def evidence_snippet(text: str, match: re.Match[str], radius: int = 110) -> str:
    start = max(0, match.start() - radius)
    end = min(len(text), match.end() + radius)
    return _short(text[start:end].strip(" .;,\n\t"), 420)


def classify_reason(texts: list[str], stage_id: str) -> dict[str, str]:
    clean = _clean(" ".join(texts))
    lowered = clean.lower().replace("ё", "е")
    for rule in REASON_RULES:
        match = re.search(rule.pattern, lowered, flags=re.IGNORECASE)
        if match:
            return {
                "reason_category": rule.name,
                "reason_confidence": "Высокая" if stage_id in FINAL_NOT_WON_STAGES else "Средняя",
                "reason_evidence": evidence_snippet(clean, match),
                "recommendation": rule.recommendation,
            }
    if stage_id in FINAL_NOT_WON_STAGES:
        confidence = "Низкая"
    else:
        confidence = "Средняя"
    return {
        "reason_category": "Причина не указана / мало контекста",
        "reason_confidence": confidence,
        "reason_evidence": "",
        "recommendation": "Сделать причину отказа обязательной: комментарий должен объяснять, почему сделка не реализована.",
    }


def crm_quality_score(row: dict[str, Any]) -> tuple[int, str]:
    score = 100
    issues: list[str] = []
    if not row.get("has_meaningful_comment"):
        score -= 30
        issues.append("нет содержательного комментария")
    if (
        row.get("context_enriched")
        and row.get("has_next_step") is False
        and row.get("loss_state") == "Не выиграна, но не закрыта"
    ):
        score -= 25
        issues.append("нет открытого следующего шага")
    if row.get("reason_category") == "Причина не указана / мало контекста":
        score -= 25
        issues.append("не указана понятная причина")
    if as_float(row.get("amount")) <= 0 and row.get("loss_state") != "Не выиграна, но не закрыта":
        score -= 10
        issues.append("нет суммы по закрытой/проваленной сделке")
    if not row.get("source"):
        score -= 10
        issues.append("нет источника")
    if row.get("days_since_update") is not None and row["days_since_update"] > 14 and row.get("loss_state") == "Не выиграна, но не закрыта":
        score -= 20
        issues.append("зависла без обновления больше 14 дней")
    score = max(0, min(100, score))
    return score, "; ".join(issues) if issues else "критичных замечаний по CRM нет"


async def call_with_page_retry(
    api: Bitrix24API, method: str, params: dict[str, Any], *, label: str, attempts: int = 6
) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            return await api.call(method, params)
        except Exception as exc:
            last_error = exc
            wait = min(20.0, 0.8 * (2 ** (attempt - 1)))
            print(f"[WARN] {label}: retry {attempt}/{attempts}: {exc}; sleep={wait:.1f}s", flush=True)
            await asyncio.sleep(wait)
    raise RuntimeError(f"{label}: failed after {attempts} attempts: {last_error}")


async def load_sales_users(api: Bitrix24API) -> tuple[list[str], dict[str, str]]:
    departments = (await api.call("department.get", {})).get("result", []) or []
    dep_by_id = {str(d.get("ID")): str(d.get("NAME") or "") for d in departments}
    sales_dep_ids = {
        str(d.get("ID"))
        for d in departments
        if str(d.get("NAME") or "").strip().lower() == SALES_DEPARTMENT_NAME.lower()
    }
    users = await api.get_all("user.get", {"FILTER": {"ACTIVE": True}})
    sales_ids: list[str] = []
    names: dict[str, str] = {}
    for user in users:
        uid = str(user.get("ID") or "")
        deps = [str(x) for x in (user.get("UF_DEPARTMENT") or [])]
        dep_names = [dep_by_id.get(x, "") for x in deps]
        if not uid or not (set(deps) & sales_dep_ids):
            continue
        if any(name.strip().lower() == FIRED_DEPARTMENT_NAME.lower() for name in dep_names):
            continue
        sales_ids.append(uid)
        names[uid] = f"{user.get('NAME', '')} {user.get('LAST_NAME', '')}".strip() or uid
    return sales_ids, names


async def fetch_deals(
    api: Bitrix24API,
    *,
    sales_ids: list[str],
    raw_path: Path,
    resume: bool,
) -> list[dict[str, Any]]:
    if resume and raw_path.exists():
        existing = json.loads(raw_path.read_text(encoding="utf-8"))
        if isinstance(existing, list) and existing:
            print(f"[CACHE] loaded raw deals: {len(existing)}", flush=True)
            return existing

    flt = {
        "CATEGORY_ID": SO_CATEGORY_ID,
        "ASSIGNED_BY_ID": sales_ids,
        "!=STAGE_SEMANTIC_ID": "S",
    }
    select = [
        "ID",
        "TITLE",
        "ASSIGNED_BY_ID",
        "STAGE_ID",
        "STAGE_SEMANTIC_ID",
        "DATE_CREATE",
        "DATE_MODIFY",
        "CLOSEDATE",
        "COMMENTS",
        "SOURCE_ID",
        "SOURCE_DESCRIPTION",
        "OPPORTUNITY",
        "CATEGORY_ID",
        "CONTACT_ID",
        "COMPANY_ID",
    ]
    deals: list[dict[str, Any]] = []
    start: int | None = 0
    total: int | None = None
    while start is not None:
        params = {
            "filter": flt,
            "select": select,
            "order": {"ID": "DESC"},
            "start": start,
        }
        res = await call_with_page_retry(api, "crm.deal.list", params, label=f"deals page {start}")
        chunk = res.get("result", []) or []
        if not chunk:
            break
        total = safe_int(res.get("total")) or total
        deals.extend(chunk)
        print(f"[DEALS] {len(deals)}/{total or '?'}", flush=True)
        if len(deals) % 1000 < len(chunk):
            raw_path.write_text(json.dumps(deals, ensure_ascii=False, indent=2), encoding="utf-8")
        next_start = res.get("next")
        start = safe_int(next_start) if next_start is not None else None
        await asyncio.sleep(0.18)
    raw_path.write_text(json.dumps(deals, ensure_ascii=False, indent=2), encoding="utf-8")
    return deals


async def enrich_context(
    api: Bitrix24API,
    deal: dict[str, Any],
    *,
    manager_names: dict[str, str],
    state_cache: dict[str, Any],
    semaphore: asyncio.Semaphore,
) -> dict[str, Any]:
    deal_id = str(deal.get("ID") or "")
    async with semaphore:
        comments_task = asyncio.create_task(fetch_timeline_comments(api, deal_id))
        activities_task = asyncio.create_task(fetch_deal_activities(api, deal_id, limit=80))
        next_steps_task = asyncio.create_task(fetch_open_next_step_activities(api, deal_id))
        comments, activities, next_steps = await asyncio.gather(
            comments_task, activities_task, next_steps_task, return_exceptions=True
        )
    if isinstance(comments, Exception):
        comments = [f"ERROR comments: {comments}"]
    if isinstance(activities, Exception):
        activities = []
    if isinstance(next_steps, Exception):
        next_steps = []

    texts: list[str] = [
        _clean(deal.get("TITLE")),
        _clean(deal.get("COMMENTS")),
        _clean(deal.get("SOURCE_DESCRIPTION")),
    ]
    texts.extend([_clean(x) for x in comments])

    calls_count = 0
    cached_transcripts = 0
    chat_count = 0
    for activity in activities:
        provider = str(activity.get("PROVIDER_ID") or "").upper()
        type_id = str(activity.get("TYPE_ID") or "")
        subject_low = str(activity.get("SUBJECT") or "").lower()
        if provider == "VOXIMPLANT_CALL" or type_id == "2":
            calls_count += 1
            transcript, _path = load_cached_transcript(
                state_cache,
                str(activity.get("ORIGIN_ID") or ""),
                deal_id,
                activity.get("ID"),
            )
            if transcript:
                cached_transcripts += 1
                texts.append(transcript)
        if any(marker in subject_low for marker in ("чат", "telegram", "whatsapp", "openline", "max")):
            chat_count += 1
        texts.extend(_extract_text_from_activity(activity))

    for activity in next_steps:
        texts.extend(_extract_text_from_activity(activity))

    summaries = extract_bitrix_gpt_summaries(
        [str(x) for x in comments] + [_clean(deal.get("COMMENTS"))], activities
    )
    if summaries.get("bitrix_chat_summary"):
        texts.insert(0, str(summaries.get("bitrix_chat_summary")))
    if summaries.get("bitrix_call_summary"):
        texts.insert(0, str(summaries.get("bitrix_call_summary")))

    texts = _dedupe_texts(texts, max_items=80)
    reason = classify_reason(texts, str(deal.get("STAGE_ID") or ""))

    return {
        "deal_id": deal_id,
        "context_enriched": True,
        "comments_count": len(comments),
        "activities_count": len(activities),
        "next_steps_count": len(next_steps),
        "calls_count": calls_count,
        "cached_transcripts_count": cached_transcripts,
        "chat_activity_count": chat_count,
        "bitrix_chat_summary": summaries.get("bitrix_chat_summary") or "",
        "bitrix_call_summary": summaries.get("bitrix_call_summary") or "",
        "context_excerpt": _short(" | ".join(texts[:10]), 1800),
        **reason,
        "has_meaningful_comment": has_meaningful_comment(deal.get("COMMENTS"), *comments),
        "has_next_step": bool(next_steps),
        "manager_name": manager_names.get(str(deal.get("ASSIGNED_BY_ID") or ""), str(deal.get("ASSIGNED_BY_ID") or "")),
    }


def base_row(deal: dict[str, Any], manager_names: dict[str, str], domain: str) -> dict[str, Any]:
    stage_id = str(deal.get("STAGE_ID") or "")
    stage_name = STAGE_NAMES.get(stage_id, stage_id)
    manager_id = str(deal.get("ASSIGNED_BY_ID") or "")
    source = _clean(deal.get("SOURCE_DESCRIPTION")) or _clean(deal.get("SOURCE_ID"))
    title = _clean(deal.get("TITLE"))
    comments = _clean(deal.get("COMMENTS"))
    base_texts = [title, comments, source, stage_name]
    reason = classify_reason(base_texts, stage_id)
    state = loss_state(stage_id)
    row = {
        "deal_id": str(deal.get("ID") or ""),
        "deal_url": deal_url_from_id(domain, str(deal.get("ID") or "")),
        "title": title,
        "manager_id": manager_id,
        "manager_name": manager_names.get(manager_id, manager_id),
        "stage_id": stage_id,
        "stage_name": stage_name,
        "loss_state": state,
        "date_create": deal.get("DATE_CREATE"),
        "date_modify": deal.get("DATE_MODIFY"),
        "close_date": deal.get("CLOSEDATE"),
        "lifetime_days": lifetime_days(deal),
        "days_since_update": days_since(deal.get("DATE_MODIFY")),
        "amount": as_float(deal.get("OPPORTUNITY")),
        "source": source,
        "comments": comments,
        "comments_count": 0,
        "activities_count": 0,
        "next_steps_count": 0,
        "calls_count": 0,
        "cached_transcripts_count": 0,
        "chat_activity_count": 0,
        "bitrix_chat_summary": "",
        "bitrix_call_summary": "",
        "context_excerpt": "",
        "context_enriched": False,
        "has_meaningful_comment": has_meaningful_comment(comments),
        "has_next_step": None,
        **reason,
    }
    score, issues = crm_quality_score(row)
    row["manager_work_score"] = score
    row["manager_work_issues"] = issues
    return row


def apply_context(row: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    out = {**row, **context}
    score, issues = crm_quality_score(out)
    out["manager_work_score"] = score
    out["manager_work_issues"] = issues
    return out


def build_summary(rows: list[dict[str, Any]], detailed_count: int) -> list[dict[str, Any]]:
    total = len(rows)
    final_not_won = sum(1 for row in rows if row.get("stage_id") in FINAL_NOT_WON_STAGES)
    open_not_won = total - final_not_won
    no_reason = sum(1 for row in rows if row.get("reason_category") == "Причина не указана / мало контекста")
    no_comment = sum(1 for row in rows if not row.get("has_meaningful_comment"))
    no_next_step_open = sum(
        1
        for row in rows
        if row.get("context_enriched")
        and row.get("loss_state") == "Не выиграна, но не закрыта"
        and row.get("has_next_step") is False
    )
    avg_score = round(sum(float(row.get("manager_work_score") or 0) for row in rows) / max(1, total), 2)
    return [
        {"metric": "Категория", "value": f"{SO_CATEGORY_ID} - {SO_CATEGORY_NAME}"},
        {"metric": "Сделок отдела продаж, не выиграны", "value": total},
        {"metric": "Закрытых/проваленных невыигранных", "value": final_not_won},
        {"metric": "Открытых невыигранных/зависших", "value": open_not_won},
        {"metric": "Сделок с детальным контекстом", "value": detailed_count},
        {"metric": "Средняя оценка работы менеджера с CRM", "value": avg_score},
        {"metric": "Без понятной причины", "value": no_reason},
        {"metric": "Без содержательного комментария", "value": no_comment},
        {"metric": "Открытых без следующего шага среди проверенных таймлайнов", "value": no_next_step_open},
        {"metric": "Дата отчета", "value": datetime.now().isoformat(timespec="seconds")},
    ]


def build_reason_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total = max(1, len(rows))
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("reason_category") or "Причина не указана / мало контекста")].append(row)

    out: list[dict[str, Any]] = []
    for reason, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        managers = Counter(str(row.get("manager_name") or "") for row in items)
        stages = Counter(str(row.get("stage_name") or "") for row in items)
        examples = "\n".join(
            f"{row.get('title')} ({row.get('deal_url')})"
            for row in items[:8]
        )
        recommendation = next(
            (rule.recommendation for rule in REASON_RULES if rule.name == reason),
            "Сделать причину отказа обязательной и разбирать сделки без причины отдельно.",
        )
        out.append(
            {
                "reason_category": reason,
                "deals_count": len(items),
                "share_percent": round(len(items) * 100.0 / total, 2),
                "top_managers": "; ".join(f"{name}: {count}" for name, count in managers.most_common(5) if name),
                "top_stages": "; ".join(f"{name}: {count}" for name, count in stages.most_common(5) if name),
                "avg_manager_work_score": round(
                    sum(float(row.get("manager_work_score") or 0) for row in items) / max(1, len(items)),
                    2,
                ),
                "recommendation": recommendation,
                "example_deals": examples,
            }
        )
    return out


def build_manager_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("manager_name") or "")].append(row)

    out: list[dict[str, Any]] = []
    for manager, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        reasons = Counter(str(row.get("reason_category") or "") for row in items)
        stages = Counter(str(row.get("stage_name") or "") for row in items)
        final_count = sum(1 for row in items if row.get("stage_id") in FINAL_NOT_WON_STAGES)
        no_comment = sum(1 for row in items if not row.get("has_meaningful_comment"))
        no_reason = sum(
            1 for row in items if row.get("reason_category") == "Причина не указана / мало контекста"
        )
        no_next = sum(
            1
            for row in items
            if row.get("context_enriched")
            and row.get("loss_state") == "Не выиграна, но не закрыта"
            and row.get("has_next_step") is False
        )
        avg_score = round(
            sum(float(row.get("manager_work_score") or 0) for row in items) / max(1, len(items)),
            2,
        )
        out.append(
            {
                "manager_name": manager,
                "deals_count": len(items),
                "final_not_won_count": final_count,
                "open_not_won_count": len(items) - final_count,
                "avg_manager_work_score": avg_score,
                "no_meaningful_comment_count": no_comment,
                "no_reason_count": no_reason,
                "open_without_next_step_count": no_next,
                "top_reasons": "; ".join(f"{name}: {count}" for name, count in reasons.most_common(5) if name),
                "top_stages": "; ".join(f"{name}: {count}" for name, count in stages.most_common(5) if name),
            }
        )
    return out


def build_stage_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("stage_name") or "")].append(row)
    total = max(1, len(rows))
    out: list[dict[str, Any]] = []
    for stage, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        reasons = Counter(str(row.get("reason_category") or "") for row in items)
        out.append(
            {
                "stage_name": stage,
                "deals_count": len(items),
                "share_percent": round(len(items) * 100.0 / total, 2),
                "avg_manager_work_score": round(
                    sum(float(row.get("manager_work_score") or 0) for row in items) / max(1, len(items)),
                    2,
                ),
                "top_reasons": "; ".join(f"{name}: {count}" for name, count in reasons.most_common(5) if name),
            }
        )
    return out


def build_recommendations(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reason_rows = build_reason_rows(rows)
    manager_rows = build_manager_rows(rows)
    recommendations: list[dict[str, Any]] = []
    if reason_rows:
        top = reason_rows[0]
        recommendations.append(
            {
                "priority": "Критично",
                "area": "Причины отказов",
                "finding": f"Главная категория: {top['reason_category']} ({top['deals_count']} сделок, {top['share_percent']}%).",
                "action": top["recommendation"],
            }
        )
    no_reason = sum(1 for row in rows if row.get("reason_category") == "Причина не указана / мало контекста")
    if no_reason:
        recommendations.append(
            {
                "priority": "Критично" if no_reason / max(1, len(rows)) > 0.25 else "Высокий",
                "area": "CRM-дисциплина",
                "finding": f"{no_reason} сделок без понятной причины отказа/нереализации.",
                "action": "Сделать обязательным комментарий причины: что хотел клиент, что предложили, почему не реализовано, следующий допустимый контакт.",
            }
        )
    for manager in manager_rows[:5]:
        if float(manager.get("avg_manager_work_score") or 0) < 70:
            recommendations.append(
                {
                    "priority": "Высокий",
                    "area": f"Менеджер: {manager['manager_name']}",
                    "finding": f"Средняя оценка CRM-работы {manager['avg_manager_work_score']}; без комментария: {manager['no_meaningful_comment_count']}; без причины: {manager['no_reason_count']}.",
                    "action": "Разобрать 10 последних сделок менеджера и закрепить стандарт закрытия сделки в СО.",
                }
            )
    return recommendations


def write_excel(
    path: Path,
    *,
    summary: list[dict[str, Any]],
    reasons: list[dict[str, Any]],
    stages: list[dict[str, Any]],
    managers: list[dict[str, Any]],
    deals: list[dict[str, Any]],
    recommendations: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name="Итоги", index=False)
        pd.DataFrame(reasons).to_excel(writer, sheet_name="Причины", index=False)
        pd.DataFrame(stages).to_excel(writer, sheet_name="Стадии", index=False)
        pd.DataFrame(managers).to_excel(writer, sheet_name="Менеджеры", index=False)
        pd.DataFrame(deals).to_excel(writer, sheet_name="Сделки", index=False)
        pd.DataFrame(recommendations).to_excel(writer, sheet_name="Рекомендации", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True, name="Arial")
        body_font = Font(name="Arial", size=10)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=idx).value or "").lower()
                width = 18
                if any(token in header for token in ("url", "comment", "summary", "excerpt", "evidence", "recommendation", "issues", "finding", "action", "example")):
                    width = 48
                elif any(token in header for token in ("title", "reason", "stage", "manager")):
                    width = 30
                ws.column_dimensions[get_column_letter(idx)].width = min(width, 65)


async def run(args: argparse.Namespace) -> tuple[Path, Path, list[dict[str, Any]]]:
    load_dotenv(args.env_path, override=True)
    raw_path = REPORTS_DIR / "so_deals_raw.json"
    api = Bitrix24API(readonly=True)
    try:
        sales_ids, manager_names = await load_sales_users(api)
        deals = await fetch_deals(api, sales_ids=sales_ids, raw_path=raw_path, resume=args.resume)
        state_cache = _load_state_cache()
        base_rows = [base_row(deal, manager_names, args.domain) for deal in deals]

        if args.detail_limit == 0:
            detail_targets = []
        else:
            if args.detail_final_only:
                detail_targets = [
                    deal
                    for deal in deals
                    if str(deal.get("STAGE_ID") or "") in FINAL_NOT_WON_STAGES
                ]
            else:
                detail_targets = [
                    deal
                    for deal in deals
                    if str(deal.get("STAGE_ID") or "") in FINAL_NOT_WON_STAGES
                    or not has_meaningful_comment(deal.get("COMMENTS"))
                ]
            if args.detail_limit and args.detail_limit > 0:
                detail_targets = detail_targets[: args.detail_limit]

        print(f"[DETAIL] targets={len(detail_targets)}", flush=True)
        semaphore = asyncio.Semaphore(max(1, args.concurrency))
        context_tasks = [
            enrich_context(
                api,
                deal,
                manager_names=manager_names,
                state_cache=state_cache,
                semaphore=semaphore,
            )
            for deal in detail_targets
        ]
        context_rows: list[dict[str, Any]] = []
        for index, task in enumerate(asyncio.as_completed(context_tasks), 1):
            try:
                context_rows.append(await task)
            except Exception as exc:
                print(f"[WARN] context failed: {exc}", flush=True)
            if index % 50 == 0 or index == len(context_tasks):
                print(f"[DETAIL] {index}/{len(context_tasks)}", flush=True)
        context_by_id = {row["deal_id"]: row for row in context_rows}
    finally:
        await api.aclose()

    rows = [
        apply_context(row, context_by_id[row["deal_id"]])
        if row["deal_id"] in context_by_id
        else row
        for row in base_rows
    ]
    rows.sort(key=lambda row: (str(row.get("manager_name") or ""), str(row.get("stage_name") or ""), str(row.get("date_modify") or "")), reverse=False)

    summary = build_summary(rows, detailed_count=len(context_by_id))
    reasons = build_reason_rows(rows)
    stages = build_stage_rows(rows)
    managers = build_manager_rows(rows)
    recommendations = build_recommendations(rows)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_out = REPORTS_DIR / f"so_deals_analysis_{ts}.json"
    xlsx_out = REPORTS_DIR / f"so_deals_analysis_{ts}.xlsx"
    payload = {
        "summary": summary,
        "reasons": reasons,
        "stages": stages,
        "managers": managers,
        "deals": rows,
        "recommendations": recommendations,
    }
    json_out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(
        xlsx_out,
        summary=summary,
        reasons=reasons,
        stages=stages,
        managers=managers,
        deals=rows,
        recommendations=recommendations,
    )
    latest_json = REPORTS_DIR / "latest_so_deals_analysis.json"
    latest_xlsx = REPORTS_DIR / "latest_so_deals_analysis.xlsx"
    shutil.copy2(json_out, latest_json)
    shutil.copy2(xlsx_out, latest_xlsx)
    return json_out, xlsx_out, rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Анализ сделок отдела продаж в категории СО")
    parser.add_argument("--env-path", default=".env")
    parser.add_argument("--domain", default=DEFAULT_DOMAIN)
    parser.add_argument("--concurrency", type=int, default=8)
    parser.add_argument("--detail-limit", type=int, default=800)
    parser.add_argument("--detail-final-only", action="store_true")
    parser.add_argument("--resume", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    json_out, xlsx_out, rows = asyncio.run(run(args))
    print(f"JSON: {json_out}")
    print(f"Excel: {xlsx_out}")
    print(f"Deals: {len(rows)}")
    print("Top reasons:")
    for row in build_reason_rows(rows)[:10]:
        print(f"- {row['reason_category']}: {row['deals_count']} ({row['share_percent']}%)")


if __name__ == "__main__":
    main()
