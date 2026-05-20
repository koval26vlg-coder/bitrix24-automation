from openpyxl import load_workbook

from pipelines.conversation_intelligence import build_conversation_intelligence
from pipelines.reporting import flatten_results


def _sample_row(**overrides):
    row = {
        "deal_id": "10",
        "deal_url": "https://example.bitrix24.ru/crm/deal/details/10/",
        "stage_name": "Потребность выявлена",
        "manager_name": "Иван Иванов",
        "activity_id": "100",
        "subject": "Исходящий звонок",
        "duration_minutes": 4.5,
        "transcript_text": (
            "Клиент: Это дорого, я не уверен, надо подумать. "
            "Менеджер: Наверное, попробуем потом вернуться."
        ),
        "has_needs_discovery": False,
        "has_next_step_phrase": False,
        "crm_checklist_percent": 55,
        "call_checklist_percent": 45,
        "overall_score": 52,
        "call_quality_score": 45,
    }
    row.update(overrides)
    return row


def test_conversation_intelligence_detects_unhandled_objection_and_risk():
    result = build_conversation_intelligence([_sample_row()])

    assert result["objection_rows"]
    assert result["objection_rows"][0]["objection_status"] == "не отработано"
    assert result["emotional_risk_rows"][0]["ci_emotional_risk_level"] in {"Высокий", "Критичный"}
    assert result["conversion_factor_rows"][0]["ci_factor_priority"] in {"Высокий", "Критичный"}
    assert result["manager_recommendation_rows"][0]["ci_main_growth_area"]


def test_conversation_intelligence_uses_lost_deal_summary_as_conversion_factor():
    result = build_conversation_intelligence(
        [_sample_row()],
        lost_reason_summary_rows=[
            {
                "loss_reason_category": "Цена/бюджет",
                "lost_deals_count": 3,
                "lost_deals_share": 30,
                "conversion_next_action": "Внедрить аргументацию ценности.",
            }
        ],
    )

    factors = {row["ci_conversion_factor"] for row in result["conversion_factor_rows"]}

    assert "Проигрыш: Цена/бюджет" in factors


def test_flatten_results_writes_conversation_intelligence_sheets(monkeypatch, tmp_path):
    monkeypatch.setattr("pipelines.reporting.REPORTS_DIR", tmp_path)

    xlsx = flatten_results(rows=[_sample_row()], manager_summary=[])
    workbook = load_workbook(xlsx, read_only=True)

    assert "Карта разговора" in workbook.sheetnames
    assert "Возражения" in workbook.sheetnames
    assert "Эмоциональные риски" in workbook.sheetnames
    assert "Факторы конверсии" in workbook.sheetnames
    assert "Рекомендации менеджерам" in workbook.sheetnames
