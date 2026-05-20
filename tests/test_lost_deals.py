from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from pipelines.lost_deals import (
    build_lost_deals_analysis,
    build_conversion_action_rows,
    build_lost_reason_summary,
    classify_loss_reason,
    lost_filter_from_base,
)
from pipelines.reporting import flatten_results


def test_lost_filter_replaces_selected_stage_with_failed_semantic():
    flt = {
        "CATEGORY_ID": 1,
        "STAGE_ID": ["C1:NEW", "C1:PREPARATION"],
        "ASSIGNED_BY_ID": "5",
        ">=DATE_CREATE": "2026-05-01",
    }

    out = lost_filter_from_base(flt)

    assert out == {
        "CATEGORY_ID": 1,
        "ASSIGNED_BY_ID": "5",
        ">=DATE_CREATE": "2026-05-01",
        "=STAGE_SEMANTIC_ID": "F",
    }


def test_lost_filter_infers_category_from_selected_stage():
    out = lost_filter_from_base({"STAGE_ID": "C7:NEW"})

    assert out["CATEGORY_ID"] == 7
    assert out["=STAGE_SEMANTIC_ID"] == "F"
    assert "STAGE_ID" not in out


def test_classify_loss_reason_price_and_tools():
    result = classify_loss_reason("Клиент сказал, что цена дорогая и бюджет не согласован")

    assert result["loss_reason_category"] == "Цена/бюджет"
    assert result["loss_reason_confidence"] == "Высокая"
    assert "ROI" in result["conversion_tools"]
    assert "цены" in result["conversion_next_action"]


def test_classify_loss_reason_default_when_no_reason_found():
    result = classify_loss_reason("Сделка закрыта без комментариев")

    assert result["loss_reason_category"] == "Причина не указана"
    assert result["loss_reason_confidence"] == "Низкая"
    assert "обязательное поле причины" in result["conversion_tools"]


def test_lost_reason_summary_and_conversion_priorities():
    rows = [
        {"loss_reason_category": "Цена/бюджет", "lost_amount": 100, "lost_lifetime_days": 2, "lost_manager_name": "А"},
        {"loss_reason_category": "Цена/бюджет", "lost_amount": 200, "lost_lifetime_days": 4, "lost_manager_name": "А"},
        {"loss_reason_category": "Нет связи с клиентом", "lost_amount": 50, "lost_manager_name": "Б"},
    ]

    summary = build_lost_reason_summary(rows)
    actions = build_conversion_action_rows(summary)

    assert summary[0]["loss_reason_category"] == "Цена/бюджет"
    assert summary[0]["lost_deals_count"] == 2
    assert summary[0]["lost_amount"] == 300
    assert summary[0]["lost_avg_lifetime_days"] == 3
    assert actions[0]["conversion_priority"] == "Критично"


@pytest.mark.asyncio
async def test_build_lost_deals_analysis_uses_filter_and_report_text(monkeypatch, tmp_path):
    filter_path = tmp_path / "filter.json"
    filter_path.write_text('{"CATEGORY_ID": 1, "STAGE_ID": "C1:NEW"}', encoding="utf-8")
    args = SimpleNamespace(
        lost_deals_analysis=True,
        mode="filter",
        filter_json=str(filter_path),
        lost_deals_limit=500,
        domain="example.bitrix24.ru",
    )

    class FakeApi:
        async def call(self, method, params):
            assert method == "crm.deal.list"
            assert params["filter"]["CATEGORY_ID"] == 1
            assert params["filter"]["=STAGE_SEMANTIC_ID"] == "F"
            assert "STAGE_ID" not in params["filter"]
            return {
                "result": [
                    {
                        "ID": "10",
                        "TITLE": "Сделка",
                        "STAGE_ID": "C1:LOSE",
                        "ASSIGNED_BY_ID": "5",
                        "OPPORTUNITY": "1000",
                    }
                ]
            }

    async def fake_user_name_map(api, ids):
        return {5: "Иван Иванов"}

    monkeypatch.setattr("pipelines.lost_deals.user_name_map", fake_user_name_map)

    result = await build_lost_deals_analysis(
        api=FakeApi(),
        args=args,
        results=[{"deal_id": "10", "transcripts_combined": "Клиент сказал, что дорого"}],
        stage_map={"C1:LOSE": "Сделка проиграна"},
    )

    assert result["rows"][0]["loss_reason_category"] == "Цена/бюджет"
    assert result["rows"][0]["lost_manager_name"] == "Иван Иванов"
    assert result["rows"][0]["lost_analysis_basis"] == "CRM + звонки отчета"
    assert result["summary_rows"][0]["lost_deals_count"] == 1
    assert result["action_rows"][0]["conversion_priority"] == "Критично"


def test_flatten_results_writes_lost_deal_sheets(monkeypatch, tmp_path):
    monkeypatch.setattr("pipelines.reporting.REPORTS_DIR", tmp_path)

    xlsx = flatten_results(
        rows=[],
        manager_summary=[],
        lost_deals_analysis={
            "rows": [
                {
                    "lost_deal_url": "https://example.bitrix24.ru/crm/deal/details/10/",
                    "lost_deal_title": "Сделка",
                    "loss_reason_category": "Цена/бюджет",
                }
            ],
            "summary_rows": [{"loss_reason_category": "Цена/бюджет", "lost_deals_count": 1}],
            "action_rows": [{"conversion_priority": "Критично", "loss_reason_category": "Цена/бюджет"}],
        },
    )

    workbook = load_workbook(xlsx, read_only=True)

    assert "Проигранные сделки" in workbook.sheetnames
    assert "Причины отказов" in workbook.sheetnames
    assert "Рост конверсии" in workbook.sheetnames
