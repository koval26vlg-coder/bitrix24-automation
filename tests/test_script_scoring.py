from openpyxl import load_workbook

from pipelines.reporting import flatten_results
from pipelines.script_scoring import (
    build_script_gap_rows,
    build_script_profile_rows,
    build_script_score_rows,
    select_script_profile_ids,
)


def _row(**overrides):
    row = {
        "deal_id": "20",
        "deal_url": "https://example.bitrix24.ru/crm/deal/details/20/",
        "stage_name": "Потребность выявлена",
        "manager_name": "Павел Клец",
        "activity_id": "200",
        "subject": "Исходящий звонок",
        "duration_minutes": 5.0,
        "transcript_text": (
            "Здравствуйте, меня зовут Павел, компания Онлайн-касса. "
            "Удобно сейчас говорить? Звоню по вашей заявке. "
            "Подскажите, какая у вас задача и когда планируете подключение? "
            "Правильно понимаю, что нужна онлайн-касса с обслуживанием? "
            "Для вашей задачи подойдет решение с поддержкой. "
            "Я отправлю коммерческое предложение сегодня и завтра перезвоню. "
            "Остались вопросы? Спасибо, до свидания."
        ),
        "call_quality_score": 80,
        "crm_checklist_percent": 80,
        "overall_score": 80,
    }
    row.update(overrides)
    return row


def test_script_scoring_detects_completed_steps():
    rows = build_script_score_rows([_row()])

    assert rows
    statuses = {row["script_status"] for row in rows}
    assert "Выполнено" in statuses
    assert any(row["script_step"] == "Поздоровался, представился" for row in rows)
    assert all(row["script_block"] for row in rows)
    assert {row["script_profile_name"] for row in rows} >= {"Продажи онлайн-касс", "Этикет общения"}


def test_script_profile_selection_prefers_objection_management():
    text = "Клиент говорит, что дорого и надо подумать."

    selected = select_script_profile_ids({}, text)

    assert selected[0] == "objection_management"
    assert "communication_etiquette" in selected


def test_script_scoring_builds_manager_gap_summary():
    bad_row = _row(
        transcript_text="Клиент: Дорого. Менеджер: молчит.",
        call_quality_score=20,
        crm_checklist_percent=30,
    )

    script_rows = build_script_score_rows([bad_row])
    gap_rows = build_script_gap_rows(script_rows)

    gap_steps = {row["script_step"] for row in gap_rows}

    assert any("Предлагает решения" in step for step in gap_steps)
    assert "Обозначил дальнейшие шаги" in gap_steps
    assert gap_rows[0]["script_priority"] in {"Критичный", "Высокий", "Средний"}


def test_script_profile_rows_show_match_and_critical_errors():
    bad_row = _row(transcript_text="Клиент: Дорого. Менеджер: Ну не знаю.")

    script_rows = build_script_score_rows([bad_row])
    profile_rows = build_script_profile_rows(script_rows)
    objection_profile = next(
        row for row in profile_rows if row["script_profile_id"] == "objection_management"
    )

    assert objection_profile["script_profile_status"] == "Не соответствует"
    assert objection_profile["script_critical_errors_count"] > 0
    assert "Критичный шаг" in objection_profile["script_critical_errors"]


def test_flatten_results_writes_script_scoring_sheets(monkeypatch, tmp_path):
    monkeypatch.setattr("pipelines.reporting.REPORTS_DIR", tmp_path)

    xlsx = flatten_results(rows=[_row()], manager_summary=[])
    workbook = load_workbook(xlsx, read_only=True)

    assert "Скрипты_итоги" in workbook.sheetnames
    assert "Скрипты_шаги" in workbook.sheetnames
    assert "Ошибки_менеджеров" in workbook.sheetnames
