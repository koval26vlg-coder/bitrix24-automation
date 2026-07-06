from __future__ import annotations

import argparse
import asyncio
import json
import os
import shutil
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bitrix24_api import Bitrix24API
from client_requests_analysis import (
    _clean,
    _dedupe_texts,
    _extract_text_from_activity,
    _short,
    classify_request,
    load_stage_names,
)
from pipelines.calls import (
    extract_bitrix_gpt_summaries,
    fetch_deal_activities,
    fetch_open_next_step_activities,
    fetch_timeline_comments,
)
from pipelines.deals import deal_url_from_id
from pipelines.paths import REPORTS_DIR
from pipelines.stages import safe_int

DEFAULT_DOMAIN = "online-kassa.bitrix24.ru"
LOCAL_TZ = timezone(timedelta(hours=3))
DEAL_ENTITY_TYPE_ID = 2
DEFAULT_CATEGORY_ID = "1"
INVOICE_STAGE_ID = "C1:UC_9NU15J"
INVOICE_STAGE_NAME = "Счет отправлен"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Топ крупных клиентов по сделкам, дошедшим до стадии `Счет отправлен`."
    )
    parser.add_argument("--start-date", default="2026-06-12")
    parser.add_argument("--end-date", default="2026-07-02")
    parser.add_argument("--category-id", default=DEFAULT_CATEGORY_ID)
    parser.add_argument("--stage-id", default=INVOICE_STAGE_ID)
    parser.add_argument("--top-n", type=int, default=10)
    parser.add_argument("--web-enrichment-json")
    parser.add_argument("--output-prefix")
    return parser.parse_args()


def day_start_iso(day: date) -> str:
    return datetime.combine(day, time.min, LOCAL_TZ).isoformat(timespec="seconds")


def day_after_start_iso(day: date) -> str:
    return datetime.combine(day + timedelta(days=1), time.min, LOCAL_TZ).isoformat(
        timespec="seconds"
    )


def parse_day(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def as_float(value: Any) -> float:
    try:
        return float(str(value or "0").replace(",", "."))
    except Exception:
        return 0.0


def parse_dt(value: Any) -> datetime | None:
    try:
        if not value:
            return None
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


def latest_transition(rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not rows:
        return None
    return sorted(rows, key=lambda row: str(row.get("CREATED_TIME") or ""))[-1]


def load_web_enrichment(path: str | None) -> dict[str, dict[str, Any]]:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        raise SystemExit(f"web enrichment file not found: {p}")
    data = json.loads(p.read_text(encoding="utf-8"))
    rows = data.get("clients", data) if isinstance(data, dict) else data
    out: dict[str, dict[str, Any]] = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        key = str(row.get("client_key") or row.get("client_name") or "").strip()
        if key:
            out[key] = row
    return out


async def fetch_invoice_transitions(
    api: Bitrix24API,
    *,
    category_id: str,
    stage_id: str,
    start_iso: str,
    end_iso: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    start: int | None = 0
    while start is not None:
        res = await api.call(
            "crm.stagehistory.list",
            {
                "entityTypeId": DEAL_ENTITY_TYPE_ID,
                "filter": {
                    "CATEGORY_ID": category_id,
                    "STAGE_ID": stage_id,
                    ">=CREATED_TIME": start_iso,
                    "<CREATED_TIME": end_iso,
                },
                "select": ["ID", "OWNER_ID", "STAGE_ID", "TYPE_ID", "CATEGORY_ID", "CREATED_TIME"],
                "order": {"ID": "ASC"},
                "start": start,
            },
        )
        result = res.get("result")
        chunk = result.get("items", []) if isinstance(result, dict) else result or []
        chunk = [row for row in chunk if isinstance(row, dict)]
        rows.extend(chunk)
        print(f"[stagehistory] {len(rows)} rows", flush=True)
        next_start = res.get("next")
        if next_start is None and isinstance(result, dict):
            next_start = result.get("next")
        start = safe_int(next_start) if next_start is not None else None
        if start is not None:
            await asyncio.sleep(0.15)
    return rows


async def fetch_deals_by_ids(api: Bitrix24API, deal_ids: list[str]) -> dict[str, dict[str, Any]]:
    select = [
        "ID",
        "TITLE",
        "ASSIGNED_BY_ID",
        "CATEGORY_ID",
        "STAGE_ID",
        "STAGE_SEMANTIC_ID",
        "DATE_CREATE",
        "DATE_MODIFY",
        "CLOSEDATE",
        "OPPORTUNITY",
        "CURRENCY_ID",
        "COMPANY_ID",
        "CONTACT_ID",
        "COMMENTS",
        "SOURCE_ID",
        "SOURCE_DESCRIPTION",
        "UF_CRM_1744698030",
    ]
    out: dict[str, dict[str, Any]] = {}
    clean = [deal_id for deal_id in dict.fromkeys(deal_ids) if str(deal_id).isdigit()]
    for index in range(0, len(clean), 50):
        batch = clean[index : index + 50]
        res = await api.call(
            "crm.deal.list",
            {"filter": {"ID": batch}, "select": select, "order": {"ID": "ASC"}},
        )
        for row in res.get("result", []) or []:
            deal_id = str(row.get("ID") or "")
            if deal_id:
                out[deal_id] = row
        print(f"[deals] {len(out)}/{len(clean)}", flush=True)
        await asyncio.sleep(0.15)
    return out


async def fetch_user_names(api: Bitrix24API, user_ids: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for uid in [x for x in dict.fromkeys(user_ids) if str(x).isdigit()]:
        try:
            res = await api.call("user.get", {"ID": int(uid)})
            rows = res.get("result") or []
            if rows and isinstance(rows[0], dict):
                user = rows[0]
                out[uid] = (
                    f"{user.get('NAME', '')} {user.get('LAST_NAME', '')}".strip() or uid
                )
        except Exception:
            out[uid] = uid
    return out


async def fetch_companies(api: Bitrix24API, company_ids: list[str]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for cid in [x for x in dict.fromkeys(company_ids) if str(x).isdigit()]:
        try:
            res = await api.call("crm.company.get", {"id": int(cid)})
            row = res.get("result") or {}
            if isinstance(row, dict):
                out[cid] = row
        except Exception as exc:
            out[cid] = {"ID": cid, "TITLE": cid, "_error": str(exc)}
        await asyncio.sleep(0.05)
    return out


async def fetch_contacts(api: Bitrix24API, contact_ids: list[str]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for cid in [x for x in dict.fromkeys(contact_ids) if str(x).isdigit()]:
        try:
            res = await api.call("crm.contact.get", {"id": int(cid)})
            row = res.get("result") or {}
            if isinstance(row, dict):
                out[cid] = row
        except Exception as exc:
            out[cid] = {"ID": cid, "_error": str(exc)}
        await asyncio.sleep(0.05)
    return out


def contact_name(row: dict[str, Any]) -> str:
    parts = [row.get("LAST_NAME"), row.get("NAME"), row.get("SECOND_NAME")]
    return " ".join(_clean(x) for x in parts if _clean(x))


def client_key_for_deal(
    deal: dict[str, Any],
    companies: dict[str, dict[str, Any]],
    contacts: dict[str, dict[str, Any]],
) -> tuple[str, str, str]:
    company_id = str(deal.get("COMPANY_ID") or "").strip()
    if company_id and company_id != "0":
        company = companies.get(company_id, {})
        name = _clean(company.get("TITLE")) or f"Компания {company_id}"
        return f"company:{company_id}", name, "company"
    contact_id = str(deal.get("CONTACT_ID") or "").strip()
    if contact_id and contact_id != "0":
        contact = contacts.get(contact_id, {})
        name = contact_name(contact) or f"Контакт {contact_id}"
        return f"contact:{contact_id}", name, "contact"
    title = _clean(deal.get("TITLE")) or f"Сделка {deal.get('ID')}"
    return f"deal:{deal.get('ID')}", title, "deal"


async def enrich_deal_context(api: Bitrix24API, deal: dict[str, Any]) -> dict[str, Any]:
    deal_id = str(deal.get("ID") or "")
    comments, activities, next_steps = await asyncio.gather(
        fetch_timeline_comments(api, deal_id),
        fetch_deal_activities(api, deal_id, limit=100),
        fetch_open_next_step_activities(api, deal_id),
    )
    activity_texts: list[str] = []
    for activity in activities:
        activity_texts.extend(_extract_text_from_activity(activity))

    summary = extract_bitrix_gpt_summaries(comments, activities)
    texts = _dedupe_texts(
        [
            _clean(deal.get("TITLE")),
            _clean(deal.get("COMMENTS")),
            _clean(deal.get("SOURCE_DESCRIPTION")),
            *comments,
            *activity_texts,
        ],
        max_items=80,
    )
    classified = classify_request(texts)
    return {
        "context_texts": texts,
        "context_excerpt": _short(" | ".join(texts[:10]), 1200),
        "comments_count": len(comments),
        "activities_count": len(activities),
        "next_steps_count": len(next_steps),
        "bitrix_chat_summary": summary.get("chat_summary", ""),
        "bitrix_call_summary": summary.get("call_summary", ""),
        "bitrix_overall_meaning": summary.get("overall_meaning", ""),
        **classified,
    }


def company_snapshot(company: dict[str, Any]) -> str:
    parts = []
    for key in ("TITLE", "INDUSTRY", "COMPANY_TYPE", "COMMENTS", "ADDRESS", "WEB"):
        value = company.get(key)
        if isinstance(value, list):
            value = "; ".join(_clean(x.get("VALUE") if isinstance(x, dict) else x) for x in value)
        text = _clean(value)
        if text:
            parts.append(f"{key}: {text}")
    return _short(" | ".join(parts), 1000)


def apply_web(row: dict[str, Any], web: dict[str, dict[str, Any]]) -> dict[str, Any]:
    hit = web.get(row["client_key"]) or web.get(row["client_name"]) or {}
    row.update(
        {
            "external_activity_summary": _clean(hit.get("activity_summary")),
            "external_sources": "; ".join(hit.get("sources") or []),
            "external_confidence": _clean(hit.get("confidence")),
            "external_notes": _clean(hit.get("notes")),
        }
    )
    return row


def write_report(payload: dict[str, Any], xlsx_path: Path) -> None:
    summary_df = pd.DataFrame(payload["client_rows"])
    deals_df = pd.DataFrame(payload["deal_rows"])
    context_df = pd.DataFrame(payload["context_rows"])
    sources_df = pd.DataFrame(payload["source_rows"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pd.DataFrame([payload["summary"]]).to_excel(writer, sheet_name="Итоги", index=False)
        summary_df.to_excel(writer, sheet_name="Топ клиентов", index=False)
        deals_df.to_excel(writer, sheet_name="Сделки", index=False)
        context_df.to_excel(writer, sheet_name="Контекст Bitrix", index=False)
        sources_df.to_excel(writer, sheet_name="Внешние источники", index=False)

    wb = load_workbook(xlsx_path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    body_font = Font(name="Arial", size=10)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, start=1):
            width = min(
                60,
                max(12, max(len(str(cell.value or "")) for cell in col[:200]) + 2),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(xlsx_path)


async def build_report(args: argparse.Namespace) -> dict[str, Any]:
    start_day = parse_day(args.start_date)
    end_day = parse_day(args.end_date)
    start_iso = day_start_iso(start_day)
    end_iso = day_after_start_iso(end_day)
    web = load_web_enrichment(args.web_enrichment_json)

    async with Bitrix24API(readonly=True) as api:
        category_id, stage_names = args.category_id, await load_stage_names(api, args.category_id)
        transitions = await fetch_invoice_transitions(
            api,
            category_id=category_id,
            stage_id=args.stage_id,
            start_iso=start_iso,
            end_iso=end_iso,
        )
        by_deal: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in transitions:
            deal_id = str(row.get("OWNER_ID") or "").strip()
            if deal_id:
                by_deal[deal_id].append(row)
        deal_ids = list(by_deal)
        deals = await fetch_deals_by_ids(api, deal_ids)
        user_names = await fetch_user_names(
            api, [str(row.get("ASSIGNED_BY_ID") or "") for row in deals.values()]
        )
        companies = await fetch_companies(
            api, [str(row.get("COMPANY_ID") or "") for row in deals.values()]
        )
        contacts = await fetch_contacts(
            api, [str(row.get("CONTACT_ID") or "") for row in deals.values() if not row.get("COMPANY_ID")]
        )

        deal_rows: list[dict[str, Any]] = []
        client_groups: dict[str, dict[str, Any]] = {}
        for deal_id, deal in deals.items():
            transition = latest_transition(by_deal.get(deal_id, [])) or {}
            client_key, client_name, client_type = client_key_for_deal(deal, companies, contacts)
            amount = as_float(deal.get("OPPORTUNITY"))
            row = {
                "deal_id": deal_id,
                "deal_url": deal_url_from_id(DEFAULT_DOMAIN, deal_id),
                "client_key": client_key,
                "client_name": client_name,
                "client_type": client_type,
                "deal_title": _clean(deal.get("TITLE")),
                "amount": amount,
                "currency": _clean(deal.get("CURRENCY_ID")) or "RUB",
                "invoice_transition_at": transition.get("CREATED_TIME"),
                "date_create": deal.get("DATE_CREATE"),
                "date_modify": deal.get("DATE_MODIFY"),
                "stage_id": deal.get("STAGE_ID"),
                "stage_name": stage_names.get(str(deal.get("STAGE_ID") or ""), deal.get("STAGE_ID")),
                "manager_id": deal.get("ASSIGNED_BY_ID"),
                "manager_name": user_names.get(str(deal.get("ASSIGNED_BY_ID") or ""), ""),
                "source": _clean(deal.get("SOURCE_DESCRIPTION")) or _clean(deal.get("SOURCE_ID")),
                "company_id": deal.get("COMPANY_ID"),
                "contact_id": deal.get("CONTACT_ID"),
                "comments": _short(deal.get("COMMENTS"), 900),
            }
            deal_rows.append(row)
            group = client_groups.setdefault(
                client_key,
                {
                    "client_key": client_key,
                    "client_name": client_name,
                    "client_type": client_type,
                    "deals": [],
                    "total_amount": 0.0,
                    "company": companies.get(str(deal.get("COMPANY_ID") or ""), {}),
                },
            )
            group["deals"].append(row)
            group["total_amount"] += amount

        ranked = sorted(
            client_groups.values(),
            key=lambda item: (float(item["total_amount"]), len(item["deals"])),
            reverse=True,
        )[: max(1, args.top_n)]

        context_rows: list[dict[str, Any]] = []
        for group in ranked:
            texts: list[str] = []
            deal_contexts: list[dict[str, Any]] = []
            for deal_row in group["deals"]:
                context = await enrich_deal_context(api, deals[deal_row["deal_id"]])
                deal_contexts.append({"deal_id": deal_row["deal_id"], **context})
                texts.extend(context.pop("context_texts", []))
                context_rows.append(
                    {
                        "client_key": group["client_key"],
                        "client_name": group["client_name"],
                        "deal_id": deal_row["deal_id"],
                        "deal_url": deal_row["deal_url"],
                        "primary_request": context.get("primary_request"),
                        "secondary_requests": context.get("secondary_requests"),
                        "request_evidence": context.get("request_evidence"),
                        "bitrix_overall_meaning": context.get("bitrix_overall_meaning"),
                        "bitrix_chat_summary": context.get("bitrix_chat_summary"),
                        "bitrix_call_summary": context.get("bitrix_call_summary"),
                        "context_excerpt": context.get("context_excerpt"),
                        "comments_count": context.get("comments_count"),
                        "activities_count": context.get("activities_count"),
                        "next_steps_count": context.get("next_steps_count"),
                    }
                )
            classified = classify_request(_dedupe_texts(texts, max_items=120))
            group["primary_request"] = classified.get("primary_request", "")
            group["request_evidence"] = classified.get("request_evidence", "")
            group["request_recommendation"] = classified.get("request_recommendation", "")
            group["bitrix_context_excerpt"] = _short(" | ".join(_dedupe_texts(texts)[:12]), 1300)

        client_rows: list[dict[str, Any]] = []
        source_rows: list[dict[str, Any]] = []
        for rank, group in enumerate(ranked, start=1):
            deal_list = sorted(group["deals"], key=lambda item: float(item["amount"]), reverse=True)
            top_deal = deal_list[0] if deal_list else {}
            row = {
                "rank": rank,
                "client_key": group["client_key"],
                "client_name": group["client_name"],
                "client_type": group["client_type"],
                "total_amount": round(float(group["total_amount"]), 2),
                "deals_count": len(group["deals"]),
                "top_deal_amount": top_deal.get("amount", 0),
                "top_deal_title": top_deal.get("deal_title", ""),
                "top_deal_url": top_deal.get("deal_url", ""),
                "managers": "; ".join(sorted({row.get("manager_name", "") for row in group["deals"] if row.get("manager_name")})),
                "stage_names": "; ".join(sorted({row.get("stage_name", "") for row in group["deals"] if row.get("stage_name")})),
                "invoice_dates": "; ".join(sorted({str(row.get("invoice_transition_at") or "")[:10] for row in group["deals"] if row.get("invoice_transition_at")})),
                "bitrix_company_snapshot": company_snapshot(group.get("company") or {}),
                "bitrix_primary_request": group.get("primary_request", ""),
                "bitrix_request_evidence": group.get("request_evidence", ""),
                "bitrix_context_excerpt": group.get("bitrix_context_excerpt", ""),
                "recommendation": group.get("request_recommendation", ""),
            }
            row = apply_web(row, web)
            client_rows.append(row)
            for src in str(row.get("external_sources") or "").split(";"):
                src = src.strip()
                if src:
                    source_rows.append(
                        {
                            "client_key": row["client_key"],
                            "client_name": row["client_name"],
                            "source_url": src,
                        }
                    )

        summary = {
            "period_start": str(start_day),
            "period_end": str(end_day),
            "period_basis": f"stage-history переход в `{args.stage_id}` / `{INVOICE_STAGE_NAME}`",
            "category_id": category_id,
            "stage_id": args.stage_id,
            "stage_name": INVOICE_STAGE_NAME,
            "stage_transition_rows": len(transitions),
            "unique_deals": len(deals),
            "clients_total": len(client_groups),
            "top_n": len(client_rows),
            "top_total_amount": round(sum(float(row["total_amount"]) for row in client_rows), 2),
            "generated_at": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
        }
        return {
            "summary": summary,
            "client_rows": client_rows,
            "deal_rows": sorted(deal_rows, key=lambda row: float(row["amount"]), reverse=True),
            "context_rows": context_rows,
            "source_rows": source_rows,
        }


def main() -> None:
    args = parse_args()
    os.environ.setdefault("BITRIX24_SOURCE_IP", "192.168.1.103")
    REPORTS_DIR.mkdir(exist_ok=True)
    stamp = datetime.now(LOCAL_TZ).strftime("%Y%m%d_%H%M%S")
    prefix = args.output_prefix or f"top_large_invoices_{stamp}"
    json_path = REPORTS_DIR / f"{prefix}.json"
    xlsx_path = REPORTS_DIR / f"{prefix}.xlsx"

    payload = asyncio.run(build_report(args))
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(payload, xlsx_path)

    latest_json = REPORTS_DIR / "latest_top_large_invoices_analysis.json"
    latest_xlsx = REPORTS_DIR / "latest_top_large_invoices_analysis.xlsx"
    try:
        shutil.copy2(json_path, latest_json)
        shutil.copy2(xlsx_path, latest_xlsx)
    except PermissionError as exc:
        print(f"[WARN] latest copy skipped: {exc}", flush=True)

    print(json.dumps({"json": str(json_path), "xlsx": str(xlsx_path), **payload["summary"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
